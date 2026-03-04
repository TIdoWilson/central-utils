from __future__ import annotations

import base64
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    from PyPDF2 import PdfReader

try:
    import pdfplumber
except Exception:  # pragma: no cover
    pdfplumber = None


TOLERANCIA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = "[$R$-416] #,##0.00"
MIN_NOTA_SAIDA_RAZAO = 10000
MAX_DIGITOS_NOTA_ENTRADA_RAZAO = 7

RE_DATA_VALOR_SALDO = re.compile(
    r"^(\d{1,2}/\d{2}/\d{4})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+-\s+(.+)$",
    re.IGNORECASE,
)
RE_NF = re.compile(r"\b(?:NF|NFS-E)\s*:?\s*(\d{3,})\b", re.IGNORECASE)
RE_LINHA_RELATORIO = re.compile(r"^\s*(\d{3,})\s+\S+\s+\d{2}/\d{2}/\d{4}\S*\s+")
RE_MONEY_ANY = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2,4}")


@dataclass(frozen=True)
class ArquivoEntrada:
    nome: str
    bytes_pdf: bytes
    linhas: list[str]


@dataclass(frozen=True)
class RegistroRazao:
    nota: str
    data_lcto: str
    valor: Decimal
    historico: str
    arquivo: str


@dataclass(frozen=True)
class RegistroRelatorio:
    nota: str
    data_lcto: str
    valor_pis: Decimal
    valor_cofins: Decimal
    arquivo: str


def normalizar_ascii(s: str) -> str:
    return normalize("NFKD", s).encode("ASCII", "ignore").decode().upper()


def valor_br_para_decimal(s: str) -> Decimal:
    return Decimal(s.replace(".", "").replace(",", "."))


def normalizar_nota(s: str) -> str:
    return re.sub(r"\D", "", s).lstrip("0") or "0"


def _linhas_do_texto(txt: str) -> list[str]:
    return [ln.strip() for ln in txt.splitlines() if ln.strip()]


def _score_linhas_extraidas(linhas: list[str]) -> tuple[int, int, int]:
    score = 0
    for ln in linhas:
        ln_norm = normalizar_ascii(ln)
        if "RELATORIO" in ln_norm or "RAZAO" in ln_norm:
            score += 2
        if "PIS" in ln_norm or "COFINS" in ln_norm:
            score += 2
        if "RECOLHER" in ln_norm or "RECUPERAR" in ln_norm:
            score += 3
        if RE_DATA_VALOR_SALDO.match(ln):
            score += 5
        if RE_LINHA_RELATORIO.match(ln):
            score += 5
    return score, len(linhas), sum(len(ln) for ln in linhas)


def _extrair_linhas_pypdf(pdf_bytes: bytes) -> list[str]:
    txt = "\n".join((pg.extract_text() or "") for pg in PdfReader(io.BytesIO(pdf_bytes)).pages)
    return _linhas_do_texto(txt)


def _extrair_linhas_pdfplumber(pdf_bytes: bytes) -> list[str]:
    if pdfplumber is None:
        return []
    partes: list[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            partes.append(page.extract_text() or "")
    return _linhas_do_texto("\n".join(partes))


def ler_linhas_pdf_bytes(pdf_bytes: bytes) -> list[str]:
    candidatos: list[list[str]] = []
    for extrator in (_extrair_linhas_pypdf, _extrair_linhas_pdfplumber):
        try:
            linhas = extrator(pdf_bytes)
        except Exception:
            linhas = []
        if linhas:
            candidatos.append(linhas)

    if not candidatos:
        return []

    return max(candidatos, key=_score_linhas_extraidas)


def detectar_tipo_arquivo_por_conteudo(linhas: list[str]) -> tuple[str | None, str | None]:
    cab = normalizar_ascii("\n".join(linhas[:220]))
    cab_compacto = re.sub(r"\s+", " ", cab)

    if "RELATORIO DE APURACAO DE PIS E COFINS" in cab_compacto:
        if " - ENTRADAS - " in cab_compacto:
            return "entrada_relatorio", "AMBOS"
        if " - SAIDA - " in cab_compacto or " - SAIDAS - " in cab_compacto:
            return "saida_relatorio", "AMBOS"
        return None, None

    if "RELATORIO: RAZAO CONTABIL" not in cab_compacto and "RELATORIO:RAZAO CONTABIL" not in cab_compacto:
        return None, None

    if "PIS A RECUPERAR" in cab_compacto:
        return "entrada_razao", "PIS"
    if "PIS A RECOLHER" in cab_compacto:
        return "saida_razao", "PIS"
    if "COFINS A RECUPERAR" in cab_compacto:
        return "entrada_razao", "COFINS"
    if "COFINS A RECOLHER" in cab_compacto:
        return "saida_razao", "COFINS"

    return None, None


def parse_razao(arquivo: ArquivoEntrada) -> list[RegistroRazao]:
    out: list[RegistroRazao] = []
    for ln in arquivo.linhas:
        m = RE_DATA_VALOR_SALDO.match(ln)
        if not m:
            continue
        data_lcto, valor_txt, _, historico = m.groups()
        nf = RE_NF.search(historico)
        if not nf:
            continue
        nota = normalizar_nota(nf.group(1))
        valor = valor_br_para_decimal(valor_txt)
        if valor == 0:
            continue
        out.append(
            RegistroRazao(
                nota=nota,
                data_lcto=data_lcto,
                valor=valor,
                historico=historico[:220],
                arquivo=arquivo.nome,
            )
        )
    return out


def filtrar_razao_por_layout(registros: list[RegistroRazao], movimento: str) -> tuple[list[RegistroRazao], int]:
    filtrados: list[RegistroRazao] = []
    descartados = 0
    for r in registros:
        if not r.nota.isdigit():
            descartados += 1
            continue
        if movimento == "saida" and int(r.nota) < MIN_NOTA_SAIDA_RAZAO:
            descartados += 1
            continue
        if movimento == "entrada" and len(r.nota) > MAX_DIGITOS_NOTA_ENTRADA_RAZAO:
            descartados += 1
            continue
        filtrados.append(r)
    return filtrados, descartados


def parse_relatorio(arquivo: ArquivoEntrada) -> list[RegistroRelatorio]:
    out: list[RegistroRelatorio] = []
    for ln in arquivo.linhas:
        if not RE_LINHA_RELATORIO.match(ln):
            continue
        partes = ln.split()
        if len(partes) < 4:
            continue
        nota = normalizar_nota(partes[0])
        m_data = re.search(r"\d{2}/\d{2}/\d{4}", ln)
        data_lcto = m_data.group(0) if m_data else ""
        valores = RE_MONEY_ANY.findall(ln)
        if len(valores) < 9:
            continue
        try:
            valor_pis = valor_br_para_decimal(valores[5])
            valor_cofins = valor_br_para_decimal(valores[8])
        except Exception:
            continue
        out.append(
            RegistroRelatorio(
                nota=nota,
                data_lcto=data_lcto,
                valor_pis=valor_pis,
                valor_cofins=valor_cofins,
                arquivo=arquivo.nome,
            )
        )
    return out


def somar_por_nota_razao(registros: list[RegistroRazao]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        out[r.nota] += r.valor
    return dict(out)


def somar_por_nota_relatorio(registros: list[RegistroRelatorio], tributo: str) -> dict[str, Decimal]:
    out: dict[str, Decimal] = defaultdict(Decimal)
    for r in registros:
        if tributo == "PIS":
            out[r.nota] += r.valor_pis
        else:
            out[r.nota] += r.valor_cofins
    return dict(out)


def filtrar_exemplo(lista: list[object], limite: int = 3) -> str:
    if not lista:
        return ""
    return " | ".join(str(x) for x in lista[:limite])


def montar_linhas_inconsistencia(
    mapa_razao: dict[str, Decimal],
    mapa_relatorio: dict[str, Decimal],
    exemplos_razao: dict[str, list[str]],
    exemplos_relatorio: dict[str, list[str]],
) -> list[list[object]]:
    out: list[list[object]] = []
    notas = sorted(set(mapa_razao) | set(mapa_relatorio))
    for nota in notas:
        vr = mapa_razao.get(nota, Decimal("0"))
        vl = mapa_relatorio.get(nota, Decimal("0"))
        dif = vr - vl

        if nota not in mapa_relatorio:
            tipo = "SO_NO_RAZAO"
        elif nota not in mapa_razao:
            tipo = "SO_NO_RELATORIO"
        elif abs(dif) > TOLERANCIA:
            tipo = "VALOR_DIVERGENTE"
        else:
            continue

        out.append(
            [
                tipo,
                nota,
                float(vr),
                float(vl),
                float(dif),
                filtrar_exemplo(exemplos_razao.get(nota, [])),
                filtrar_exemplo(exemplos_relatorio.get(nota, [])),
            ]
        )
    return out


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            txt = "" if val is None else str(val)
            if len(txt) > max_len:
                max_len = len(txt)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 70))


def escrever_aba_inconsistencias(ws, linhas: list[list[object]]) -> None:
    ws.append(
        [
            "TIPO",
            "NOTA",
            "VALOR_RAZAO",
            "VALOR_RELATORIO",
            "DIFERENCA",
            "EXEMPLOS_RAZAO",
            "EXEMPLOS_RELATORIO",
        ]
    )
    for row in linhas:
        ws.append(row)

    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for c in row:
            if isinstance(c.value, (int, float)):
                c.number_format = XLSX_FORMATO_MOEDA_BR
    autoajustar_colunas(ws)


def gerar_xlsx_bytes(sheets: dict[str, list[list[object]]]) -> bytes:
    wb = Workbook()
    primeira = True
    for nome, linhas in sheets.items():
        if primeira:
            ws = wb.active
            ws.title = nome
            primeira = False
        else:
            ws = wb.create_sheet(nome)
        escrever_aba_inconsistencias(ws, linhas)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def obter_arquivo_unico(lista: list[ArquivoEntrada], descricao: str) -> ArquivoEntrada:
    if not lista:
        raise ValueError(f"Arquivo obrigatorio nao encontrado para: {descricao}")
    if len(lista) > 1:
        nomes = ", ".join(p.nome for p in lista)
        raise ValueError(f"Encontrado mais de um arquivo para {descricao}: {nomes}")
    return lista[0]


def _modo_para_movimentos(modo: str) -> list[tuple[str, str]]:
    m = (modo or "AUTO").upper()
    movimentos: list[tuple[str, str]] = []
    if m in {"AUTO", "RECUPERAR"}:
        movimentos.append(("entrada", "RECUPERAR"))
    if m in {"AUTO", "RECOLHER"}:
        movimentos.append(("saida", "RECOLHER"))
    if not movimentos:
        raise ValueError("Modo invalido. Use AUTO, RECUPERAR ou RECOLHER.")
    return movimentos


def conciliar_pis_cofins(arquivos: list[tuple[str, bytes]], modo: str = "AUTO") -> dict[str, Any]:
    if len(arquivos) < 3:
        raise ValueError("Envie no minimo 3 PDFs (PIS Razao, COFINS Razao e Relatorio).")

    entrada: list[ArquivoEntrada] = []
    for nome, conteudo in arquivos:
        if not conteudo:
            continue
        linhas = ler_linhas_pdf_bytes(conteudo)
        entrada.append(ArquivoEntrada(nome=nome, bytes_pdf=conteudo, linhas=linhas))

    buckets: dict[str, Any] = {
        "entrada_razao": {"PIS": [], "COFINS": []},
        "saida_razao": {"PIS": [], "COFINS": []},
        "entrada_relatorio": [],
        "saida_relatorio": [],
    }

    for arq in entrada:
        tipo, tributo = detectar_tipo_arquivo_por_conteudo(arq.linhas)
        if tipo is None:
            continue
        if tipo in {"entrada_razao", "saida_razao"} and tributo in {"PIS", "COFINS"}:
            buckets[tipo][tributo].append(arq)
        elif tipo in {"entrada_relatorio", "saida_relatorio"}:
            buckets[tipo].append(arq)

    movimentos = _modo_para_movimentos(modo)
    sheets: dict[str, list[list[object]]] = {}
    resumo: list[dict[str, Any]] = []
    preview_incons: list[dict[str, Any]] = []

    for mov, rotulo in movimentos:
        chave_razao = "entrada_razao" if mov == "entrada" else "saida_razao"
        chave_rel = "entrada_relatorio" if mov == "entrada" else "saida_relatorio"

        if not buckets[chave_razao]["PIS"] or not buckets[chave_razao]["COFINS"] or not buckets[chave_rel]:
            continue

        razao_pis_pdf = obter_arquivo_unico(buckets[chave_razao]["PIS"], f"{rotulo} Razao PIS")
        razao_cof_pdf = obter_arquivo_unico(buckets[chave_razao]["COFINS"], f"{rotulo} Razao COFINS")
        rel_pdf = obter_arquivo_unico(buckets[chave_rel], f"{rotulo} Relatorio")

        regs_razao_pis = parse_razao(razao_pis_pdf)
        regs_razao_cof = parse_razao(razao_cof_pdf)
        regs_relatorio = parse_relatorio(rel_pdf)

        regs_razao_pis, desc_pis = filtrar_razao_por_layout(regs_razao_pis, mov)
        regs_razao_cof, desc_cof = filtrar_razao_por_layout(regs_razao_cof, mov)

        if not regs_razao_pis or not regs_razao_cof or not regs_relatorio:
            raise ValueError(f"Sem registros validos para {rotulo}.")

        for tributo, regs_razao, descartados in (
            ("PIS", regs_razao_pis, desc_pis),
            ("COFINS", regs_razao_cof, desc_cof),
        ):
            mapa_razao = somar_por_nota_razao(regs_razao)
            mapa_rel = somar_por_nota_relatorio(regs_relatorio, tributo)

            ex_razao: dict[str, list[str]] = defaultdict(list)
            ex_rel: dict[str, list[str]] = defaultdict(list)
            for r in regs_razao:
                ex_razao[r.nota].append(f"{r.data_lcto} {r.valor:.2f}")
            for r in regs_relatorio:
                v = r.valor_pis if tributo == "PIS" else r.valor_cofins
                ex_rel[r.nota].append(f"{r.data_lcto} {v:.2f}")

            linhas = montar_linhas_inconsistencia(mapa_razao, mapa_rel, dict(ex_razao), dict(ex_rel))
            nome_aba = f"{rotulo}_{tributo}"
            sheets[nome_aba] = linhas

            resumo.append(
                {
                    "movimento": rotulo,
                    "tributo": tributo,
                    "arquivo_razao": regs_razao[0].arquivo if regs_razao else "",
                    "arquivo_relatorio": rel_pdf.nome,
                    "registros_razao": len(regs_razao),
                    "registros_relatorio": len(regs_relatorio),
                    "descartados_razao": descartados,
                    "inconsistencias": len(linhas),
                }
            )
            for ln in linhas[:200]:
                preview_incons.append(
                    {
                        "movimento": rotulo,
                        "tributo": tributo,
                        "tipo": ln[0],
                        "nota": ln[1],
                        "valor_razao": ln[2],
                        "valor_relatorio": ln[3],
                        "diferenca": ln[4],
                    }
                )

    if not sheets:
        raise ValueError("Nenhum trio completo identificado para processamento.")

    xlsx_bytes = gerar_xlsx_bytes(sheets)
    filename = f"Conciliacao_PIS_COFINS_{datetime.now().strftime('%H-%M-%S_%d-%m-%Y')}.xlsx"
    return {
        "ok": True,
        "filename": filename,
        "xlsxBase64": base64.b64encode(xlsx_bytes).decode("ascii"),
        "resumo": resumo,
        "inconsistencias": preview_incons,
    }
