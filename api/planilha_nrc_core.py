from __future__ import annotations

import io
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


def _normalize_text(value: Any) -> str:
    text = str(value or '').strip().lower()
    text = ''.join(
        ch for ch in unicodedata.normalize('NFD', text)
        if unicodedata.category(ch) != 'Mn'
    )
    return ' '.join(text.split())


def _parse_period(periodo: str) -> tuple[int, int, int]:
    raw = str(periodo or '').strip()
    parts = raw.split('/')
    if len(parts) != 2:
        raise ValueError('Periodo invalido. Use o formato MM/AAAA.')
    try:
        month = int(parts[0])
        year = int(parts[1])
    except ValueError as exc:
        raise ValueError('Periodo invalido. Use o formato MM/AAAA.') from exc
    if month < 1 or month > 12:
        raise ValueError('Periodo invalido. Mes deve estar entre 01 e 12.')
    if year < 1900 or year > 2999:
        raise ValueError('Periodo invalido. Ano fora do intervalo suportado.')
    return month, year, year * 12 + month


def _parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None

    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _build_mappings(raw_mappings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for idx, row in enumerate(raw_mappings or []):
        historico = str(row.get('historico') or '').strip()
        agrupamento = str(row.get('agrupamento') or '').strip()
        if not historico or not agrupamento:
            continue
        out.append(
            {
                'index': idx,
                'historico': historico,
                'agrupamento': agrupamento,
                'historico_norm': _normalize_text(historico),
            }
        )
    return out


def _header_columns(sheet) -> tuple[int, int, int] | None:
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    wanted = {
        'data': None,
        'historico': None,
        'agrupamento': None,
    }

    for idx, header in enumerate(headers, start=1):
        name = _normalize_text(header)
        if name == 'data':
            wanted['data'] = idx
        elif name == 'historico':
            wanted['historico'] = idx
        elif name == 'agrupamento':
            wanted['agrupamento'] = idx

    if wanted['data'] and wanted['historico'] and wanted['agrupamento']:
        return wanted['data'], wanted['historico'], wanted['agrupamento']
    return None


def processar_planilha_nrc(
    arquivo_bytes: bytes,
    periodo_inicial: str,
    periodo_final: str,
    mappings: list[dict[str, Any]],
) -> dict[str, Any]:
    if not arquivo_bytes:
        raise ValueError('Arquivo vazio.')

    _, _, start_serial = _parse_period(periodo_inicial)
    _, _, end_serial = _parse_period(periodo_final)
    if start_serial > end_serial:
        raise ValueError('Periodo inicial nao pode ser maior que o periodo final.')

    rules = _build_mappings(mappings)

    try:
        wb = load_workbook(io.BytesIO(arquivo_bytes), data_only=False, keep_vba=True)
    except Exception as exc:
        raise ValueError('Nao foi possivel ler a planilha. Envie um arquivo Excel valido.') from exc

    sheets_processadas: list[str] = []
    conflitos: list[dict[str, Any]] = []
    linhas_no_periodo = 0
    linhas_alteradas = 0

    for sheet in wb.worksheets:
        cols = _header_columns(sheet)
        if not cols:
            continue

        col_data, col_historico, col_agrupamento = cols
        sheets_processadas.append(sheet.title)

        for row_idx in range(2, sheet.max_row + 1):
            data_value = sheet.cell(row=row_idx, column=col_data).value
            row_date = _parse_date_value(data_value)
            if not row_date:
                continue

            row_serial = row_date.year * 12 + row_date.month
            if row_serial < start_serial or row_serial > end_serial:
                continue

            linhas_no_periodo += 1
            historico_value = sheet.cell(row=row_idx, column=col_historico).value
            historico_text = str(historico_value or '').strip()
            historico_norm = _normalize_text(historico_text)
            if not historico_norm:
                continue

            matches = [rule for rule in rules if rule['historico_norm'] and rule['historico_norm'] in historico_norm]
            if not matches:
                continue

            matches_sorted = sorted(matches, key=lambda item: (-len(item['historico_norm']), item['index']))
            winner = matches_sorted[0]

            if len(matches_sorted) > 1:
                second = matches_sorted[1]
                tie = len(winner['historico_norm']) == len(second['historico_norm'])
                conflitos.append(
                    {
                        'aba': sheet.title,
                        'linha': row_idx,
                        'historico': historico_text,
                        'regras': [rule['historico'] for rule in matches_sorted],
                        'criterio': 'empate-primeira-regra' if tie else 'mais-especifica',
                        'aplicada': winner['historico'],
                    }
                )

            current_group = str(sheet.cell(row=row_idx, column=col_agrupamento).value or '').strip()
            if current_group != winner['agrupamento']:
                sheet.cell(row=row_idx, column=col_agrupamento).value = winner['agrupamento']
                linhas_alteradas += 1

    if not sheets_processadas:
        raise ValueError('Nenhuma aba de extrato foi encontrada com as colunas Data, Historico e Agrupamento.')

    output = io.BytesIO()
    wb.save(output)

    return {
        'ok': True,
        'xlsx_bytes': output.getvalue(),
        'resumo': {
            'abasProcessadas': sheets_processadas,
            'totalAbasProcessadas': len(sheets_processadas),
            'linhasNoPeriodo': linhas_no_periodo,
            'linhasAlteradas': linhas_alteradas,
            'totalRegras': len(rules),
            'totalConflitos': len(conflitos),
            'conflitos': conflitos[:500],
        },
    }
