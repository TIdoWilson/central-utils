from __future__ import annotations

import calendar
import contextlib
import io
import json
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

JSON_MARKER = "__LRENASUL_JSON__"

try:  # local-only: Windows + Excel are the preferred path
    import win32com.client
except Exception:  # pragma: no cover - fallback optional
    win32com = None

try:
    from python_calamine import load_workbook as load_calamine_workbook
except Exception:  # pragma: no cover - fallback optional
    load_calamine_workbook = None

try:
    import xlrd
except Exception:  # pragma: no cover - fallback optional
    xlrd = None

try:
    from openpyxl import load_workbook as load_openpyxl_workbook
except Exception:  # pragma: no cover - fallback optional
    load_openpyxl_workbook = None


def normalize_text(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_text_key(value: Any) -> str:
    return normalize_text(value).lower()


def normalize_digits(value: Any) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def parse_decimal(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, datetime):
        return None

    text = normalize_text(value)
    if not text:
        return None

    text = text.replace("R$", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def format_money(value: Any) -> str:
    decimal_value = parse_decimal(value)
    if decimal_value is None:
        return "0,00"
    quantized = decimal_value.quantize(Decimal("0.01"))
    return f"{quantized:.2f}".replace(".", ",")


def month_end_from(value: date) -> date:
    last_day = calendar.monthrange(value.year, value.month)[1]
    return date(value.year, value.month, last_day)


def to_ddmmyyyy(value: date) -> str:
    return value.strftime("%d%m%Y")


def _matrix_from_used_range(used_range) -> List[List[Any]]:
    values = used_range.Value
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [[values]]
    if values and not isinstance(values[0], tuple):
        return [list(values)]
    return [list(row) if isinstance(row, tuple) else [row] for row in values]


def _com_rows(input_path: Path) -> tuple[List[List[Any]], str]:
    if win32com is None:
        raise RuntimeError("Excel/COM indisponivel no ambiente atual.")

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(
            str(input_path),
            ReadOnly=True,
            UpdateLinks=0,
            IgnoreReadOnlyRecommended=True,
        )
        worksheet = workbook.Worksheets(1)
        rows = _matrix_from_used_range(worksheet.UsedRange)
        return rows, worksheet.Name
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        try:
            excel.Quit()
        except Exception:
            pass


def _calamine_rows(input_path: Path) -> tuple[List[List[Any]], str]:
    if load_calamine_workbook is None:
        raise RuntimeError("python-calamine nao instalado para leitura de planilhas.")

    workbook = load_calamine_workbook(str(input_path))
    sheet_name = workbook.sheet_names[0]
    worksheet = workbook.get_sheet_by_name(sheet_name)
    rows = worksheet.to_python()
    return [list(row) if isinstance(row, (list, tuple)) else [row] for row in rows], sheet_name


def _xlrd_rows(input_path: Path) -> tuple[List[List[Any]], str]:
    if xlrd is None:
        raise RuntimeError("xlrd nao instalado para leitura de .xls.")

    workbook = xlrd.open_workbook(str(input_path))
    worksheet = workbook.sheet_by_index(0)
    rows: List[List[Any]] = []

    for row_idx in range(worksheet.nrows):
      row: List[Any] = []
      for col_idx in range(worksheet.ncols):
          ctype = worksheet.cell_type(row_idx, col_idx)
          value = worksheet.cell_value(row_idx, col_idx)
          if ctype == xlrd.XL_CELL_EMPTY:
              row.append("")
              continue
          if ctype == xlrd.XL_CELL_DATE:
              try:
                  row.append(xlrd.xldate_as_datetime(value, workbook.datemode))
              except Exception:
                  row.append(value)
              continue
          row.append(value)
      rows.append(row)

    return rows, worksheet.name


def _openpyxl_rows(input_path: Path) -> tuple[List[List[Any]], str]:
    if load_openpyxl_workbook is None:
        raise RuntimeError("openpyxl nao instalado para leitura de .xlsx/.xlsm.")

    workbook = load_openpyxl_workbook(str(input_path), data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: List[List[Any]] = []

    try:
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return rows, worksheet.title


def read_workbook_rows(input_path: Path) -> tuple[List[List[Any]], str]:
    suffix = input_path.suffix.lower()
    if suffix in {".xls", ".xlsx", ".xlsm"} and load_calamine_workbook is not None:
        try:
            return _calamine_rows(input_path)
        except Exception:
            pass

    if suffix in {".xls", ".xlsx", ".xlsm"} and win32com is not None:
        try:
            return _com_rows(input_path)
        except Exception:
            if suffix == ".xls" and xlrd is not None:
                return _xlrd_rows(input_path)
            if suffix in {".xlsx", ".xlsm"} and load_openpyxl_workbook is not None:
                return _openpyxl_rows(input_path)
            raise

    if suffix == ".xls":
        return _xlrd_rows(input_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _openpyxl_rows(input_path)

    raise RuntimeError(f"Formato de arquivo nao suportado: {suffix or 'desconhecido'}.")


def parse_centers_list(raw: Any) -> set[str]:
    return {
        normalize_digits(part)
        for part in normalize_text(raw).split(",")
        if normalize_digits(part)
    }


def center_type_for(center_number: str, config: Dict[str, Any]) -> str:
    adm = parse_centers_list((config.get("centrosCusto") or {}).get("adm", "2,4"))
    prod = parse_centers_list((config.get("centrosCusto") or {}).get("producao", "1,5,6,7"))
    number = normalize_digits(center_number)
    if number in adm:
        return "adm"
    if number in prod:
        return "producao"
    return ""


def _config_depara_rows(config: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = config.get("dePara")
    if not isinstance(rows, list) or not rows:
        rows = config.get("deParaRows")
    return [row for row in rows if isinstance(row, dict)]


def _find_depara_row(config: Dict[str, Any], rubrica: Any) -> Optional[Dict[str, Any]]:
    rubrica_key = normalize_digits(rubrica)
    if not rubrica_key:
        return None

    for row in _config_depara_rows(config):
        row_key = normalize_digits(row.get("rubrica") or row.get("codigo") or "")
        if row_key != rubrica_key:
            continue
        return row

    return None


def _mapping_accounts(row: Optional[Dict[str, Any]]) -> Dict[str, str]:
    row = row or {}
    return {
        "debito_prod": normalize_text(row.get("contaDebitoProducao") or row.get("debitoProducao") or row.get("debito_producao") or ""),
        "credito_prod": normalize_text(row.get("contaCreditoProducao") or row.get("creditoProducao") or row.get("credito_producao") or ""),
        "debito_adm": normalize_text(row.get("contaDebitoAdm") or row.get("debitoAdm") or row.get("debito_adm") or ""),
        "credito_adm": normalize_text(row.get("contaCreditoAdm") or row.get("creditoAdm") or row.get("credito_adm") or ""),
    }


def lookup_mapping(config: Dict[str, Any], rubrica: Any, center_type: str) -> Optional[Dict[str, str]]:
    rubrica_key = normalize_digits(rubrica)
    if not rubrica_key:
        return None

    row = _find_depara_row(config, rubrica_key)
    if not row:
        return None

    accounts = _mapping_accounts(row)
    if center_type == "adm":
        debito = accounts["debito_adm"]
        credito = accounts["credito_adm"]
    else:
        debito = accounts["debito_prod"]
        credito = accounts["credito_prod"]

    return {
        "rubrica": rubrica_key,
        "nome": normalize_text(row.get("nome") or row.get("name") or ""),
        "debito": debito,
        "credito": credito,
    }


def detect_competence(rows: List[List[Any]]) -> date:
    for row in rows[:20]:
        for idx, cell in enumerate(row[:12]):
            text = normalize_text_key(cell)
            if "compet" not in text:
                continue
            for probe in row[idx + 1 : idx + 5]:
                if isinstance(probe, datetime):
                    return probe.date()
                if isinstance(probe, date):
                    return probe
                txt = normalize_text(probe)
                if not txt:
                    continue
                for fmt in ("%d/%m/%Y", "%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%y"):
                    try:
                        parsed = datetime.strptime(txt, fmt)
                        return parsed.date()
                    except ValueError:
                        continue
    return date.today()


def _row_text(row: Iterable[Any]) -> str:
    return " | ".join(normalize_text(cell) for cell in row if normalize_text(cell))


def _first_nonempty(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def _as_float(value: Any) -> float:
    decimal_value = parse_decimal(value)
    return float(decimal_value) if decimal_value is not None else 0.0


def parse_summary_rows(rows: List[List[Any]], config: Dict[str, Any], source_name: str) -> Dict[str, Any]:
    competence = detect_competence(rows)
    competence_end = month_end_from(competence)
    competence_text = to_ddmmyyyy(competence_end)

    preview_eventos: List[Dict[str, Any]] = []
    preview_pendencias: List[Dict[str, Any]] = []
    entries: List[Dict[str, Any]] = []
    centers_seen: set[str] = set()
    pending_index: Dict[str, Dict[str, Any]] = {}
    current_center_number = ""
    current_center_name = ""
    current_section = ""

    skip_patterns = (
        "total:",
        "total ",
        "liquido c.custo",
        "liquido ccusto",
        "resumo do centro de custo",
        "resumo",
        "informativa",
        "informativo",
    )

    def register_pending(rubrica: str, nome: str, center_label: str, motivo: str, value: Decimal) -> None:
        rubrica_key = normalize_digits(rubrica)
        if not rubrica_key:
            return

        current = pending_index.get(rubrica_key)
        if current is None:
            current = {
                "rubrica": rubrica_key,
                "nome": normalize_text(nome),
                "centro": normalize_text(center_label),
                "motivo": normalize_text(motivo),
                "valor_total": Decimal("0"),
                "centros": [],
                "centros_seen": set(),
            }
            pending_index[rubrica_key] = current

        if normalize_text(nome) and not current["nome"]:
            current["nome"] = normalize_text(nome)
        if normalize_text(motivo) and not current["motivo"]:
            current["motivo"] = normalize_text(motivo)

        center_text = normalize_text(center_label)
        if center_text and center_text not in current["centros_seen"]:
            current["centros_seen"].add(center_text)
            current["centros"].append(center_text)

        current["valor_total"] += Decimal(value)

    for row in rows:
        row_text = _row_text(row)
        row_text_key = normalize_text_key(row_text)

        center_match = re.search(r"c\.?\s*custo:\s*(\d+)\s*-\s*(.+)$", row_text, re.IGNORECASE)
        if center_match:
            current_center_number = normalize_digits(center_match.group(1))
            current_center_name = normalize_text(center_match.group(2))
            centers_seen.add(current_center_number)
            current_section = ""
            continue

        if "proventos" in row_text_key:
            current_section = "PROVENTOS"
            continue
        if "descontos" in row_text_key:
            current_section = "DESCONTOS"
            continue

        if not current_center_number:
            continue
        if not row_text_key:
            continue
        if any(pattern in row_text_key for pattern in skip_patterns):
            continue

        rubrica = normalize_digits(row[0] if len(row) > 0 else "")
        nome = _first_nonempty(row[3] if len(row) > 3 else "", row[1] if len(row) > 1 else "")
        if not rubrica or not nome:
            continue

        star_flag = normalize_text(row[62] if len(row) > 62 else "")
        if star_flag == "*":
            continue

        value_raw = row[56] if len(row) > 56 else None
        if parse_decimal(value_raw) is None:
            value_raw = row[45] if len(row) > 45 else None
        value = parse_decimal(value_raw) or Decimal("0")
        if value <= 0:
            continue

        center_type = center_type_for(current_center_number, config)
        preview_item = {
            "rubrica": rubrica,
            "nome": nome,
            "centro": f"{current_center_number} - {current_center_name}".strip(" -"),
            "centroNumero": current_center_number,
            "centroNome": current_center_name,
            "valor": format_money(value),
        }

        depara_row = _find_depara_row(config, rubrica)
        accounts = _mapping_accounts(depara_row)
        has_complete_depara = bool(accounts["debito_prod"] and accounts["credito_prod"] and accounts["debito_adm"] and accounts["credito_adm"])

        if not has_complete_depara:
            register_pending(
                rubrica,
                nome,
                f"{current_center_number} - {current_center_name}".strip(" -"),
                "rubrica_sem_conta_cadastrada",
                value,
            )
        if not center_type:
            preview_item["status"] = "pendencia"
            preview_item["motivo"] = "centro_nao_classificado"
            register_pending(
                rubrica,
                nome,
                f"{current_center_number} - {current_center_name}".strip(" -"),
                "centro_nao_classificado",
                value,
            )
            preview_eventos.append(preview_item)
            continue

        mapping = lookup_mapping(config, rubrica, center_type)
        if not mapping or not mapping.get("debito") or not mapping.get("credito"):
            preview_item["status"] = "pendencia"
            preview_item["motivo"] = "rubrica_sem_conta_cadastrada"
            register_pending(
                rubrica,
                nome,
                f"{current_center_number} - {current_center_name}".strip(" -"),
                "rubrica_sem_conta_cadastrada",
                value,
            )
            preview_eventos.append(preview_item)
            continue

        entry = {
            "rubrica": rubrica,
            "nome": nome,
            "centerType": center_type,
            "centerNumber": current_center_number,
            "centerName": current_center_name,
            "debitAccount": mapping["debito"],
            "creditAccount": mapping["credito"],
            "complement": f"{rubrica} - {nome}",
            "history": nome,
            "value": float(value),
            "missing": False,
        }
        entries.append(entry)

        preview_item["status"] = "mapeado"
        preview_item["motivo"] = ""
        preview_item["debito"] = mapping["debito"]
        preview_item["credito"] = mapping["credito"]
        preview_eventos.append(preview_item)

    total_valor = sum(Decimal(str(row.get("value", 0) or 0)) for row in entries)
    preview_pendencias = []
    for pending in pending_index.values():
        centros = pending.get("centros") or []
        centro_text = " | ".join(centros)
        preview_pendencias.append({
          "rubrica": pending.get("rubrica", ""),
          "nome": pending.get("nome", ""),
          "centro": centro_text,
          "motivo": pending.get("motivo", ""),
          "valor": format_money(pending.get("valor_total", Decimal("0"))),
      })

    preview_pendencias.sort(key=lambda row: (normalize_digits(row.get("rubrica")), normalize_text(row.get("nome"))))
    pode_gerar_txt = len(preview_pendencias) == 0 and len(entries) > 0

    resumo = {
        "source": source_name,
        "competenceDate": competence_end.isoformat(),
        "competenceDateText": competence_text,
        "total_registros": len(preview_eventos),
        "total_pendencias": len(preview_pendencias),
        "total_centros": len(centers_seen),
        "total_valor": float(total_valor),
        "validado": len(preview_pendencias) == 0,
        "pode_gerar_txt": pode_gerar_txt,
        "gerou_txt": False,
        "message": (
            "Nenhuma conta em falta para este arquivo."
            if pode_gerar_txt
            else f"Validacao encontrou {len(preview_pendencias)} pendencia(s)."
        ),
        "rows": entries,
        "preview_eventos": preview_eventos,
        "preview_pendencias": preview_pendencias,
        "preview_linhas": [],
    }

    return resumo


def processar_lotes_renasul(payload: Dict[str, Any]) -> Dict[str, Any]:
    config = payload.get("config") if isinstance(payload.get("config"), dict) else {}
    files = payload.get("files") if isinstance(payload.get("files"), list) else []
    if not files:
        raise ValueError("Nenhum arquivo foi informado.")

    input_file = files[0] if isinstance(files[0], dict) else {}
    file_path = Path(str(input_file.get("path") or "").strip())
    if not file_path.is_file():
        raise FileNotFoundError(f"Arquivo de entrada nao encontrado: {file_path}")

    rows, sheet_name = read_workbook_rows(file_path)
    resumo = parse_summary_rows(rows, config, source_name=str(input_file.get("name") or file_path.name))
    resumo["sheetName"] = sheet_name

    return {
        "ok": True,
        "resumo": resumo,
    }


def _read_payload() -> Dict[str, Any]:
    raw = sys.stdin.read()
    if not raw.strip():
        return {}
    return json.loads(raw)


def main() -> int:
    try:
        payload = _read_payload()
        stdout_buffer = io.StringIO()
        stderr_buffer = io.StringIO()
        with contextlib.redirect_stdout(stdout_buffer), contextlib.redirect_stderr(stderr_buffer):
            result = processar_lotes_renasul(payload)
        warning_text = "\n".join(
            part.strip()
            for part in [stdout_buffer.getvalue(), stderr_buffer.getvalue()]
            if part and part.strip()
        ).strip()
        if warning_text:
            sys.stderr.write(warning_text + "\n")
        sys.stdout.write(f"{JSON_MARKER}\n")
        sys.stdout.write(json.dumps(result, ensure_ascii=False))
        return 0
    except Exception as exc:
        error = {
            "ok": False,
            "error": f"Erro ao validar lotes Renasul: {exc}",
        }
        sys.stdout.write(f"{JSON_MARKER}\n")
        sys.stdout.write(json.dumps(error, ensure_ascii=False))
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
