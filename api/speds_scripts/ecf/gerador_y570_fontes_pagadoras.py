from __future__ import annotations

import argparse
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from tkinter import Tk, filedialog, messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_CODES_PATH = SCRIPT_DIR / "assets" / "codigos_aceitos.xlsx"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "Saidas"
CHECK_YES = "☑"
CHECK_NO = "☐"
MAIN_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECONDARY_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WARNING_FILL = PatternFill("solid", fgColor="FDE9D9")
ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
LOCKED_FILL = PatternFill("solid", fgColor="EDEDED")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


@dataclass(frozen=True)
class AcceptedCode:
    code: str
    description: str
    orgao_publico: str
    irrf: str
    csll: str


@dataclass
class SourceRow:
    cnpj: str
    nome: str
    code: str
    orgao_publico: str
    rendimento_original: Decimal
    tributo_retido_original: Decimal
    rendimento_ecf: Decimal
    aliquota: Decimal | None
    irrf: Decimal
    csll: Decimal
    code_found: bool
    editable_orgao_publico: bool
    source_count: int = 1


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value: object) -> str:
    return strip_accents(clean_text(value)).casefold()


def normalize_code(value: object) -> str:
    text = clean_text(value)
    digits = re.sub(r"\D", "", text)
    if not digits:
        return ""
    return digits.zfill(4)[-4:]


def normalize_flag(value: object) -> str:
    text = strip_accents(clean_text(value)).casefold()
    if text in {"sim", "s", "yes", "true", "1"}:
        return "Sim"
    if text in {"sim ou nao", "sim ou nao.", "sim ou nao ", "sim ou nao/nao"} or "sim ou nao" in text:
        return "Sim ou Nao"
    if text.startswith("n") or text in {"false", "0"}:
        return "Nao"
    return clean_text(value)

def to_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = clean_text(value)
    if not text:
        return Decimal("0")
    text = text.replace(".", "").replace(",", "")
    text = re.sub(r"[^\d\-]", "", text)
    if not text or text == "-":
        return Decimal("0")
    return Decimal(text)


def cents_to_money(value: object) -> Decimal:
    cents = to_decimal(value)
    return (cents / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def money_to_float(value: Decimal | None) -> float:
    if value is None:
        return 0.0
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def select_input_files() -> tuple[list[Path], Path, Path]:
    root = Tk()
    root.withdraw()
    root.update()

    messagebox.showinfo(
        "Y570 ECF",
        "Selecione os TXT das fontes pagadoras.\n"
        "A planilha de codigos aceitos e fixa do sistema e, por fim, escolha o destino do XLSX.",
    )

    txt_paths = filedialog.askopenfilenames(
        title="Selecione os TXT das fontes pagadoras",
        filetypes=[("Arquivos TXT", "*.txt *.TXT"), ("Todos os arquivos", "*.*")],
    )
    if not txt_paths:
        root.destroy()
        raise RuntimeError("Nenhum TXT foi selecionado.")

    codes_path = DEFAULT_CODES_PATH
    if not codes_path.exists():
        root.destroy()
        raise RuntimeError(f"Planilha de codigos aceitos nao encontrada: {codes_path}")

    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    default_name = f"Y570_fontes_pagadoras_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = filedialog.asksaveasfilename(
        title="Salvar XLSX gerado",
        defaultextension=".xlsx",
        initialdir=str(DEFAULT_OUTPUT_DIR),
        initialfile=default_name,
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
    )
    root.destroy()

    if not output_path:
        raise RuntimeError("Nenhum destino de saida foi selecionado.")

    return [Path(p) for p in txt_paths], Path(codes_path), Path(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Gera o Y570 da ECF a partir dos TXT das fontes pagadoras.")
    parser.add_argument(
        "--txt",
        action="append",
        default=[],
        help="Caminho de um TXT das fontes pagadoras. Pode ser repetido.",
    )
    parser.add_argument(
        "--codes",
        help="Opcional. Sobrescreve a planilha fixa de codigos aceitos (.xlsx).",
    )
    parser.add_argument(
        "--output",
        help="Caminho do XLSX de saida.",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Nao abre popups e exige que os caminhos sejam informados por argumento.",
    )
    return parser.parse_args()


def resolve_inputs(args: argparse.Namespace) -> tuple[list[Path], Path, Path, bool]:
    if args.no_gui:
        txt_paths = [Path(item) for item in (args.txt or []) if str(item).strip()]
        codes_path = Path(args.codes) if args.codes else DEFAULT_CODES_PATH
        output_path = Path(args.output) if args.output else None
        if not txt_paths:
            raise RuntimeError("Em modo --no-gui, informe ao menos um --txt.")
        if not codes_path.exists():
            raise RuntimeError(f"Planilha de codigos aceitos nao encontrada: {codes_path}")
        if output_path is None:
            raise RuntimeError("Em modo --no-gui, informe --output.")
        return txt_paths, codes_path, output_path, False

    txt_paths, codes_path, output_path = select_input_files()
    return txt_paths, codes_path, output_path, True


def parse_txt_file(path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    with path.open("r", encoding="utf-8", errors="ignore") as handle:
        for raw_line in handle:
            line = raw_line.rstrip("\r\n")
            if len(line) < 213:
                continue
            if not line[:14].isdigit():
                continue

            cnpj = line[0:14]
            nome = clean_text(line[17:168])
            code = normalize_code(line[177:181])
            if not code:
                continue

            rendimento_original = cents_to_money(line[182:197])
            tributo_retido_original = cents_to_money(line[198:213])
            rows.append(
                {
                    "cnpj": cnpj,
                    "nome": nome,
                    "code": code,
                    "rendimento_original": rendimento_original,
                    "tributo_retido_original": tributo_retido_original,
                    "source_file": path.name,
                }
            )
    return rows


def read_accepted_codes(path: Path) -> dict[str, AcceptedCode]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    accepted: dict[str, AcceptedCode] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_code(row[0] if len(row) > 0 else None)
        if not code:
            continue
        accepted[code] = AcceptedCode(
            code=code,
            description=clean_text(row[1] if len(row) > 1 else ""),
            orgao_publico=normalize_flag(row[2] if len(row) > 2 else ""),
            irrf=normalize_flag(row[3] if len(row) > 3 else ""),
            csll=normalize_flag(row[4] if len(row) > 4 else ""),
        )
    return accepted


def aggregate_sources(txt_paths: list[Path]) -> tuple[list[SourceRow], dict[str, dict[str, object]]]:
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    names_by_group: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    files_by_group: dict[tuple[str, str], set[str]] = defaultdict(set)

    for path in txt_paths:
        for item in parse_txt_file(path):
            key = (item["cnpj"], item["code"])
            bucket = grouped.setdefault(
                key,
                {
                    "cnpj": item["cnpj"],
                    "code": item["code"],
                    "rendimento_original": Decimal("0"),
                    "tributo_retido_original": Decimal("0"),
                },
            )
            bucket["rendimento_original"] = bucket["rendimento_original"] + item["rendimento_original"]
            bucket["tributo_retido_original"] = (
                bucket["tributo_retido_original"] + item["tributo_retido_original"]
            )
            names_by_group[key][item["nome"]] += 1
            files_by_group[key].add(item["source_file"])

    rows: list[SourceRow] = []
    metadata: dict[str, dict[str, object]] = {}
    for (cnpj, code), bucket in sorted(
        grouped.items(),
        key=lambda item: (item[0][0], item[0][1]),
    ):
        nome = names_by_group[(cnpj, code)].most_common(1)[0][0]
        metadata_key = f"{cnpj}|{code}"
        metadata[metadata_key] = {
            "source_files": sorted(files_by_group[(cnpj, code)]),
            "occurrences": sum(names_by_group[(cnpj, code)].values()),
        }
        rows.append(
            SourceRow(
                cnpj=cnpj,
                nome=nome,
                code=code,
                orgao_publico=CHECK_NO,
                rendimento_original=bucket["rendimento_original"],
                tributo_retido_original=bucket["tributo_retido_original"],
                rendimento_ecf=bucket["rendimento_original"],
                aliquota=None,
                irrf=Decimal("0"),
                csll=Decimal("0"),
                code_found=False,
                editable_orgao_publico=False,
                source_count=sum(names_by_group[(cnpj, code)].values()),
            )
        )
    return rows, metadata


def update_row_calculations(row: SourceRow, accepted: dict[str, AcceptedCode]) -> None:
    code_meta = accepted.get(row.code)
    row.code_found = code_meta is not None
    row.editable_orgao_publico = bool(code_meta and code_meta.orgao_publico == "Sim ou Nao")
    if code_meta:
        row.orgao_publico = CHECK_YES if code_meta.orgao_publico == "Sim" else CHECK_NO
    else:
        row.orgao_publico = CHECK_NO

    if row.aliquota is None:
        row.irrf = Decimal("0")
        row.csll = Decimal("0")
        return

    aliquota = row.aliquota / Decimal("100")
    row.irrf = (
        (row.rendimento_ecf * aliquota).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if code_meta and code_meta.irrf == "Sim"
        else Decimal("0")
    )
    row.csll = (
        (row.rendimento_ecf * aliquota).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if code_meta and code_meta.csll == "Sim"
        else Decimal("0")
    )


def build_code_rows(rows: list[SourceRow], accepted: dict[str, AcceptedCode]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    found_codes = sorted({row.code for row in rows})
    code_rows: list[dict[str, object]] = []
    missing_rows: list[dict[str, object]] = []

    for code in found_codes:
        meta = accepted.get(code)
        count = sum(1 for row in rows if row.code == code)
        if meta is None:
            missing_rows.append(
                {
                    "codigo": code,
                    "descricao": "",
                    "ocorrencias": count,
                }
            )
            continue

        code_rows.append(
            {
                "codigo": code,
                "descricao": meta.description,
                "orgao_publico": meta.orgao_publico,
                "irrf": meta.irrf,
                "csll": meta.csll,
                "aliquota": None,
                "ocorrencias": count,
            }
        )

    return code_rows, missing_rows


def autosize_columns(ws, widths: dict[int, int]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[chr(64 + index)].width = width


def style_header(ws, row_idx: int, fill: PatternFill) -> None:
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_table_cell(cell, locked: bool = True, fill: PatternFill | None = None) -> None:
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.protection = Protection(locked=locked, hidden=False)
    if fill is not None:
        cell.fill = fill


def add_hidden_list_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("_listas")
    ws.sheet_state = "hidden"
    ws["A1"] = CHECK_NO
    ws["A2"] = CHECK_YES


def write_summary_sheet(wb: Workbook, txt_paths: list[Path], codes_path: Path, output_path: Path, rows: list[SourceRow], code_rows: list[dict[str, object]], missing_rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Resumo", 0)
    ws["A1"] = "Y570 - Fontes pagadoras"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Arquivos TXT"
    ws["B3"] = len(txt_paths)
    ws["A4"] = "Planilha de codigos"
    ws["B4"] = codes_path.name
    ws["A5"] = "Saida"
    ws["B5"] = output_path.name
    ws["A7"] = "Fontes agrupadas"
    ws["B7"] = len(rows)
    ws["A8"] = "Codigos localizados na base aceita"
    ws["B8"] = len(code_rows)
    ws["A9"] = "Codigos nao localizados"
    ws["B9"] = len(missing_rows)
    ws["A11"] = "Uso"
    ws["B11"] = "Preencha a coluna Aliquota (%) na aba Codigos e ajuste os campos editaveis."
    ws["B12"] = "Checkbox: " + CHECK_YES + " = Sim, " + CHECK_NO + " = Nao."
    ws["B13"] = "O rendimento ECF abre igual ao rendimento original e pode ser alterado."
    for row in range(3, 14):
        ws[f"A{row}"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90


def write_codes_sheet(wb: Workbook, code_rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Codigos")
    headers = [
        "Codigo",
        "Descricao",
        "Orgao publico aceito",
        "IRRF",
        "CSLL",
        "Aliquota (%)",
        "Ocorrencias",
    ]
    ws.append(headers)
    style_header(ws, 1, MAIN_HEADER_FILL)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(len(code_rows) + 1, 1)}"
    aliquota_validation = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="100",
        allow_blank=True,
    )
    ws.add_data_validation(aliquota_validation)

    for idx, item in enumerate(code_rows, start=2):
        ws.cell(idx, 1, item["codigo"])
        ws.cell(idx, 2, item["descricao"])
        ws.cell(idx, 3, item["orgao_publico"])
        ws.cell(idx, 4, item["irrf"])
        ws.cell(idx, 5, item["csll"])
        ws.cell(idx, 6, None)
        ws.cell(idx, 7, item["ocorrencias"])
        for col in range(1, 8):
            locked = col != 6
            fill = EDITABLE_FILL if col == 6 else None
            style_table_cell(ws.cell(idx, col), locked=locked, fill=fill)
            if col == 6:
                ws.cell(idx, col).number_format = "0.00"
                aliquota_validation.add(ws.cell(idx, col).coordinate)

    if not code_rows:
        ws["A2"] = "Nenhum codigo aceito encontrado nos TXT."
        ws["A2"].fill = WARNING_FILL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14


def write_missing_codes_sheet(wb: Workbook, missing_rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Nao_localizados")
    headers = ["Codigo", "Descricao", "Ocorrencias"]
    ws.append(headers)
    style_header(ws, 1, MAIN_HEADER_FILL)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max(len(missing_rows) + 1, 1)}"

    for idx, item in enumerate(missing_rows, start=2):
        ws.cell(idx, 1, item["codigo"])
        ws.cell(idx, 2, "NAO ENCONTRADO NA PLANILHA DE CODIGOS ACEITOS")
        ws.cell(idx, 3, item["ocorrencias"])
        for col in range(1, 4):
            style_table_cell(ws.cell(idx, col), locked=True, fill=ERROR_FILL if col == 2 else None)

    if not missing_rows:
        ws["A2"] = "Nenhum codigo ficou fora da lista aceita."
        ws["A2"].fill = WARNING_FILL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 14


def write_sources_sheet(wb: Workbook, rows: list[SourceRow], accepted: dict[str, AcceptedCode]) -> None:
    ws = wb.create_sheet("Fontes")
    headers = [
        "CNPJ",
        "Nome",
        "Codigo DARF",
        "Orgao publico",
        "Rendimento original",
        "Tributo retido original",
        "Rendimento ECF",
        "IRRF",
        "CSLL",
        "Ocorrencias",
    ]
    ws.append(headers)
    style_header(ws, 1, MAIN_HEADER_FILL)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{max(len(rows) + 1, 1)}"

    checkbox_validation = DataValidation(
        type="list",
        formula1='"' + CHECK_NO + "," + CHECK_YES + '"',
        allow_blank=False,
    )
    ws.add_data_validation(checkbox_validation)

    for idx, row in enumerate(rows, start=2):
        update_row_calculations(row, accepted)

        ws.cell(idx, 1, row.cnpj)
        ws.cell(idx, 2, row.nome)
        ws.cell(idx, 3, row.code)
        ws.cell(idx, 4, row.orgao_publico)
        ws.cell(idx, 5, money_to_float(row.rendimento_original))
        ws.cell(idx, 6, money_to_float(row.tributo_retido_original))
        ws.cell(idx, 7, money_to_float(row.rendimento_ecf))
        ws.cell(idx, 8, f"=IFERROR(IF(VLOOKUP($C{idx},Codigos!$A:$F,4,FALSE)=\"Sim\",ROUND($G{idx}*VLOOKUP($C{idx},Codigos!$A:$F,6,FALSE)/100,2),0),0)")
        ws.cell(idx, 9, f"=IFERROR(IF(VLOOKUP($C{idx},Codigos!$A:$F,5,FALSE)=\"Sim\",ROUND($G{idx}*VLOOKUP($C{idx},Codigos!$A:$F,6,FALSE)/100,2),0),0)")
        ws.cell(idx, 10, row.source_count)

        for col in range(1, 11):
            cell = ws.cell(idx, col)
            locked = True
            fill = None

            if col == 4:
                locked = not row.editable_orgao_publico
                fill = EDITABLE_FILL if row.editable_orgao_publico else LOCKED_FILL
                if row.editable_orgao_publico:
                    checkbox_validation.add(cell.coordinate)
            elif col == 7:
                locked = False
                fill = EDITABLE_FILL
            elif col in {8, 9}:
                locked = True
                fill = None
            elif col in {1, 2, 3, 5, 6, 10}:
                locked = True
                fill = None

            style_table_cell(cell, locked=locked, fill=fill)

            if col in {5, 6, 7, 8, 9}:
                cell.number_format = '#,##0.00'

        if not row.code_found:
            for col in range(1, 11):
                ws.cell(idx, col).fill = ERROR_FILL

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12


def create_workbook(
    txt_paths: list[Path],
    codes_path: Path,
    output_path: Path,
    rows: list[SourceRow],
    code_rows: list[dict[str, object]],
    missing_rows: list[dict[str, object]],
    accepted: dict[str, AcceptedCode],
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    add_hidden_list_sheet(wb)
    write_summary_sheet(wb, txt_paths, codes_path, output_path, rows, code_rows, missing_rows)
    write_codes_sheet(wb, code_rows)
    write_missing_codes_sheet(wb, missing_rows)
    write_sources_sheet(wb, rows, accepted)

    for ws in wb.worksheets:
        if ws.title == "_listas":
            continue
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.protection.sheet = True
        ws.protection.enable()

    wb.save(output_path)


def build_context(
    txt_paths: list[Path],
    accepted: dict[str, AcceptedCode],
) -> tuple[list[SourceRow], list[dict[str, object]], list[dict[str, object]]]:
    rows, _metadata = aggregate_sources(txt_paths)
    for row in rows:
        code_meta = accepted.get(row.code)
        if code_meta is not None and code_meta.orgao_publico == "Sim":
            row.orgao_publico = CHECK_YES
        elif code_meta is not None and code_meta.orgao_publico == "Nao":
            row.orgao_publico = CHECK_NO
        elif code_meta is not None and code_meta.orgao_publico == "Sim ou Nao":
            row.orgao_publico = CHECK_NO
            row.editable_orgao_publico = True
        else:
            row.orgao_publico = CHECK_NO
            row.editable_orgao_publico = False

        row.aliquota = None
        update_row_calculations(row, accepted)

    code_rows, missing_rows = build_code_rows(rows, accepted)
    return rows, code_rows, missing_rows


def show_finish_message(
    output_path: Path,
    rows: list[SourceRow],
    code_rows: list[dict[str, object]],
    missing_rows: list[dict[str, object]],
) -> None:
    root = Tk()
    root.withdraw()
    messagebox.showinfo(
        "Y570 ECF",
        "Arquivo gerado com sucesso.\n\n"
        f"Saida: {output_path}\n"
        f"Fontes agrupadas: {len(rows)}\n"
        f"Codigos localizados: {len(code_rows)}\n"
        f"Codigos nao localizados: {len(missing_rows)}\n\n"
        "Abra a aba 'Codigos' para preencher a alíquota e a aba 'Fontes' para ajustar os campos editaveis.",
    )
    root.destroy()


def main() -> int:
    args = parse_args()
    use_gui = not args.no_gui
    try:
        txt_paths, codes_path, output_path, use_gui = resolve_inputs(args)
        accepted = read_accepted_codes(codes_path)
        rows, code_rows, missing_rows = build_context(txt_paths, accepted)
        create_workbook(txt_paths, codes_path, output_path, rows, code_rows, missing_rows, accepted)
        if use_gui:
            show_finish_message(output_path, rows, code_rows, missing_rows)
        else:
            print(f"Arquivo gerado com sucesso: {output_path}")
            print(f"Fontes agrupadas: {len(rows)}")
            print(f"Codigos localizados: {len(code_rows)}")
            print(f"Codigos nao localizados: {len(missing_rows)}")
        return 0
    except Exception as exc:
        if use_gui:
            root = Tk()
            root.withdraw()
            messagebox.showerror("Y570 ECF", f"Falha ao gerar o arquivo.\n\n{exc}")
            root.destroy()
        else:
            print(f"Falha ao gerar o arquivo: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
