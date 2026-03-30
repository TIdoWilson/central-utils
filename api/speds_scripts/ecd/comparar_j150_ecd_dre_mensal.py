#!/usr/bin/env python3
"""
Compara os saldos do registro J150 da ECD com a DRE mensal do Excel.

Fluxo:
1. O usuario seleciona o arquivo TXT da ECD.
2. O usuario seleciona o arquivo XLSX com a DRE mensal.
3. O script usa o layout do registro J150 do diretorio `ecd/`.
4. Os registros J150 sao separados em 12 blocos mensais pelo reinicio do
   campo NU_ORDEM.
5. Cada bloco mensal e comparado com a DRE do Excel na mesma ordem do mes.
6. O resultado e salvo em uma planilha XLSX com resumo, comparacoes e
   registros nao mapeados.

Observacao:
- O registro J150 nao traz um campo de mes explicito; por isso o script usa a
  ordem dos blocos mensais no TXT e assume que o primeiro bloco corresponde a
  Janeiro, o segundo a Fevereiro e assim por diante.
"""

from __future__ import annotations

import argparse
import difflib
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
from tkinter import Tk, filedialog, messagebox


BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[3]
LAYOUT_DIR = BASE_DIR / "ecd"
if not LAYOUT_DIR.exists():
    LAYOUT_DIR = PROJECT_ROOT / "api" / "layouts" / "speds" / "ecd"
DEFAULT_MONTHS = [
    "Janeiro",
    "Fevereiro",
    "Marco",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]
DEFAULT_TOLERANCE = Decimal("0.01")
DEFAULT_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


@dataclass
class J150Record:
    block_index: int
    line_number: int
    nu_ordem: int
    cod_agl: str
    ind_cod_agl: str
    nivel_agl: int
    cod_agl_sup: str
    descr_cod_agl: str
    vl_cta_ini: Decimal
    ind_dc_cta_ini: str
    vl_cta_fin: Decimal
    ind_dc_cta_fin: str
    ind_grp_dre: str
    nota_exp_ref: str


@dataclass
class ExcelRow:
    row_number: int
    descricao: str
    descricao_norm: str
    meses: list[Decimal | None]
    total: Decimal | None


@dataclass
class MatchResult:
    mes: str
    ecd: J150Record
    excel: ExcelRow
    valor_mes_excel: Decimal | None
    valor_total_excel: Decimal | None
    diff_mes: Decimal | None
    diff_total: Decimal | None
    status: str


@dataclass
class J100Record:
    block_index: int
    line_number: int
    cod_agl: str
    ind_cod_agl: str
    nivel_agl: int
    cod_agl_sup: str
    ind_grp_bal: str
    descr_cod_agl: str
    vl_cta_ini: Decimal
    ind_dc_cta_ini: str
    vl_cta_fin: Decimal
    ind_dc_cta_fin: str
    nota_exp_ref: str


@dataclass
class BalanceRow:
    row_number: int
    descricao: str
    descricao_norm: str
    valor_final: Decimal | None
    valor_inicial: Decimal | None


@dataclass
class BalanceMatch:
    row_number: int
    excel: BalanceRow
    ecd_inicial: J100Record | None
    ecd_final: J100Record | None
    valor_ecd_inicial: Decimal | None
    valor_ecd_final: Decimal | None
    valor_excel_inicial: Decimal | None
    valor_excel_final: Decimal | None
    diff_inicial: Decimal | None
    diff_final: Decimal | None
    status: str


@dataclass
class BalanceReport:
    matches: list[BalanceMatch]
    unmatched_ecd_rows: list[dict[str, object]]
    unmatched_excel_rows: list[dict[str, object]]
    summary: dict[str, int]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compara o J150 da ECD com a DRE mensal do Excel."
    )
    parser.add_argument("--ecd", help="Caminho do TXT da ECD.")
    parser.add_argument("--excel", help="Caminho do XLSX da DRE mensal.")
    parser.add_argument("--balanco", help="Caminho do XLSX do balanco patrimonial.")
    parser.add_argument("--output", help="Caminho do arquivo XLSX de saida.")
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Nao abre popups e usa apenas os argumentos informados.",
    )
    return parser.parse_args()


def make_root() -> Tk:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return root


def choose_file_dialog(
    title: str,
    filetypes: list[tuple[str, str]],
    initialdir: str | None = None,
) -> Path | None:
    root = make_root()
    selected = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes,
        initialdir=initialdir,
    )
    root.destroy()
    if not selected:
        return None
    return Path(selected)


def choose_save_dialog(title: str, defaultextension: str, filetypes: list[tuple[str, str]], initialfile: str) -> Path | None:
    root = make_root()
    selected = filedialog.asksaveasfilename(
        title=title,
        defaultextension=defaultextension,
        filetypes=filetypes,
        initialfile=initialfile,
    )
    root.destroy()
    if not selected:
        return None
    return Path(selected)


def read_text_with_fallback(path: Path) -> tuple[str, str]:
    last_error: Exception | None = None
    for encoding in DEFAULT_ENCODINGS:
        try:
            return path.read_text(encoding=encoding), encoding
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    raise last_error or UnicodeError(f"Nao foi possivel ler o arquivo: {path}")


def split_sped_line(line: str) -> list[str]:
    return line.lstrip("\ufeff").rstrip("\r\n").split("|")


def get_field(parts: list[str], index: int) -> str:
    if index >= len(parts):
        return ""
    return parts[index]


def normalize_digits(value: str | None) -> str:
    if not value:
        return ""
    return "".join(ch for ch in str(value) if ch.isdigit())


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().strip()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return " ".join(text.split())


def normalize_match_text(value: str | None) -> str:
    text = normalize_text(value)
    aliases = {
        "LUCRO LIQUIDO DO EXERCICIO": "RESULTADO LIQUIDO DO EXERCICIO",
        "LUCRO PREJUIZO LIQUIDO DO EXERCICIO": "RESULTADO LIQUIDO DO EXERCICIO",
        "PREJUIZO LIQUIDO DO EXERCICIO": "RESULTADO LIQUIDO DO EXERCICIO",
    }
    return aliases.get(text, text)


def is_present(value: str | None) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text not in {"-", "--", "N/A"}


def parse_decimal_br(value: str | int | float | Decimal | None) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if text in {"-", "--", "N/A"}:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except Exception:  # noqa: BLE001
        return None


def format_decimal_br(value: Decimal | None, places: int = 2) -> str:
    if value is None:
        return ""
    quant = Decimal("1").scaleb(-places)
    normalized = value.quantize(quant, rounding=ROUND_HALF_UP)
    text = format(normalized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text.replace(".", ",")


def decimal_diff(left: Decimal | None, right: Decimal | None) -> Decimal | None:
    if left is None or right is None:
        return None
    return left - right


def comparison_value(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return abs(value)


def within_tolerance(diff: Decimal | None, tolerance: Decimal) -> bool:
    if diff is None:
        return False
    return abs(diff) <= tolerance


def align_by_description(
    left_items: list[object],
    right_items: list[object],
    left_desc_getter,
    right_desc_getter,
) -> tuple[dict[int, int], list[int], list[int]]:
    left_norm = [normalize_match_text(left_desc_getter(item)) for item in left_items]
    right_norm = [normalize_match_text(right_desc_getter(item)) for item in right_items]
    matcher = difflib.SequenceMatcher(a=left_norm, b=right_norm, autojunk=False)
    mapping: dict[int, int] = {}
    matched_left: set[int] = set()
    matched_right: set[int] = set()

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for i, j in zip(range(i1, i2), range(j1, j2)):
                mapping[i] = j
                matched_left.add(i)
                matched_right.add(j)

    unmatched_left = [i for i in range(len(left_items)) if i not in matched_left]
    unmatched_right = [j for j in range(len(right_items)) if j not in matched_right]
    return mapping, unmatched_left, unmatched_right


def load_layout_fields(layout_name: str) -> list[str]:
    path = LAYOUT_DIR / layout_name
    if not path.exists():
        raise FileNotFoundError(f"Layout nao encontrado: {path}")
    data = json.loads(path.read_text(encoding="utf-8"))
    return [campo["campo"] for campo in data["campos"]]


def build_field_index(field_names: list[str]) -> dict[str, int]:
    return {name: idx + 1 for idx, name in enumerate(field_names)}


def parse_j150_blocks(path: Path) -> tuple[list[list[J150Record]], str]:
    layout_fields = load_layout_fields("J150.json")
    index = build_field_index(layout_fields)
    text, encoding = read_text_with_fallback(path)
    blocks: list[list[J150Record]] = []
    current: list[J150Record] = []
    line_number = 0

    for raw_line in text.splitlines():
        line_number += 1
        if not raw_line.startswith("|J150|"):
            continue
        parts = split_sped_line(raw_line)
        nivel_text = get_field(parts, index["NIVEL_AGL"])
        nu_ordem_text = get_field(parts, index["NU_ORDEM"])
        try:
            nivel_agl = int(nivel_text or "0")
        except ValueError:
            continue
        if nivel_agl > 5:
            continue
        try:
            nu_ordem = int(nu_ordem_text or "0")
        except ValueError:
            nu_ordem = 0

        if nu_ordem == 0 and current:
            blocks.append(current)
            current = []

        rec = J150Record(
            block_index=len(blocks) + 1,
            line_number=line_number,
            nu_ordem=nu_ordem,
            cod_agl=get_field(parts, index["COD_AGL"]).strip(),
            ind_cod_agl=get_field(parts, index["IND_COD_AGL"]).strip(),
            nivel_agl=nivel_agl,
            cod_agl_sup=get_field(parts, index["COD_AGL_SUP"]).strip(),
            descr_cod_agl=get_field(parts, index["DESCR_COD_AGL"]).strip(),
            vl_cta_ini=parse_decimal_br(get_field(parts, index["VL_CTA_INI"])) or Decimal("0"),
            ind_dc_cta_ini=get_field(parts, index["IND_DC_CTA_INI"]).strip(),
            vl_cta_fin=parse_decimal_br(get_field(parts, index["VL_CTA_FIN"])) or Decimal("0"),
            ind_dc_cta_fin=get_field(parts, index["IND_DC_CTA_FIN"]).strip(),
            ind_grp_dre=get_field(parts, index["IND_GRP_DRE"]).strip(),
            nota_exp_ref=get_field(parts, index["NOTA_EXP_REF"]).strip(),
        )
        current.append(rec)

    if current:
        blocks.append(current)
    return blocks, encoding


def detect_header_row(ws) -> int:
    months_norm = [normalize_text(month) for month in DEFAULT_MONTHS]
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [normalize_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 14) + 1)]
        if values[1:13] == months_norm and normalize_text(ws.cell(row_idx, 14).value) == "TOTAL":
            return row_idx
    return 6


def parse_excel_rows(path: Path) -> tuple[list[ExcelRow], list[str], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = detect_header_row(ws)
    month_names = [
        str(ws.cell(header_row, col).value).strip()
        for col in range(2, 14)
    ]
    rows: list[ExcelRow] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        descricao_raw = ws.cell(row_idx, 1).value
        if descricao_raw is None:
            continue
        descricao = str(descricao_raw).strip()
        if not descricao:
            continue
        # Ignore obvious report headers/footers.
        if normalize_text(descricao) in {
            "PAGINA",
            "DATA",
            "HORA",
            "EXERCICIO 2025",
            "CONTABILIDADE",
        }:
            continue

        meses = [parse_decimal_br(ws.cell(row_idx, col).value) for col in range(2, 14)]
        total = parse_decimal_br(ws.cell(row_idx, 14).value)
        rows.append(
            ExcelRow(
                row_number=row_idx,
                descricao=descricao,
                descricao_norm=normalize_text(descricao),
                meses=meses,
                total=total,
            )
        )
    return rows, month_names, ws.title


def parse_j100_blocks(path: Path) -> tuple[list[list[J100Record]], str]:
    layout_fields = load_layout_fields("J100.json")
    index = build_field_index(layout_fields)
    text, encoding = read_text_with_fallback(path)
    blocks: list[list[J100Record]] = []
    current: list[J100Record] = []
    line_number = 0

    for raw_line in text.splitlines():
        line_number += 1
        if not raw_line.startswith("|J100|"):
            continue
        parts = split_sped_line(raw_line)
        nivel_text = get_field(parts, index["NIVEL_AGL"])
        cod_agl = get_field(parts, index["COD_AGL"]).strip()
        try:
            nivel_agl = int(nivel_text or "0")
        except ValueError:
            continue
        if nivel_agl > 5:
            continue

        if cod_agl == "BAL_1" and current:
            blocks.append(current)
            current = []

        rec = J100Record(
            block_index=len(blocks) + 1,
            line_number=line_number,
            cod_agl=cod_agl,
            ind_cod_agl=get_field(parts, index["IND_COD_AGL"]).strip(),
            nivel_agl=nivel_agl,
            cod_agl_sup=get_field(parts, index["COD_AGL_SUP"]).strip(),
            ind_grp_bal=get_field(parts, index["IND_GRP_BAL"]).strip(),
            descr_cod_agl=get_field(parts, index["DESCR_COD_AGL"]).strip(),
            vl_cta_ini=parse_decimal_br(get_field(parts, index["VL_CTA_INI"])) or Decimal("0"),
            ind_dc_cta_ini=get_field(parts, index["IND_DC_CTA_INI"]).strip(),
            vl_cta_fin=parse_decimal_br(get_field(parts, index["VLR_CTA_FIN"])) or Decimal("0"),
            ind_dc_cta_fin=get_field(parts, index["IND_DC_CTA_FIN"]).strip(),
            nota_exp_ref=get_field(parts, index["NOTA_EXP_REF"]).strip(),
        )
        current.append(rec)

    if current:
        blocks.append(current)
    return blocks, encoding


def detect_balance_header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        col3 = normalize_text(ws.cell(row_idx, 3).value)
        col4 = normalize_text(ws.cell(row_idx, 4).value)
        if col3.endswith("2025") and col4.endswith("2024"):
            return row_idx
    return 6


def parse_balance_rows(path: Path) -> tuple[list[BalanceRow], str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = detect_balance_header_row(ws)
    rows: list[BalanceRow] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        descricao_raw = ws.cell(row_idx, 1).value
        if descricao_raw is None:
            continue
        descricao = str(descricao_raw).strip()
        if not descricao:
            continue
        if normalize_text(descricao) in {
            "PAGINA",
            "DATA",
            "HORA",
            "CONSOLIDACAO EMPRESA",
            "CONTABILIDADE",
            "BALANCO PATRIMONIAL",
        }:
            continue
        valor_final = parse_decimal_br(ws.cell(row_idx, 3).value)
        valor_inicial = parse_decimal_br(ws.cell(row_idx, 4).value)
        rows.append(
            BalanceRow(
                row_number=row_idx,
                descricao=descricao,
                descricao_norm=normalize_text(descricao),
                valor_final=valor_final,
                valor_inicial=valor_inicial,
            )
        )
    return rows, ws.title, ws.cell(header_row, 3).value or "12/2025"


def compare_balance(
    ecd_blocks: list[list[J100Record]],
    balance_rows: list[BalanceRow],
    tolerance: Decimal,
) -> BalanceReport:
    if not ecd_blocks:
        raise RuntimeError("Nao foram encontrados blocos J100 validos na ECD.")
    if not balance_rows:
        raise RuntimeError("Nao foram encontradas linhas de balanco no Excel.")

    ini_block = ecd_blocks[0]
    fin_block = ecd_blocks[-1]
    mapping_ini, unmatched_ini_idx, unmatched_excel_ini = align_by_description(
        ini_block,
        balance_rows,
        lambda rec: rec.descr_cod_agl,
        lambda row: row.descricao,
    )
    mapping_fin, unmatched_fin_idx, unmatched_excel_fin = align_by_description(
        fin_block,
        balance_rows,
        lambda rec: rec.descr_cod_agl,
        lambda row: row.descricao,
    )

    reverse_mapping_ini = {excel_idx: ecd_idx for ecd_idx, excel_idx in mapping_ini.items()}
    reverse_mapping_fin = {excel_idx: ecd_idx for ecd_idx, excel_idx in mapping_fin.items()}

    matches: list[BalanceMatch] = []
    unmatched_ecd_rows: dict[tuple[object, ...], dict[str, object]] = {}
    unmatched_excel_rows: dict[tuple[object, ...], dict[str, object]] = {}
    summary = {
        "blocks": len(ecd_blocks),
        "excel_rows": len(balance_rows),
        "matched_pairs": 0,
        "ok": 0,
        "divergent": 0,
        "unmatched_ecd": 0,
        "unmatched_excel": 0,
    }

    for idx, row in enumerate(balance_rows):
        ecd_ini_idx = reverse_mapping_ini.get(idx)
        ecd_fin_idx = reverse_mapping_fin.get(idx)
        ecd_ini = ini_block[ecd_ini_idx] if ecd_ini_idx is not None and ecd_ini_idx < len(ini_block) else None
        ecd_fin = fin_block[ecd_fin_idx] if ecd_fin_idx is not None and ecd_fin_idx < len(fin_block) else None
        valor_ecd_ini = comparison_value(ecd_ini.vl_cta_ini) if ecd_ini else None
        valor_ecd_fin = comparison_value(ecd_fin.vl_cta_fin) if ecd_fin else None
        valor_excel_ini = comparison_value(row.valor_inicial)
        valor_excel_fin = comparison_value(row.valor_final)
        diff_ini = decimal_diff(valor_ecd_ini, valor_excel_ini)
        diff_fin = decimal_diff(valor_ecd_fin, valor_excel_fin)
        ok_ini = within_tolerance(diff_ini, tolerance)
        ok_fin = within_tolerance(diff_fin, tolerance)
        if ecd_ini is not None and ecd_fin is not None and ok_ini and ok_fin:
            status = "OK"
            summary["ok"] += 1
        elif ecd_ini is not None or ecd_fin is not None:
            status = "DIVERGENTE"
            summary["divergent"] += 1
        else:
            status = "SEM_MAPEAMENTO"
            summary["divergent"] += 1

        if ecd_ini is not None or ecd_fin is not None:
            summary["matched_pairs"] += 1

        matches.append(
            BalanceMatch(
                row_number=row.row_number,
                excel=row,
                ecd_inicial=ecd_ini,
                ecd_final=ecd_fin,
                valor_ecd_inicial=valor_ecd_ini,
                valor_ecd_final=valor_ecd_fin,
                valor_excel_inicial=valor_excel_ini,
                valor_excel_final=valor_excel_fin,
                diff_inicial=diff_ini,
                diff_final=diff_fin,
                status=status,
            )
        )

    for ecd_idx in unmatched_ini_idx:
        ecd = ini_block[ecd_idx]
        key = (
            ecd.line_number,
            ecd.cod_agl,
            ecd.descr_cod_agl,
            format_decimal_br(ecd.vl_cta_ini),
            "INI",
        )
        entry = unmatched_ecd_rows.setdefault(
            key,
            {
                "line_number": ecd.line_number,
                "cod_agl": ecd.cod_agl,
                "descricao": ecd.descr_cod_agl,
                "valor": ecd.vl_cta_ini,
                "fase": "Inicial",
                "meses": set(),
            },
        )
        entry["meses"].add("Inicial")

    for ecd_idx in unmatched_fin_idx:
        ecd = fin_block[ecd_idx]
        key = (
            ecd.line_number,
            ecd.cod_agl,
            ecd.descr_cod_agl,
            format_decimal_br(ecd.vl_cta_fin),
            "FIN",
        )
        entry = unmatched_ecd_rows.setdefault(
            key,
            {
                "line_number": ecd.line_number,
                "cod_agl": ecd.cod_agl,
                "descricao": ecd.descr_cod_agl,
                "valor": ecd.vl_cta_fin,
                "fase": "Final",
                "meses": set(),
            },
        )
        entry["meses"].add("Final")

    for excel_idx in set(unmatched_excel_ini) | set(unmatched_excel_fin):
        row = balance_rows[excel_idx]
        key = (row.row_number, row.descricao)
        entry = unmatched_excel_rows.setdefault(
            key,
            {
                "row_number": row.row_number,
                "descricao": row.descricao,
                "meses": set(),
            },
        )
        entry["meses"].add("Inicial/Final")

    summary["unmatched_ecd"] = len(unmatched_ecd_rows)
    summary["unmatched_excel"] = len(unmatched_excel_rows)
    return BalanceReport(
        matches=matches,
        unmatched_ecd_rows=list(unmatched_ecd_rows.values()),
        unmatched_excel_rows=list(unmatched_excel_rows.values()),
        summary=summary,
    )


def align_sequences(ecd_records: list[J150Record], excel_rows: list[ExcelRow]) -> tuple[dict[int, int], list[int], list[int]]:
    return align_by_description(
        ecd_records,
        excel_rows,
        lambda rec: rec.descr_cod_agl,
        lambda row: row.descricao,
    )


def build_matches(
    ecd_blocks: list[list[J150Record]],
    excel_rows: list[ExcelRow],
    month_names: list[str],
    tolerance: Decimal,
) -> tuple[list[MatchResult], list[dict[str, object]], list[dict[str, object]], dict[str, int]]:
    all_matches: list[MatchResult] = []
    unmatched_ecd_rows: dict[tuple[object, ...], dict[str, object]] = {}
    unmatched_excel_rows: dict[tuple[object, ...], dict[str, object]] = {}
    summary = {
        "blocks": len(ecd_blocks),
        "excel_rows": len(excel_rows),
        "matched_pairs": 0,
        "month_ok": 0,
        "total_ok": 0,
        "divergent": 0,
        "unmatched_ecd": 0,
        "unmatched_excel": 0,
    }

    for idx, block in enumerate(ecd_blocks):
        mes = month_names[idx] if idx < len(month_names) else f"Mes_{idx + 1}"
        mapping, unmatched_ecd_idx, unmatched_excel_idx = align_sequences(block, excel_rows)

        for ecd_idx, excel_idx in mapping.items():
            ecd = block[ecd_idx]
            excel = excel_rows[excel_idx]
            valor_ecd_base = comparison_value(ecd.vl_cta_fin)
            valor_mes_excel = excel.meses[idx] if idx < len(excel.meses) else None
            valor_total_excel = excel.total
            valor_mes_excel_base = comparison_value(valor_mes_excel)
            valor_total_excel_base = comparison_value(valor_total_excel)
            diff_mes = decimal_diff(valor_ecd_base, valor_mes_excel_base)
            diff_total = decimal_diff(valor_ecd_base, valor_total_excel_base)
            ok_mes = within_tolerance(diff_mes, tolerance)
            ok_total = within_tolerance(diff_total, tolerance)
            if ok_mes:
                status = "OK"
                summary["month_ok"] += 1
            else:
                status = "DIVERGENTE"
                summary["divergent"] += 1
            if ok_total:
                summary["total_ok"] += 1
            summary["matched_pairs"] += 1
            all_matches.append(
                MatchResult(
                    mes=mes,
                    ecd=ecd,
                excel=excel,
                valor_mes_excel=valor_mes_excel,
                valor_total_excel=valor_total_excel,
                diff_mes=diff_mes,
                diff_total=diff_total,
                status=status,
            )
        )

        for ecd_idx in unmatched_ecd_idx:
            ecd = block[ecd_idx]
            key = (
                ecd.line_number,
                ecd.nu_ordem,
                ecd.nivel_agl,
                ecd.cod_agl,
                ecd.descr_cod_agl,
                format_decimal_br(ecd.vl_cta_fin),
            )
            entry = unmatched_ecd_rows.setdefault(
                key,
                {
                    "line_number": ecd.line_number,
                    "nu_ordem": ecd.nu_ordem,
                    "nivel_agl": ecd.nivel_agl,
                    "cod_agl": ecd.cod_agl,
                    "descricao": ecd.descr_cod_agl,
                    "valor": ecd.vl_cta_fin,
                    "meses": set(),
                },
            )
            entry["meses"].add(mes)

        for excel_idx in unmatched_excel_idx:
            excel = excel_rows[excel_idx]
            key = (excel.row_number, excel.descricao)
            entry = unmatched_excel_rows.setdefault(
                key,
                {
                    "row_number": excel.row_number,
                    "descricao": excel.descricao,
                    "meses": set(),
                },
            )
            entry["meses"].add(mes)

    summary["unmatched_ecd"] = len(unmatched_ecd_rows)
    summary["unmatched_excel"] = len(unmatched_excel_rows)
    return all_matches, list(unmatched_ecd_rows.values()), list(unmatched_excel_rows.values()), summary


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 55)


def write_balance_only_report(
    output_path: Path,
    ecd_path: Path,
    balanco_path: Path,
    balance_report: BalanceReport,
    balance_sheet_name: str | None = None,
) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_bal = wb.create_sheet("Comparacao_Balanco")
    ws_bal_ecd = wb.create_sheet("Nao_Mapeadas_Balanco_ECD")
    ws_bal_xlsx = wb.create_sheet("Nao_Mapeadas_Balanco_XLSX")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    bal_ok_fill = PatternFill("solid", fgColor="E2F0D9")
    bal_bad_fill = PatternFill("solid", fgColor="F4CCCC")
    bal_warn_fill = PatternFill("solid", fgColor="FFF2CC")

    def style_header(ws) -> None:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_summary.append(["Campo", "Valor"])
    ws_summary.append(["Arquivo ECD", str(ecd_path)])
    ws_summary.append(["Arquivo Excel DRE", "(nao informado)"])
    ws_summary.append(["Arquivo Balanco", str(balanco_path)])
    ws_summary.append(["Aba Balanco", balance_sheet_name or "(nao informado)"])
    ws_summary.append(["Blocos J100", balance_report.summary["blocks"]])
    ws_summary.append(["Linhas Excel Balanco", balance_report.summary["excel_rows"]])
    ws_summary.append(["Pares mapeados Balanco", balance_report.summary["matched_pairs"]])
    ws_summary.append(["OK Balanco", balance_report.summary["ok"]])
    ws_summary.append(["Divergencias Balanco", balance_report.summary["divergent"]])
    ws_summary.append(["Nao mapeadas Balanco ECD", balance_report.summary["unmatched_ecd"]])
    ws_summary.append(["Nao mapeadas Balanco Excel", balance_report.summary["unmatched_excel"]])
    style_header(ws_summary)
    ws_summary.add_table(
        Table(
            displayName="tblResumoInfoBalanco",
            ref=f"A1:B{ws_summary.max_row}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    ws_summary.freeze_panes = "A2"
    autosize_columns(ws_summary)

    ws_bal.append(
        [
            "Linha_Excel",
            "Descricao_Excel",
            "Valor_Excel_12_2025",
            "Valor_Excel_12_2024",
            "Linha_ECD_Inicial",
            "Descricao_ECD_Inicial",
            "Valor_ECD_Inicial",
            "Linha_ECD_Final",
            "Descricao_ECD_Final",
            "Valor_ECD_Final",
            "Dif_Inicial",
            "Dif_Final",
            "Status",
        ]
    )
    for match in balance_report.matches:
        ws_bal.append(
            [
                match.row_number,
                match.excel.descricao,
                None if match.valor_excel_final is None else float(match.valor_excel_final),
                None if match.valor_excel_inicial is None else float(match.valor_excel_inicial),
                None if match.ecd_inicial is None else match.ecd_inicial.line_number,
                None if match.ecd_inicial is None else match.ecd_inicial.descr_cod_agl,
                None if match.valor_ecd_inicial is None else float(match.valor_ecd_inicial),
                None if match.ecd_final is None else match.ecd_final.line_number,
                None if match.ecd_final is None else match.ecd_final.descr_cod_agl,
                None if match.valor_ecd_final is None else float(match.valor_ecd_final),
                None if match.diff_inicial is None else float(match.diff_inicial),
                None if match.diff_final is None else float(match.diff_final),
                match.status,
            ]
        )
    style_header(ws_bal)
    ws_bal.freeze_panes = "A2"
    ws_bal.auto_filter.ref = ws_bal.dimensions
    for col in ("C", "D", "G", "J", "K", "L"):
        for cell in ws_bal[col][1:]:
            cell.number_format = "#,##0.00"
    ws_bal.conditional_formatting.add(
        f"A2:M{ws_bal.max_row}",
        FormulaRule(formula=['$M2="OK"'], fill=bal_ok_fill),
    )
    ws_bal.conditional_formatting.add(
        f"A2:M{ws_bal.max_row}",
        FormulaRule(formula=['$M2="DIVERGENTE"'], fill=bal_bad_fill),
    )
    ws_bal.conditional_formatting.add(
        f"A2:M{ws_bal.max_row}",
        FormulaRule(formula=['$M2="SEM_MAPEAMENTO"'], fill=bal_warn_fill),
    )
    ws_bal.add_table(
        Table(
            displayName="tblComparacaoBalancoOnly",
            ref=f"A1:M{ws_bal.max_row}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    autosize_columns(ws_bal)

    ws_bal_ecd.append(["Fase", "Linha_ECD", "COD_AGL", "Descricao_ECD", "Valor_ECD"])
    for row in balance_report.unmatched_ecd_rows:
        ws_bal_ecd.append(
            [
                row["fase"],
                row["line_number"],
                row["cod_agl"],
                row["descricao"],
                float(row["valor"]),
            ]
        )
    style_header(ws_bal_ecd)
    ws_bal_ecd.freeze_panes = "A2"
    ws_bal_ecd.auto_filter.ref = ws_bal_ecd.dimensions
    for cell in ws_bal_ecd["E"][1:]:
        cell.number_format = "#,##0.00"
    if ws_bal_ecd.max_row > 1:
        ws_bal_ecd.add_table(
            Table(
                displayName="tblNaoMapeadasBalancoECDOnly",
                ref=f"A1:E{ws_bal_ecd.max_row}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
        )
    autosize_columns(ws_bal_ecd)

    ws_bal_xlsx.append(["Linha_Excel", "Descricao_Excel"])
    for row in balance_report.unmatched_excel_rows:
        ws_bal_xlsx.append([row["row_number"], row["descricao"]])
    style_header(ws_bal_xlsx)
    ws_bal_xlsx.freeze_panes = "A2"
    ws_bal_xlsx.auto_filter.ref = ws_bal_xlsx.dimensions
    if ws_bal_xlsx.max_row > 1:
        ws_bal_xlsx.add_table(
            Table(
                displayName="tblNaoMapeadasBalancoXLSXOnly",
                ref=f"A1:B{ws_bal_xlsx.max_row}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
        )
    autosize_columns(ws_bal_xlsx)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_report(
    output_path: Path,
    matches: list[MatchResult],
    unmatched_ecd_rows: list[dict[str, object]],
    unmatched_excel_rows: list[dict[str, object]],
    summary: dict[str, int],
    ecd_path: Path,
    excel_path: Path,
    month_names: list[str],
    excel_sheet_name: str,
    balanco_path: Path | None = None,
    balance_report: BalanceReport | None = None,
    balance_sheet_name: str | None = None,
) -> None:
    if excel_path is None:
        if balanco_path is None or balance_report is None:
            raise RuntimeError("Nao ha dados suficientes para gerar o relatorio.")
        write_balance_only_report(
            output_path=output_path,
            ecd_path=ecd_path,
            balanco_path=balanco_path,
            balance_report=balance_report,
            balance_sheet_name=balance_sheet_name,
        )
        return

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_match = wb.create_sheet("Comparacao")
    ws_ecd = wb.create_sheet("Nao_Mapeadas_ECD")
    ws_xlsx = wb.create_sheet("Nao_Mapeadas_XLSX")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    bad_fill = PatternFill("solid", fgColor="F4CCCC")

    def style_header(ws) -> None:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    month_order = {month: idx for idx, month in enumerate(month_names)}
    sort_months = lambda months: ", ".join(sorted(months, key=lambda m: month_order.get(m, 999)))  # noqa: E731

    # Summary sheet
    ws_summary.append(["Campo", "Valor"])
    ws_summary.append(["Arquivo ECD", str(ecd_path)])
    ws_summary.append(["Arquivo Excel", str(excel_path)])
    ws_summary.append(["Aba Excel", excel_sheet_name])
    ws_summary.append(["Critério", "Comparação por valor absoluto"])
    ws_summary.append(["Blocos mensais ECD", summary["blocks"]])
    ws_summary.append(["Linhas Excel", summary["excel_rows"]])
    ws_summary.append(["Pares mapeados", summary["matched_pairs"]])
    ws_summary.append(["OK no mes", summary["month_ok"]])
    ws_summary.append(["OK no total", summary["total_ok"]])
    ws_summary.append(["Divergencias", summary["divergent"]])
    ws_summary.append(["Nao mapeadas ECD", summary["unmatched_ecd"]])
    ws_summary.append(["Nao mapeadas Excel", summary["unmatched_excel"]])
    ws_summary.append([])
    ws_summary.append(["Mes", "Quantidade de comparacoes"])
    per_month: dict[str, int] = {}
    for match in matches:
        per_month[match.mes] = per_month.get(match.mes, 0) + 1
    for month in month_names:
        ws_summary.append([month, per_month.get(month, 0)])
    style_header(ws_summary)
    for cell in ws_summary[13]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.add_table(
        Table(
            displayName="tblResumoInfo",
            ref=f"A1:B13",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    ws_summary.add_table(
        Table(
            displayName="tblResumoMeses",
            ref=f"A15:B27",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    ws_summary.freeze_panes = "A2"
    autosize_columns(ws_summary)

    # Comparisons sheet
    ws_match.append(
        [
            "Mes",
            "Linha_ECD",
            "NU_ORDEM",
            "Nivel",
            "COD_AGL",
            "COD_AGL_SUP",
            "Descricao_ECD",
            "Valor_ECD",
            "IND_DC_FIN",
            "IND_GRP_DRE",
            "Linha_Excel",
            "Descricao_Excel",
            "Valor_Excel_Mes",
            "Valor_Excel_Total",
            "Dif_Mes",
            "Dif_Total",
            "Status",
        ]
    )
    for match in matches:
        ws_match.append(
            [
                match.mes,
                match.ecd.line_number,
                match.ecd.nu_ordem,
                match.ecd.nivel_agl,
                match.ecd.cod_agl,
                match.ecd.cod_agl_sup,
                match.ecd.descr_cod_agl,
                float(match.ecd.vl_cta_fin),
                match.ecd.ind_dc_cta_fin,
                match.ecd.ind_grp_dre,
                match.excel.row_number,
                match.excel.descricao,
                None if match.valor_mes_excel is None else float(match.valor_mes_excel),
                None if match.valor_total_excel is None else float(match.valor_total_excel),
                None if match.diff_mes is None else float(match.diff_mes),
                None if match.diff_total is None else float(match.diff_total),
                match.status,
            ]
        )

    style_header(ws_match)
    ws_match.freeze_panes = "A2"
    ws_match.auto_filter.ref = ws_match.dimensions
    for col in ("H", "M", "N", "O", "P"):
        for cell in ws_match[col][1:]:
            cell.number_format = '#,##0.00'
    ws_match.conditional_formatting.add(
        f"A2:Q{ws_match.max_row}",
        FormulaRule(
            formula=['$Q2="OK"'],
            fill=ok_fill,
        ),
    )
    ws_match.conditional_formatting.add(
        f"A2:Q{ws_match.max_row}",
        FormulaRule(
            formula=['$Q2="DIVERGENTE"'],
            fill=bad_fill,
        ),
    )
    ws_match.add_table(
        Table(
            displayName="tblComparacao",
            ref=f"A1:Q{ws_match.max_row}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    autosize_columns(ws_match)

    # Unmatched ECD
    ws_ecd.append(["Meses", "Linha_ECD", "NU_ORDEM", "Nivel", "COD_AGL", "Descricao_ECD", "Valor_ECD"])
    for row in unmatched_ecd_rows:
        ws_ecd.append(
            [
                sort_months(row["meses"]),
                row["line_number"],
                row["nu_ordem"],
                row["nivel_agl"],
                row["cod_agl"],
                row["descricao"],
                float(row["valor"]),
            ]
        )
    style_header(ws_ecd)
    ws_ecd.freeze_panes = "A2"
    ws_ecd.auto_filter.ref = ws_ecd.dimensions
    for cell in ws_ecd["G"][1:]:
        cell.number_format = '#,##0.00'
    ws_ecd.add_table(
        Table(
            displayName="tblNaoMapeadasECD",
            ref=f"A1:G{ws_ecd.max_row}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    autosize_columns(ws_ecd)

    # Unmatched Excel
    ws_xlsx.append(["Meses", "Linha_Excel", "Descricao_Excel"])
    for row in unmatched_excel_rows:
        ws_xlsx.append([sort_months(row["meses"]), row["row_number"], row["descricao"]])
    style_header(ws_xlsx)
    ws_xlsx.freeze_panes = "A2"
    ws_xlsx.add_table(
        Table(
            displayName="tblNaoMapeadasXLSX",
            ref=f"A1:C{ws_xlsx.max_row}",
            tableStyleInfo=TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            ),
        )
    )
    autosize_columns(ws_xlsx)

    if balance_report is not None:
        ws_bal = wb.create_sheet("Comparacao_Balanco")
        ws_bal_ecd = wb.create_sheet("Nao_Mapeadas_Balanco_ECD")
        ws_bal_xlsx = wb.create_sheet("Nao_Mapeadas_Balanco_XLSX")

        bal_ok_fill = PatternFill("solid", fgColor="E2F0D9")
        bal_bad_fill = PatternFill("solid", fgColor="F4CCCC")
        bal_warn_fill = PatternFill("solid", fgColor="FFF2CC")

        ws_bal.append(
            [
                "Linha_Excel",
                "Descricao_Excel",
                "Valor_Excel_12_2025",
                "Valor_Excel_12_2024",
                "Linha_ECD_Inicial",
                "Descricao_ECD_Inicial",
                "Valor_ECD_Inicial",
                "Linha_ECD_Final",
                "Descricao_ECD_Final",
                "Valor_ECD_Final",
                "Dif_Inicial",
                "Dif_Final",
                "Status",
            ]
        )
        for match in balance_report.matches:
            ws_bal.append(
                [
                    match.row_number,
                    match.excel.descricao,
                    None if match.valor_excel_final is None else float(match.valor_excel_final),
                    None if match.valor_excel_inicial is None else float(match.valor_excel_inicial),
                    None if match.ecd_inicial is None else match.ecd_inicial.line_number,
                    None if match.ecd_inicial is None else match.ecd_inicial.descr_cod_agl,
                    None if match.valor_ecd_inicial is None else float(match.valor_ecd_inicial),
                    None if match.ecd_final is None else match.ecd_final.line_number,
                    None if match.ecd_final is None else match.ecd_final.descr_cod_agl,
                    None if match.valor_ecd_final is None else float(match.valor_ecd_final),
                    None if match.diff_inicial is None else float(match.diff_inicial),
                    None if match.diff_final is None else float(match.diff_final),
                    match.status,
                ]
            )
        style_header(ws_bal)
        ws_bal.freeze_panes = "A2"
        ws_bal.auto_filter.ref = ws_bal.dimensions
        for col in ("C", "D", "G", "J", "K", "L"):
            for cell in ws_bal[col][1:]:
                cell.number_format = "#,##0.00"
        ws_bal.conditional_formatting.add(
            f"A2:M{ws_bal.max_row}",
            FormulaRule(formula=['$M2="OK"'], fill=bal_ok_fill),
        )
        ws_bal.conditional_formatting.add(
            f"A2:M{ws_bal.max_row}",
            FormulaRule(formula=['$M2="DIVERGENTE"'], fill=bal_bad_fill),
        )
        ws_bal.conditional_formatting.add(
            f"A2:M{ws_bal.max_row}",
            FormulaRule(formula=['$M2="SEM_MAPEAMENTO"'], fill=bal_warn_fill),
        )
        ws_bal.add_table(
            Table(
                displayName="tblComparacaoBalanco",
                ref=f"A1:M{ws_bal.max_row}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
        )
        autosize_columns(ws_bal)

        ws_bal_ecd.append(["Fase", "Linha_ECD", "COD_AGL", "Descricao_ECD", "Valor_ECD"])
        for row in balance_report.unmatched_ecd_rows:
            ws_bal_ecd.append(
                [
                    row["fase"],
                    row["line_number"],
                    row["cod_agl"],
                    row["descricao"],
                    float(row["valor"]),
                ]
            )
        style_header(ws_bal_ecd)
        ws_bal_ecd.freeze_panes = "A2"
        ws_bal_ecd.auto_filter.ref = ws_bal_ecd.dimensions
        for cell in ws_bal_ecd["E"][1:]:
            cell.number_format = "#,##0.00"
        ws_bal_ecd.add_table(
            Table(
                displayName="tblNaoMapeadasBalancoECD",
                ref=f"A1:E{ws_bal_ecd.max_row}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
        )
        autosize_columns(ws_bal_ecd)

        ws_bal_xlsx.append(["Linha_Excel", "Descricao_Excel"])
        for row in balance_report.unmatched_excel_rows:
            ws_bal_xlsx.append([row["row_number"], row["descricao"]])
        style_header(ws_bal_xlsx)
        ws_bal_xlsx.freeze_panes = "A2"
        ws_bal_xlsx.auto_filter.ref = ws_bal_xlsx.dimensions
        ws_bal_xlsx.add_table(
            Table(
                displayName="tblNaoMapeadasBalancoXLSX",
                ref=f"A1:B{ws_bal_xlsx.max_row}",
                tableStyleInfo=TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                ),
            )
        )
        autosize_columns(ws_bal_xlsx)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def resolve_inputs(args: argparse.Namespace) -> tuple[Path, Path | None, Path | None, Path]:
    ecd_path = Path(args.ecd) if args.ecd else None
    excel_path = Path(args.excel) if args.excel else None
    balanco_path = Path(args.balanco) if getattr(args, "balanco", None) else None
    output_path = Path(args.output) if args.output else None

    if not args.no_gui:
        if ecd_path is None:
            ecd_path = choose_file_dialog(
                "Selecione o TXT da ECD",
                [("Arquivos TXT", "*.txt"), ("Todos os arquivos", "*.*")],
            )
        if excel_path is None:
            excel_path = choose_file_dialog(
                "Selecione o Excel da DRE mensal (opcional)",
                [("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
            )
        if balanco_path is None:
            balanco_path = choose_file_dialog(
                "Selecione o Excel do balanco patrimonial (opcional)",
                [("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")],
                initialdir=str(excel_path.parent) if excel_path else str(ecd_path.parent) if ecd_path else None,
            )
        if output_path is None:
            output_path = choose_save_dialog(
                "Salvar relatorio",
                ".xlsx",
                [("Arquivos Excel", "*.xlsx")],
                "comparacao_j150_ecd_dre.xlsx",
            )

    if ecd_path is None:
        raise SystemExit("E necessario informar o TXT da ECD.")
    if excel_path is None and balanco_path is None:
        raise SystemExit("E necessario informar ao menos o XLSX da DRE ou o XLSX do balanco.")

    if output_path is None:
        output_base = excel_path or balanco_path or ecd_path
        output_path = output_base.with_name("comparacao_j150_ecd_dre.xlsx")

    return ecd_path, excel_path, balanco_path, output_path


def main() -> int:
    args = parse_args()
    try:
        ecd_path, excel_path, balanco_path, output_path = resolve_inputs(args)
        ecd_encoding: str | None = None
        matches: list[MatchResult] = []
        unmatched_ecd_rows: list[dict[str, object]] = []
        unmatched_excel_rows: list[dict[str, object]] = []
        summary = {
            "blocks": 0,
            "excel_rows": 0,
            "matched_pairs": 0,
            "month_ok": 0,
            "total_ok": 0,
            "divergent": 0,
            "unmatched_ecd": 0,
            "unmatched_excel": 0,
        }
        month_names: list[str] = []
        sheet_name: str | None = None
        balance_report: BalanceReport | None = None
        balance_sheet_name: str | None = None
        if excel_path is not None:
            ecd_blocks, ecd_encoding = parse_j150_blocks(ecd_path)
            excel_rows, month_names, sheet_name = parse_excel_rows(excel_path)
            if not ecd_blocks:
                raise RuntimeError("Nao foram encontrados blocos J150 validos na ECD.")
            if not excel_rows:
                raise RuntimeError("Nao foram encontradas linhas de DRE no Excel.")
            matches, unmatched_ecd_rows, unmatched_excel_rows, summary = build_matches(
                ecd_blocks=ecd_blocks,
                excel_rows=excel_rows,
                month_names=month_names,
                tolerance=DEFAULT_TOLERANCE,
            )
        if balanco_path is not None:
            balance_rows, balance_sheet_name, _header = parse_balance_rows(balanco_path)
            balance_blocks, balance_encoding = parse_j100_blocks(ecd_path)
            if ecd_encoding is None:
                ecd_encoding = balance_encoding
            balance_report = compare_balance(
                ecd_blocks=balance_blocks,
                balance_rows=balance_rows,
                tolerance=DEFAULT_TOLERANCE,
            )
        if ecd_encoding is None:
            ecd_encoding = "desconhecida"

        write_report(
            output_path=output_path,
            matches=matches,
            unmatched_ecd_rows=unmatched_ecd_rows,
            unmatched_excel_rows=unmatched_excel_rows,
            summary=summary,
            ecd_path=ecd_path,
            excel_path=excel_path,
            month_names=month_names,
            excel_sheet_name=sheet_name,
            balanco_path=balanco_path,
            balance_report=balance_report,
            balance_sheet_name=balance_sheet_name,
        )

        message = (
            f"Relatorio gerado com sucesso em:\n{output_path}\n\n"
            f"Codificacao da ECD: {ecd_encoding}\n"
            f"Blocos mensais encontrados: {summary['blocks']}\n"
            f"Pares mapeados: {summary['matched_pairs']}\n"
            f"Divergencias: {summary['divergent']}"
        )
        if balance_report is not None:
            message += (
                "\n\nBalanço:\n"
                f"Blocos J100 encontrados: {balance_report.summary['blocks']}\n"
                f"Pares mapeados: {balance_report.summary['matched_pairs']}\n"
                f"OK: {balance_report.summary['ok']}\n"
                f"Divergencias: {balance_report.summary['divergent']}"
            )
        if excel_path is None and balance_report is not None:
            message = (
                f"Relatorio gerado com sucesso em:\n{output_path}\n\n"
                f"Codificacao da ECD: {ecd_encoding}\n\n"
                "Balanco:\n"
                f"Blocos J100 encontrados: {balance_report.summary['blocks']}\n"
                f"Pares mapeados: {balance_report.summary['matched_pairs']}\n"
                f"OK: {balance_report.summary['ok']}\n"
                f"Divergencias: {balance_report.summary['divergent']}"
            )
        if not args.no_gui:
            root = make_root()
            messagebox.showinfo("Comparacao concluida", message, parent=root)
            root.destroy()
        else:
            print(message)
        return 0
    except Exception as exc:  # noqa: BLE001
        if not args.no_gui:
            root = make_root()
            messagebox.showerror("Erro", str(exc), parent=root)
            root.destroy()
        else:
            print(f"Erro: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
