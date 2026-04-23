from __future__ import annotations

import base64
import importlib.util
import io
import re
import sys
import tempfile
import unicodedata
import zipfile
from decimal import Decimal, InvalidOperation
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

try:
    import fitz  # PyMuPDF
except Exception:
    try:
        import pymupdf as fitz  # fallback
    except Exception:
        fitz = None

try:
    from PIL import Image
except Exception:
    Image = None

try:
    import pytesseract
except Exception:
    pytesseract = None


BASE_DIR = Path(__file__).resolve().parents[2]
TOLERANCIA_CONSISTENCIA = Decimal('0.01')
ORDEM_CAMPOS_EDITAVEIS = [
    'comissao_motorista',
    'dsr',
    'horas_extras',
    'diarias_refeicoes',
    'salario',
    'estadia_tempo_espera',
    'premio_disco_tacografo',
    'total_remuneracao_mensal',
]
CAMPOS_SOMA_REMUNERACAO = [
    'comissao_motorista',
    'dsr',
    'horas_extras',
    'diarias_refeicoes',
    'salario',
    'estadia_tempo_espera',
    'premio_disco_tacografo',
]

PADRAO_VALOR_BR_GERAL = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}')
TESSERACT_CANDIDATOS = [
    Path(r'C:\Program Files\Tesseract-OCR\tesseract.exe'),
    Path(r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe'),
    Path(r'C:\ProgramData\chocolatey\bin\tesseract.exe'),
    Path(r'C:\ProgramData\chocolatey\lib\tesseract\tools\tesseract.exe'),
]


def _normalizar_nome(valor: str) -> str:
    base = unicodedata.normalize('NFKD', str(valor or '').upper().strip())
    sem_acentos = ''.join(ch for ch in base if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', sem_acentos)


def _normalizar_cpf(valor: str) -> str:
    return re.sub(r'\D', '', valor or '')


def _formatar_codigo(valor: str) -> str:
    digitos = re.sub(r'\D', '', valor or '')
    return digitos.zfill(5)[-5:] if digitos else '00000'


def _formatar_cpf_visual(cpf: str) -> str:
    dig = _normalizar_cpf(cpf)
    if len(dig) != 11:
        return cpf or ''
    return f'{dig[:3]}.{dig[3:6]}.{dig[6:9]}-{dig[9:]}'


def _normalizar_linha_ocr(valor: str) -> str:
    base = unicodedata.normalize('NFKD', str(valor or ''))
    base = ''.join(ch for ch in base if not unicodedata.combining(ch))
    base = base.upper()
    return re.sub(r'\s+', ' ', base).strip()


def _decimal_de_br(valor: str) -> Decimal:
    bruto = str(valor or '').strip().replace('.', '').replace(',', '.')
    try:
        return Decimal(bruto)
    except InvalidOperation:
        return Decimal('0')


def _decimal_para_br(valor: Decimal) -> str:
    texto = f'{valor:,.2f}'
    return texto.replace(',', '#').replace('.', ',').replace('#', '.')


def _normalizar_valor_br_texto(valor: Any) -> str:
    bruto = str(valor or '').strip().replace(' ', '')
    if not bruto:
        return '0,00'
    if ',' in bruto:
        return _decimal_para_br(_decimal_de_br(bruto))
    try:
        return _decimal_para_br(Decimal(bruto))
    except Exception:
        return '0,00'


def _slug_nome_arquivo(valor: str) -> str:
    base = _normalizar_nome(valor or 'funcionario').replace(' ', '_')
    base = re.sub(r'[^A-Z0-9_]+', '', base)
    return (base or 'FUNCIONARIO')[:50]


def montar_chave_funcionario(nome: str, cpf: str) -> str:
    return f'{_normalizar_nome(nome)}|{_normalizar_cpf(cpf)}'


def localizar_script_bandeira() -> Path:
    pasta_python = BASE_DIR / 'python'
    pasta_robos = BASE_DIR / 'Robôs Leonardo'

    candidatos = sorted(
        pasta_python.glob('Cart*o Horas Bandeira Transportes/extrair_remuneracao_motorista_ultima_pagina.py')
    )
    if candidatos:
        return candidatos[0]

    candidatos = sorted(
        pasta_robos.glob('Cart*o Horas Bandeira Transportes/extrair_remuneracao_motorista_ultima_pagina.py')
    )
    if candidatos:
        return candidatos[0]

    candidatos = sorted(pasta_python.rglob('extrair_remuneracao_motorista_ultima_pagina.py'))
    for item in candidatos:
        if 'BANDEIRA' in _normalizar_nome(str(item.parent)):
            return item

    candidatos = sorted(pasta_robos.rglob('extrair_remuneracao_motorista_ultima_pagina.py'))
    for item in candidatos:
        if 'BANDEIRA' in _normalizar_nome(str(item.parent)):
            return item

    raise FileNotFoundError('Não foi possível localizar o script do Cartão Horas Bandeira Transportes.')


def localizar_lista_funcionarios() -> Path:
    pasta_python = BASE_DIR / 'python'
    pasta_robos = BASE_DIR / 'Robôs Leonardo'

    candidatos = sorted(
        pasta_python.glob('Cart*o Horas Bandeira Transportes/Lista funcion*rios.xlsx')
    )
    if candidatos:
        return candidatos[0]

    candidatos = sorted(
        pasta_robos.glob('Cart*o Horas Bandeira Transportes/Lista funcion*rios.xlsx')
    )
    if candidatos:
        return candidatos[0]

    candidatos = sorted(pasta_python.rglob('Lista*funcion*rios*.xlsx'))
    for item in candidatos:
        if 'BANDEIRA' in _normalizar_nome(str(item.parent)):
            return item

    candidatos = sorted(pasta_robos.rglob('Lista*funcion*rios*.xlsx'))
    for item in candidatos:
        if 'BANDEIRA' in _normalizar_nome(str(item.parent)):
            return item

    raise FileNotFoundError('Não foi possível localizar a planilha Lista funcionários.xlsx do Bandeira.')


SCRIPT_BANDEIRA = localizar_script_bandeira()
LISTA_FUNCIONARIOS = localizar_lista_funcionarios()


def carregar_modulo_bandeira():
    spec = importlib.util.spec_from_file_location('cartao_horas_bandeira_script', SCRIPT_BANDEIRA)
    if spec is None or spec.loader is None:
        raise RuntimeError(f'Não foi possível carregar o script: {SCRIPT_BANDEIRA}')
    modulo = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = modulo
    spec.loader.exec_module(modulo)
    return modulo


def _configurar_tesseract_cmd() -> None:
    if pytesseract is None:
        return

    cmd_atual = str(getattr(pytesseract.pytesseract, 'tesseract_cmd', '') or '').strip()
    if cmd_atual and Path(cmd_atual).exists():
        return

    # Regra operacional: não depender de PATH do Windows.
    for candidato in TESSERACT_CANDIDATOS:
        if candidato.exists():
            pytesseract.pytesseract.tesseract_cmd = str(candidato)
            return

    raise RuntimeError(
        'OCR indisponível: executável do Tesseract não localizado nos caminhos fixos esperados.'
    )


def _extrair_ocr_linhas_pdf(caminho_pdf: Path) -> list[str]:
    if fitz is None or Image is None or pytesseract is None:
        raise RuntimeError('OCR indisponível neste ambiente (fitz/pillow/pytesseract).')

    _configurar_tesseract_cmd()

    doc = fitz.open(str(caminho_pdf))
    try:
        if doc.page_count < 1:
            return []
        page = doc.load_page(0)
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
        texto = pytesseract.image_to_string(img, lang='por+eng')
    finally:
        doc.close()

    linhas = [re.sub(r'[ \t]+', ' ', linha).strip() for linha in texto.replace('\r', '').splitlines()]
    return [linha for linha in linhas if linha]


def _extrair_ultimo_valor_br(texto: str) -> str | None:
    encontrados = PADRAO_VALOR_BR_GERAL.findall(str(texto or ''))
    return encontrados[-1] if encontrados else None


def _buscar_valor_por_palavras(linhas: list[str], palavras: list[str]) -> str | None:
    palavras_norm = [_normalizar_linha_ocr(p) for p in palavras]
    for linha in linhas:
        linha_norm = _normalizar_linha_ocr(linha)
        if all(p in linha_norm for p in palavras_norm):
            valor = _extrair_ultimo_valor_br(linha_norm)
            if valor:
                return valor
    return None


def _resultado_extracao_obj(modulo: Any, payload: dict[str, Any]) -> Any:
    try:
        return modulo.ResultadoExtracao(**payload)
    except TypeError:
        payload2 = dict(payload)
        payload2.pop('origem_matricula', None)
        payload2.pop('matricula_foi_heuristica', None)
        payload2.pop('matricula_candidatas', None)
        try:
            return modulo.ResultadoExtracao(**payload2)
        except TypeError:
            return SimpleNamespace(**payload2)


def _extrair_resultado_por_ocr(caminho_pdf: Path, modulo: Any) -> Any:
    linhas = _extrair_ocr_linhas_pdf(caminho_pdf)
    if not linhas:
        raise ValueError('OCR sem conteúdo legível na página.')

    texto_compacto = '\n'.join(linhas)
    texto_norm = _normalizar_linha_ocr(texto_compacto)

    empresa = ''
    empregado = ''
    cpf = ''
    periodo_inicial = ''
    periodo_final = ''

    m_empregado = re.search(r'EMPREGADO\s*[:;]\s*(.+?)(?:\s+CONTA\s+CORRENTE|\s+CPF\s*[:;])', texto_norm)
    if m_empregado:
        empregado = m_empregado.group(1).strip()

    m_empresa = re.search(r'EMPRESA\s*[:;]\s*(.+?)(?:\s+EMPREGADO\s*[:;]|\s+RONDONOPOLIS|\s+MT)', texto_norm)
    if m_empresa:
        empresa = m_empresa.group(1).strip()
        if 'BANDEIRA' in texto_norm:
            empresa = f'BANDEIRA TRANSPORTES {empresa}'.strip()
    elif 'BANDEIRA' in texto_norm:
        empresa = 'BANDEIRA TRANSPORTES'

    m_cpf = re.search(r'CPF\s*[:;]\s*(\d{11})', texto_norm)
    if m_cpf:
        cpf = m_cpf.group(1)

    datas = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', texto_norm)
    if len(datas) >= 2:
        def _fmt(d: str) -> str:
            dia, mes, ano = d.split('/')
            return f'{int(dia):02d}/{int(mes):02d}/{ano}'

        periodo_inicial, periodo_final = _fmt(datas[0]), _fmt(datas[1])

    valores = {
        'comissao_motorista': _buscar_valor_por_palavras(linhas, ['Comissao', 'Motorista']) or '0,00',
        'dsr': _buscar_valor_por_palavras(linhas, ['Descanso', 'DSR']) or '0,00',
        'horas_extras': _buscar_valor_por_palavras(linhas, ['Horas', 'Extras']) or '0,00',
        'diarias_refeicoes': _buscar_valor_por_palavras(linhas, ['Diarias', 'Refei']) or '0,00',
        'salario': _buscar_valor_por_palavras(linhas, ['Salario']) or '0,00',
        'estadia_tempo_espera': _buscar_valor_por_palavras(linhas, ['Estadia']) or '0,00',
        'premio_disco_tacografo': _buscar_valor_por_palavras(linhas, ['Premio', 'Tac']) or '0,00',
        'total_remuneracao_mensal': _buscar_valor_por_palavras(linhas, ['Total', 'Remuneração', 'Motorista']) or '0,00',
    }

    payload = {
        'empresa': empresa,
        'periodo_inicial': periodo_inicial,
        'periodo_final': periodo_final,
        'empregado': empregado,
        'cpf': cpf,
        'matricula': '',
        'valores': valores,
        'origem_matricula': 'ocr',
        'matricula_foi_heuristica': False,
        'matricula_candidatas': [],
    }
    return _resultado_extracao_obj(modulo, payload)


def _montar_indices_lista_funcionarios() -> tuple[dict[str, str], dict[str, str]]:
    por_cpf: dict[str, str] = {}
    por_nome_cpf: dict[str, str] = {}

    if not LISTA_FUNCIONARIOS.exists():
        return por_cpf, por_nome_cpf

    workbook = load_workbook(LISTA_FUNCIONARIOS, data_only=True, read_only=True)
    try:
        for ws in workbook.worksheets:
            if ws.max_row < 2:
                continue

            col_cpf = ws.max_column if ws.max_column >= 1 else 1
            for row in range(2, ws.max_row + 1):
                codigo = _formatar_codigo(str(ws.cell(row, 1).value or ''))
                if codigo == '00000':
                    continue

                nome_col2 = str(ws.cell(row, 2).value or '').strip()
                nome_col3 = str(ws.cell(row, 3).value or '').strip() if ws.max_column >= 3 else ''
                cpf = _normalizar_cpf(str(ws.cell(row, col_cpf).value or ''))

                if len(cpf) == 11:
                    por_cpf[cpf] = codigo
                    if nome_col2:
                        por_nome_cpf[montar_chave_funcionario(nome_col2, cpf)] = codigo
                    if nome_col3:
                        por_nome_cpf[montar_chave_funcionario(nome_col3, cpf)] = codigo
    finally:
        workbook.close()

    return por_cpf, por_nome_cpf


def _resolver_matricula(
    nome: str,
    cpf: str,
    matricula_extraida: str,
    matricula_manual: str,
    indice_por_cpf: dict[str, str],
    indice_por_nome_cpf: dict[str, str],
) -> tuple[str, str]:
    if matricula_manual:
        return _formatar_codigo(matricula_manual), 'informada_usuario'

    codigo_pdf = _formatar_codigo(matricula_extraida)
    if codigo_pdf != '00000':
        return codigo_pdf, 'pdf'

    chave = montar_chave_funcionario(nome, cpf)
    codigo_lista = indice_por_nome_cpf.get(chave)
    if not codigo_lista:
        codigo_lista = indice_por_cpf.get(_normalizar_cpf(cpf), '')
    codigo_lista = _formatar_codigo(codigo_lista)
    if codigo_lista != '00000':
        return codigo_lista, 'lista_funcionarios'

    return '00000', 'nao_encontrada'


def _aplicar_overrides_valores(
    valores: dict[str, str],
    registro_id: str,
    chave: str,
    overrides_por_registro: dict[str, Any],
) -> dict[str, str]:
    resultado = dict(valores or {})
    if not isinstance(overrides_por_registro, dict):
        return resultado

    payload = overrides_por_registro.get(registro_id)
    if payload is None:
        payload = overrides_por_registro.get(chave)
    if not isinstance(payload, dict):
        return resultado

    for campo in ORDEM_CAMPOS_EDITAVEIS:
        if campo in payload:
            resultado[campo] = _normalizar_valor_br_texto(payload.get(campo))
    return resultado


def _garantir_campos_editaveis(valores: dict[str, Any]) -> dict[str, str]:
    out: dict[str, str] = {}
    for campo in ORDEM_CAMPOS_EDITAVEIS:
        out[campo] = _normalizar_valor_br_texto((valores or {}).get(campo, '0,00'))
    return out


def _avaliar_consistencia(valores: dict[str, str]) -> tuple[Decimal, Decimal, bool]:
    soma = Decimal('0')
    for campo in CAMPOS_SOMA_REMUNERACAO:
        soma += _decimal_de_br(valores.get(campo, '0,00'))

    total = _decimal_de_br(valores.get('total_remuneracao_mensal', '0,00'))
    return soma, total, abs(soma - total) <= TOLERANCIA_CONSISTENCIA


def _normalizar_chave_acao(valor: str) -> str:
    return str(valor or '').strip()


def processar_pdf_cartao_bandeira(
    pdf_bytes: bytes,
    nome_arquivo: str,
    ids_por_chave: dict[str, str] | None = None,
    confirmados: list[str] | set[str] | None = None,
    removidos: list[str] | set[str] | None = None,
    overrides_por_registro: dict[str, Any] | None = None,
) -> dict[str, Any]:
    modulo = carregar_modulo_bandeira()
    ids_por_chave = ids_por_chave or {}
    overrides_por_registro = overrides_por_registro or {}

    confirmados_set = {_normalizar_chave_acao(item) for item in (confirmados or []) if _normalizar_chave_acao(item)}
    removidos_set = {_normalizar_chave_acao(item) for item in (removidos or []) if _normalizar_chave_acao(item)}

    indice_por_cpf, indice_por_nome_cpf = _montar_indices_lista_funcionarios()

    sufixo = Path(nome_arquivo or 'arquivo.pdf').suffix or '.pdf'
    funcionarios: list[dict[str, Any]] = []
    pendencias: list[dict[str, Any]] = []
    linhas_preview: list[str] = []
    arquivos_txt_gerados: list[tuple[str, str]] = []
    assinaturas_paginas_vistas: set[str] = set()
    total_duplicados_ignorados = 0
    total_paginas_com_extracao = 0
    arquivo_base64 = ''
    arquivo_mime_type = ''
    tipo_saida = 'zip'

    with tempfile.TemporaryDirectory(prefix='cartao_bandeira_') as tmp_dir:
        caminho_pdf = Path(tmp_dir) / f'entrada{sufixo}'
        caminho_pdf.write_bytes(pdf_bytes)

        reader = PdfReader(str(caminho_pdf))
        total_paginas = len(reader.pages)
        if total_paginas == 0:
            raise ValueError('PDF sem páginas.')

        for idx_pagina, pagina in enumerate(reader.pages, start=1):
            writer = PdfWriter()
            writer.add_page(pagina)
            caminho_pagina = Path(tmp_dir) / f'pagina_{idx_pagina:04d}.pdf'
            with caminho_pagina.open('wb') as fp:
                writer.write(fp)

            resultado = None
            erro_ocr = None
            erro_texto = None

            # Fluxo principal: OCR. Fallback textual para PDFs que vierem com camada de texto.
            try:
                resultado = _extrair_resultado_por_ocr(caminho_pagina, modulo)
            except Exception as exc_ocr:
                erro_ocr = str(exc_ocr)
                try:
                    resultado = modulo.extrair_dados_ultima_pagina(caminho_pagina, matricula_forcada='')
                except Exception as exc_texto:
                    erro_texto = str(exc_texto)

            if resultado is None:
                chave_erro = f'PAGINA_{idx_pagina}'
                registro_id_erro = f'{idx_pagina}:{chave_erro}'
                removido_usuario = registro_id_erro in removidos_set or chave_erro in removidos_set
                item_erro = {
                    'pagina': idx_pagina,
                    'registroId': registro_id_erro,
                    'chave': chave_erro,
                    'nome': '',
                    'cpf': '',
                    'matricula': '00000',
                    'origemMatricula': 'nao_encontrada',
                    'requerPreenchimento': True,
                    'consistenciaOk': False,
                    'somaRemuneracao': '0,00',
                    'totalRemuneracao': '0,00',
                    'inconsistencias': [
                        f'Falha OCR da página {idx_pagina}: {erro_ocr or "-"}',
                        f'Falha de extração textual da página {idx_pagina}: {erro_texto or "-"}',
                    ],
                    'bloqueiaGeracao': True,
                    'linhasGeradas': 0,
                    'camposRemuneracao': [],
                    'valoresEditaveis': _garantir_campos_editaveis({}),
                    'confirmadoUsuario': False,
                    'removidoUsuario': removido_usuario,
                    'statusRegistro': 'removido' if removido_usuario else 'pendente',
                    'duplicadoIgnorado': False,
                }
                funcionarios.append(item_erro)
                if not removido_usuario:
                    pendencias.append(item_erro)
                continue

            total_paginas_com_extracao += 1
            nome_res = str(getattr(resultado, 'empregado', '') or '')
            cpf_res = str(getattr(resultado, 'cpf', '') or '')
            assinatura = '|'.join([
                _normalizar_nome(nome_res),
                _normalizar_cpf(cpf_res),
                _normalizar_nome(str(getattr(resultado, 'periodo_inicial', '') or '')),
                _normalizar_nome(str(getattr(resultado, 'periodo_final', '') or '')),
            ])
            if assinatura in assinaturas_paginas_vistas and assinatura != '|||':
                total_duplicados_ignorados += 1
                funcionarios.append({
                    'pagina': idx_pagina,
                    'registroId': f'{idx_pagina}:DUPLICADO_{idx_pagina}',
                    'chave': f'DUPLICADO_{idx_pagina}',
                    'nome': nome_res,
                    'cpf': cpf_res,
                    'matricula': _formatar_codigo(str(getattr(resultado, 'matricula', '') or '')),
                    'origemMatricula': 'duplicado_ignorado',
                    'requerPreenchimento': False,
                    'consistenciaOk': True,
                    'somaRemuneracao': '0,00',
                    'totalRemuneracao': '0,00',
                    'inconsistencias': ['Página ignorada por cópia duplicada de funcionário já processado.'],
                    'bloqueiaGeracao': False,
                    'linhasGeradas': 0,
                    'camposRemuneracao': [],
                    'valoresEditaveis': _garantir_campos_editaveis({}),
                    'duplicadoIgnorado': True,
                    'confirmadoUsuario': False,
                    'removidoUsuario': False,
                    'statusRegistro': 'duplicado_ignorado',
                })
                continue
            assinaturas_paginas_vistas.add(assinatura)

            chave = montar_chave_funcionario(nome_res, cpf_res)
            registro_id = f'{idx_pagina}:{chave}'
            confirmado_usuario = (registro_id in confirmados_set) or (chave in confirmados_set)
            removido_usuario = (registro_id in removidos_set) or (chave in removidos_set)

            valores = _garantir_campos_editaveis(getattr(resultado, 'valores', {}) or {})
            valores = _aplicar_overrides_valores(valores, registro_id=registro_id, chave=chave, overrides_por_registro=overrides_por_registro)
            resultado.valores = valores

            matricula_manual = str(ids_por_chave.get(chave, '') or '').strip()
            matricula_final, origem_matricula = _resolver_matricula(
                nome=nome_res,
                cpf=cpf_res,
                matricula_extraida=str(getattr(resultado, 'matricula', '') or ''),
                matricula_manual=matricula_manual,
                indice_por_cpf=indice_por_cpf,
                indice_por_nome_cpf=indice_por_nome_cpf,
            )
            resultado.matricula = matricula_final

            soma_remuneracao, total_remuneracao, consistencia_ok = _avaliar_consistencia(valores)
            pendencia_matricula = matricula_final == '00000'
            pendencia_consistencia = not consistencia_ok
            inconsistencias: list[str] = []
            if pendencia_matricula:
                inconsistencias.append('Matrícula não identificada. Informe a matrícula para liberar a geração.')
            if pendencia_consistencia:
                inconsistencias.append(
                    f'Soma da remuneracao difere do total. Soma: {_decimal_para_br(soma_remuneracao)} | Total: {_decimal_para_br(total_remuneracao)}.'
                )

            if removido_usuario:
                pendencia_matricula = False
                pendencia_consistencia = False
                inconsistencias = []
            elif confirmado_usuario and not pendencia_matricula:
                # Confirmado pelo usuário: a divergência de valores deixa de bloquear a geração.
                pendencia_consistencia = False
                inconsistencias = []

            # Regra operacional: divergência entre soma dos fatores e total sempre bloqueia.
            bloqueia_geracao = (pendencia_matricula or pendencia_consistencia) and not removido_usuario
            if confirmado_usuario and not pendencia_matricula:
                bloqueia_geracao = False

            linhas_funcionario: list[str] = []
            if not bloqueia_geracao and not removido_usuario:
                try:
                    linhas_funcionario = list(modulo.montar_linhas_txt(resultado) or [])
                except Exception:
                    linhas_funcionario = []

                if linhas_funcionario:
                    linhas_preview.extend(linhas_funcionario)
                    cpf_limpo = _normalizar_cpf(cpf_res) or 'SEMCPF'
                    matricula_limpa = _formatar_codigo(matricula_final)
                    nome_limpo = _slug_nome_arquivo(nome_res)
                    nome_txt = f'{idx_pagina:03d}_{nome_limpo}_{cpf_limpo}_{matricula_limpa}.txt'
                    arquivos_txt_gerados.append((nome_txt, '\n'.join(linhas_funcionario)))

            status_registro = 'ok'
            if removido_usuario:
                status_registro = 'removido'
            elif bloqueia_geracao:
                status_registro = 'pendente'
            elif confirmado_usuario:
                status_registro = 'confirmado_usuario'

            item = {
                'pagina': idx_pagina,
                'registroId': registro_id,
                'chave': chave,
                'nome': nome_res,
                'cpf': cpf_res,
                'matricula': matricula_final,
                'origemMatricula': origem_matricula,
                'requerPreenchimento': matricula_final == '00000',
                'consistenciaOk': consistencia_ok,
                'somaRemuneracao': _decimal_para_br(soma_remuneracao),
                'totalRemuneracao': _decimal_para_br(total_remuneracao),
                'inconsistencias': inconsistencias,
                'bloqueiaGeracao': bloqueia_geracao,
                'linhasGeradas': len(linhas_funcionario),
                'camposRemuneracao': list(CAMPOS_SOMA_REMUNERACAO),
                'valoresEditaveis': valores,
                'duplicadoIgnorado': False,
                'confirmadoUsuario': confirmado_usuario,
                'removidoUsuario': removido_usuario,
                'statusRegistro': status_registro,
            }
            funcionarios.append(item)
            if bloqueia_geracao:
                pendencias.append(item)

    if total_paginas_com_extracao == 0:
        raise RuntimeError(
            'Falha completa: nenhuma página teve extração válida (OCR/texto).'
        )

    pode_gerar_txt = len(arquivos_txt_gerados) > 0 and len(pendencias) == 0
    nome_base = Path(nome_arquivo or 'arquivo.pdf').stem
    nome_saida = f'{nome_base}.zip'
    if pode_gerar_txt:
        if len(arquivos_txt_gerados) == 1:
            tipo_saida = 'txt'
            nome_saida = f'{nome_base}.txt'
            arquivo_mime_type = 'text/plain;charset=utf-8'
            arquivo_base64 = base64.b64encode(arquivos_txt_gerados[0][1].encode('utf-8')).decode('ascii')
        else:
            tipo_saida = 'zip'
            arquivo_mime_type = 'application/zip'
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
                for nome_txt, conteudo_txt in arquivos_txt_gerados:
                    zf.writestr(nome_txt, conteudo_txt.encode('utf-8'))
            arquivo_base64 = base64.b64encode(buffer.getvalue()).decode('ascii')

    return {
        'nome_saida': nome_saida,
        'arquivo_base64': arquivo_base64,
        'arquivo_mime_type': arquivo_mime_type,
        'tipo_saida': tipo_saida,
        'linhas': linhas_preview,
        'funcionarios': funcionarios,
        'pendencias': pendencias,
        'pode_gerar_txt': pode_gerar_txt,
        'bloqueado_geracao': not pode_gerar_txt,
        'mensagem_bloqueio': (
            'Geração travada: corrija matrícula faltante e/ou ajuste os fatores para que a soma da remuneração seja igual ao total de remuneração.'
            if not pode_gerar_txt
            else ''
        ),
        'total_paginas': total_paginas,
        'total_funcionarios': len([f for f in funcionarios if not bool(f.get('duplicadoIgnorado'))]),
        'total_pendencias': len(pendencias),
        'total_duplicados_ignorados': total_duplicados_ignorados,
        'total_arquivos_txt': len(arquivos_txt_gerados),
        'lista_funcionarios_arquivo': str(LISTA_FUNCIONARIOS),
    }


def salvar_funcionario_lista(nome: str, cpf: str, matricula: str) -> dict[str, Any]:
    if not LISTA_FUNCIONARIOS.exists():
        raise FileNotFoundError('Planilha de funcionários não encontrada.')

    nome_limpo = str(nome or '').strip()
    cpf_limpo = _normalizar_cpf(cpf)
    codigo = _formatar_codigo(matricula)

    if not nome_limpo:
        raise ValueError('Nome do funcionário é obrigatório.')
    if len(cpf_limpo) != 11:
        raise ValueError('CPF inválido. Informe 11 dígitos.')
    if codigo == '00000':
        raise ValueError('Matrícula inválida. Informe um código numérico válido.')

    workbook = load_workbook(LISTA_FUNCIONARIOS)

    encontrados: list[tuple[int, Any, int]] = []
    nome_norm = _normalizar_nome(nome_limpo)

    for idx, ws in enumerate(workbook.worksheets):
        if ws.max_row < 2 or ws.max_column < 4:
            continue

        col_nome = 2
        col_nome_completo = 3
        col_cpf = ws.max_column

        for row in range(2, ws.max_row + 1):
            cpf_row = _normalizar_cpf(str(ws.cell(row, col_cpf).value or ''))
            if cpf_row != cpf_limpo:
                continue

            nome_row = _normalizar_nome(str(ws.cell(row, col_nome).value or ''))
            nome_comp_row = _normalizar_nome(str(ws.cell(row, col_nome_completo).value or ''))
            if nome_norm and nome_norm not in {nome_row, nome_comp_row}:
                continue

            encontrados.append((idx, ws, row))

    if encontrados:
        _, ws_alvo, row_alvo = sorted(encontrados, key=lambda item: (item[0], item[2]))[-1]
        ws_alvo.cell(row_alvo, 1, codigo)
        if not str(ws_alvo.cell(row_alvo, 2).value or '').strip():
            ws_alvo.cell(row_alvo, 2, nome_limpo)
        if ws_alvo.max_column >= 3 and not str(ws_alvo.cell(row_alvo, 3).value or '').strip():
            ws_alvo.cell(row_alvo, 3, nome_limpo)
        ws_alvo.cell(row_alvo, ws_alvo.max_column, _formatar_cpf_visual(cpf_limpo))

        workbook.save(LISTA_FUNCIONARIOS)
        return {
            'acao': 'atualizado',
            'aba': ws_alvo.title,
            'linha': row_alvo,
            'codigo': codigo,
            'arquivo': str(LISTA_FUNCIONARIOS),
        }

    ws = workbook.worksheets[0]
    nova_linha = ws.max_row + 1
    ws.cell(nova_linha, 1, codigo)
    ws.cell(nova_linha, 2, nome_limpo)
    if ws.max_column >= 3:
        ws.cell(nova_linha, 3, nome_limpo)
    ws.cell(nova_linha, ws.max_column, _formatar_cpf_visual(cpf_limpo))

    workbook.save(LISTA_FUNCIONARIOS)
    return {
        'acao': 'criado',
        'aba': ws.title,
        'linha': nova_linha,
        'codigo': codigo,
        'arquivo': str(LISTA_FUNCIONARIOS),
    }
