import argparse
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


FONTE_CORPO = Font(name="Calibri", size=10)
FONTE_CABECALHO = Font(name="Calibri", size=12, bold=True)
CABECALHOS = [
    "NF",
    "FORNECEDOR",
    "DESCRICAO",
    "NCM",
    "CFOP",
    "CST_ICMS",
    "CST_PIS",
    "CST_COFINS",
]
ENCODINGS_CANDIDATOS = ("utf-8-sig", "utf-8", "latin1")


class ExtratorFiscalError(Exception):
    def __init__(self, code, message, details=None):
        super().__init__(message)
        self.code = code
        self.details = details or []


def garantir_pasta(caminho):
    os.makedirs(caminho, exist_ok=True)


def escolher_encoding(caminho_arquivo):
    ultimo_erro = None
    for encoding in ENCODINGS_CANDIDATOS:
        try:
            with open(caminho_arquivo, "r", encoding=encoding) as handle:
                handle.readline()
            return encoding
        except UnicodeDecodeError as exc:
            ultimo_erro = exc
            continue
    raise ExtratorFiscalError(
        "encoding_error",
        "Nao foi possivel ler o arquivo SPED com os encodings suportados.",
        [str(ultimo_erro)] if ultimo_erro else [],
    )


def parse_sped(caminho_arquivo):
    fornecedores = {}
    produtos = {}
    linhas = []
    periodo = ""
    encoding_utilizado = escolher_encoding(caminho_arquivo)

    with open(caminho_arquivo, "r", encoding=encoding_utilizado) as handle:
        fornecedor_atual = None
        nf_atual = ""
        is_entrada = False

        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue

            fields = line.split("|")
            if len(fields) < 2:
                continue

            reg = fields[1]

            if reg == "0000":
                if len(fields) >= 6:
                    periodo = f"{fields[4]} - {fields[5]}"
            elif reg == "0150":
                if len(fields) >= 4:
                    fornecedores[fields[2]] = fields[3]
            elif reg == "0200":
                if len(fields) >= 9:
                    produtos[fields[2]] = (fields[3], fields[8])
            elif reg == "C100":
                if len(fields) >= 9:
                    if fields[2] == "0":
                        is_entrada = True
                        fornecedor_atual = fornecedores.get(fields[4], fields[4])
                        nf_atual = fields[8]
                    else:
                        is_entrada = False
                        fornecedor_atual = None
                        nf_atual = ""
            elif reg == "C170":
                if not is_entrada or len(fields) < 12:
                    continue

                cod_item = fields[3]
                descricao, ncm = produtos.get(cod_item, (cod_item, ""))
                cst_pis = fields[25].strip() if len(fields) > 25 else ""
                cst_cofins = fields[31].strip() if len(fields) > 31 else ""

                linhas.append(
                    {
                        "NF": nf_atual,
                        "FORNECEDOR": fornecedor_atual or "",
                        "DESCRICAO": descricao,
                        "NCM": ncm,
                        "CFOP": fields[11],
                        "CST_ICMS": fields[10],
                        "CST_PIS": cst_pis,
                        "CST_COFINS": cst_cofins,
                    }
                )

    return {
        "periodo": periodo,
        "linhas": linhas,
        "encoding": encoding_utilizado,
    }


def gerar_nome_excel(periodo):
    data_inicio = str(periodo or "").split(" - ")[0].strip()
    if len(data_inicio) == 8 and data_inicio.isdigit():
        mes = data_inicio[2:4]
        ano = data_inicio[4:8]
        return f"compras_{mes}{ano}.xlsx"
    return "compras_indefinido.xlsx"


def salvar_excel(dados, arquivo_saida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    for col, titulo in enumerate(CABECALHOS, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = FONTE_CABECALHO
        cell.alignment = Alignment(horizontal="center")

    for row_idx, reg in enumerate(dados, 2):
        ws.cell(row=row_idx, column=1, value=reg["NF"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=2, value=reg["FORNECEDOR"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=3, value=reg["DESCRICAO"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=4, value=reg["NCM"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=5, value=reg["CFOP"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=6, value=reg["CST_ICMS"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=7, value=reg["CST_PIS"]).font = FONTE_CORPO
        ws.cell(row=row_idx, column=8, value=reg["CST_COFINS"]).font = FONTE_CORPO

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(arquivo_saida)


def escrever_log(log_dir, payload):
    garantir_pasta(log_dir)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(log_dir, f"execucao_{timestamp}.json")
    with open(log_path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
    return log_path


def processar_arquivo(caminho_arquivo, output_dir, log_dir):
    if not os.path.exists(caminho_arquivo):
        raise ExtratorFiscalError("file_not_found", "Arquivo SPED nao encontrado.")

    garantir_pasta(output_dir)
    garantir_pasta(log_dir)

    resultado = parse_sped(caminho_arquivo)
    periodo = resultado["periodo"]
    linhas = resultado["linhas"]
    encoding_utilizado = resultado["encoding"]

    if not linhas:
        raise ExtratorFiscalError(
            "no_items",
            "Nenhum item de compra encontrado no arquivo.",
            [f"Encoding utilizado: {encoding_utilizado}"],
        )

    nome_excel = gerar_nome_excel(periodo)
    output_path = os.path.join(output_dir, nome_excel)
    salvar_excel(linhas, output_path)

    log_payload = {
        "ok": True,
        "arquivo_entrada": caminho_arquivo,
        "arquivo_saida": output_path,
        "periodo": periodo,
        "total_itens": len(linhas),
        "encoding": encoding_utilizado,
        "executado_em": datetime.now().isoformat(),
    }
    log_path = escrever_log(log_dir, log_payload)

    return {
        "ok": True,
        "message": "Arquivo processado com sucesso.",
        "periodo": periodo,
        "totalItems": len(linhas),
        "encoding": encoding_utilizado,
        "outputPath": output_path,
        "outputFileName": nome_excel,
        "logPath": log_path,
    }


def main():
    parser = argparse.ArgumentParser(description="Extrator fiscal SPED")
    parser.add_argument("--input", required=True, help="Caminho do arquivo SPED TXT.")
    parser.add_argument("--output-dir", required=True, help="Diretorio de saida do XLSX.")
    parser.add_argument("--log-dir", required=True, help="Diretorio para logs da execucao.")
    args = parser.parse_args()

    try:
        result = processar_arquivo(args.input, args.output_dir, args.log_dir)
        print(json.dumps(result, ensure_ascii=False))
        return 0
    except ExtratorFiscalError as exc:
        log_path = escrever_log(
            args.log_dir,
            {
                "ok": False,
                "code": exc.code,
                "message": str(exc),
                "details": exc.details,
                "arquivo_entrada": args.input,
                "executado_em": datetime.now().isoformat(),
            },
        )
        payload = {
            "ok": False,
            "code": exc.code,
            "message": str(exc),
            "details": exc.details,
            "logPath": log_path,
        }
        print(json.dumps(payload, ensure_ascii=False))
        return 1
    except Exception as exc:
        log_path = escrever_log(
            args.log_dir,
            {
                "ok": False,
                "code": "unexpected_error",
                "message": "Falha inesperada ao processar o SPED.",
                "details": [f"{type(exc).__name__}: {exc}"],
                "arquivo_entrada": args.input,
                "executado_em": datetime.now().isoformat(),
            },
        )
        payload = {
            "ok": False,
            "code": "unexpected_error",
            "message": "Falha inesperada ao processar o SPED.",
            "details": [f"{type(exc).__name__}: {exc}"],
            "logPath": log_path,
        }
        print(json.dumps(payload, ensure_ascii=False))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
