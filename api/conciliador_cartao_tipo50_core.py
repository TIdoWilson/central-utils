from __future__ import annotations

import base64
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, List, Tuple
from unicodedata import normalize

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    from PyPDF2 import PdfReader


TOLERANCIA_DIFERENCA = Decimal("0.10")
XLSX_FORMATO_MOEDA_BR = '[$R$-416] #,##0.00'

RE_MONEY_2_DEC = re.compile(r"^(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_MONEY_ANY = re.compile(r"(\d{1,3}(?:\.\d{3})*,\d{2})")
RE_INICIO_LIVRO = re.compile(
    r"^(\d{2}/\d{2}/\d{4})\s+(NFe|NF|CTe)\s+\d+\s+(\d+)\s+"
    r"\d{2}/\d{2}/\d{4}\s+\d+\s+[A-Z]{2}\s+00\s*/\s*00\s+([\d\.,]+)\s+(\d\.\d{3})"
)
RE_CONTINUACAO_LIVRO = re.compile(r"^([\d\.,]+)\s+(\d\.\d{3})(?:\s|$)")
RE_INICIO_RELATORIO = re.compile(r"^(\d{2}/\d{2}/\d{3})(.*)$")
RE_CHAVE_RELATORIO = re.compile(r"^\s*(\d+)\s+(\d\.\d{3})\s+(.+)$")


@dataclass(frozen=True)
class Registro:
    nota: str
    cfop: str
    valor: Decimal


def gerar_nome_saida() -> str:
    agora = datetime.now()
    return f"Conciliação_{agora.strftime('%H-%M-%S')}_{agora.strftime('%d-%m-%Y')}.xlsx"


def ler_linhas_pdf_bytes(pdf_bytes: bytes) -> List[str]:
    leitor = PdfReader(io.BytesIO(pdf_bytes))
    texto = "\n".join((pagina.extract_text() or "") for pagina in leitor.pages)
    return [linha.strip() for linha in texto.splitlines() if linha.strip()]


def valor_br_para_decimal(valor_br: str) -> Decimal:
    return Decimal(valor_br.replace(".", "").replace(",", "."))


def normalizar_nota(nota: str) -> str:
    return re.sub(r"\D", "", nota)


def extrair_valor_duas_casas(token: str) -> Decimal | None:
    match = RE_MONEY_2_DEC.match(token)
    if not match:
        return None
    return valor_br_para_decimal(match.group(1))


def extrair_total_rodape_livro(linhas: List[str]) -> Decimal | None:
    for linha in linhas:
        if linha.startswith("Total Geral"):
            match = RE_MONEY_ANY.search(linha)
            if match:
                return valor_br_para_decimal(match.group(1))
    return None


def extrair_total_rodape_relatorio(linhas: List[str]) -> Decimal | None:
    for i, linha in enumerate(linhas):
        if linha.startswith("Registros.:"):
            for prox in linhas[i + 1 : i + 20]:
                match = RE_MONEY_ANY.search(prox)
                if match:
                    return valor_br_para_decimal(match.group(1))
    return None


def identificar_tipo_pdf(linhas: List[str]) -> str | None:
    cabecalho = normalize("NFKD", "\n".join(linhas[:60])).encode("ASCII", "ignore").decode().upper()
    texto_completo = normalize("NFKD", "\n".join(linhas)).encode("ASCII", "ignore").decode().upper()

    if "REGISTRO DE ENTRADAS - MODELO P1" in cabecalho or "TOTAL GERAL" in texto_completo:
        return "livro"

    if "RELATORIO DO ARQUIVO MAGN" in cabecalho and "REGISTROS.:" in texto_completo:
        return "tipo50"

    return None


def parse_livro_registro(linhas: List[str]) -> List[Registro]:
    registros: List[Registro] = []
    nota_atual: str | None = None

    for linha in linhas:
        inicio = RE_INICIO_LIVRO.match(linha)
        if inicio:
            _, _, nota, token_valor, cfop = inicio.groups()
            nota_atual = nota
            valor = extrair_valor_duas_casas(token_valor)
            if valor is not None:
                registros.append(Registro(nota=nota, cfop=cfop, valor=valor))
            continue

        continuacao = RE_CONTINUACAO_LIVRO.match(linha)
        if continuacao and nota_atual:
            token_valor, cfop = continuacao.groups()
            valor = extrair_valor_duas_casas(token_valor)
            if valor is not None:
                registros.append(Registro(nota=nota_atual, cfop=cfop, valor=valor))

    return registros


def parse_relatorio_tipo50(linhas: List[str]) -> List[Registro]:
    registros: List[Registro] = []

    for linha in linhas:
        inicio = RE_INICIO_RELATORIO.match(linha)
        if not inicio:
            continue

        _, resto = inicio.groups()
        chave = RE_CHAVE_RELATORIO.match(resto)
        if not chave:
            continue

        nota, cfop, sufixo = chave.groups()
        primeiro_token = sufixo.split()[0] if sufixo.split() else ""
        valor = extrair_valor_duas_casas(primeiro_token)
        if valor is None:
            continue
        registros.append(Registro(nota=nota, cfop=cfop, valor=valor))

    return registros


def agrupar_por_nota(registros: List[Registro]) -> Dict[str, Decimal]:
    acumulado: Dict[str, Decimal] = defaultdict(Decimal)
    for registro in registros:
        nota_normalizada = normalizar_nota(registro.nota)
        if nota_normalizada:
            acumulado[nota_normalizada] += registro.valor
    return dict(acumulado)


def agrupar_por_nota_cfop(registros: List[Registro]) -> Dict[Tuple[str, str], Decimal]:
    acumulado: Dict[Tuple[str, str], Decimal] = defaultdict(Decimal)
    for registro in registros:
        nota_normalizada = normalizar_nota(registro.nota)
        if nota_normalizada:
            acumulado[(nota_normalizada, registro.cfop)] += registro.valor
    return dict(acumulado)


def anular_pares_opostos_linhas(linhas: List[List[object]], idx_diferenca: int) -> List[List[object]]:
    saldos: Dict[int, List[List[object]]] = {}
    resultado: List[List[object]] = []

    for linha in linhas:
        valor = linha[idx_diferenca]
        if not isinstance(valor, (int, float)):
            resultado.append(linha)
            continue

        cents = int(round(valor * 100))
        oposto = -cents
        if oposto in saldos and saldos[oposto]:
            saldos[oposto].pop()
            continue
        saldos.setdefault(cents, []).append(linha)

    for bucket in saldos.values():
        resultado.extend(bucket)
    return resultado


def anular_pares_opostos_cfop(linhas: List[List[object]]) -> List[List[object]]:
    # linhas: [nota, cfop, valor_livro, valor_relatorio]
    # anula apenas dentro da mesma NOTA+CFOP
    saldos: Dict[Tuple[str, str, int], List[List[object]]] = {}
    resultado: List[List[object]] = []

    for linha in linhas:
        nota = str(linha[0])
        cfop = str(linha[1])
        vl = linha[2]
        vr = linha[3]

        if not isinstance(vl, (int, float)) or not isinstance(vr, (int, float)):
            resultado.append(linha)
            continue

        diferenca = float(vl) - float(vr)
        cents = int(round(diferenca * 100))
        oposto = -cents

        chave_oposta = (nota, cfop, oposto)
        if chave_oposta in saldos and saldos[chave_oposta]:
            saldos[chave_oposta].pop()
            continue

        chave = (nota, cfop, cents)
        saldos.setdefault(chave, []).append(linha)

    for bucket in saldos.values():
        resultado.extend(bucket)
    return resultado


def autoajustar_colunas(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            texto = "" if val is None else str(val)
            if len(texto) > max_len:
                max_len = len(texto)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 60))


def gerar_xlsx_bytes(linhas_divergencias: List[List[object]], linhas_cfop: List[List[object]]) -> bytes:
    wb = Workbook()
    ws_div = wb.active
    ws_div.title = "DIVERGENCIAS"
    ws_div.append(["TIPO", "NOTA", "DIFERENÇA"])

    for linha in linhas_divergencias:
        ws_div.append(linha)

    ws_div["E1"] = "DIVERGÊNCIA TOTAL"
    ws_div["E2"] = "=SUM(C:C)"
    ws_div["E2"].number_format = XLSX_FORMATO_MOEDA_BR
    ws_div.freeze_panes = "A2"

    for row in ws_div.iter_rows(min_row=2, max_row=ws_div.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = XLSX_FORMATO_MOEDA_BR

    ws_cfop = wb.create_sheet("CFOP_DIFERENTE")
    ws_cfop.append(["NOTA", "CFOP", "VALOR_LIVRO", "VALOR_RELATÓRIO"])

    for linha in linhas_cfop:
        ws_cfop.append(linha)

    ws_cfop.freeze_panes = "A2"
    for row in ws_cfop.iter_rows(min_row=2, max_row=ws_cfop.max_row, min_col=3, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = XLSX_FORMATO_MOEDA_BR

    for ws in (ws_div, ws_cfop):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.upper()
        autoajustar_colunas(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _to_records_div(rows: List[List[object]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "tipo": r[0],
                "nota": r[1],
                "diferenca": float(r[2]),
            }
        )
    return out


def _to_records_cfop(rows: List[List[object]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in rows:
        out.append(
            {
                "nota": r[0],
                "cfop": r[1],
                "valor_livro": float(r[2]),
                "valor_relatorio": float(r[3]),
            }
        )
    return out


def conciliar_cartao_tipo50(
    arquivo_a_bytes: bytes,
    arquivo_a_nome: str,
    arquivo_b_bytes: bytes,
    arquivo_b_nome: str,
) -> Dict[str, Any]:
    linhas_pdf_1 = ler_linhas_pdf_bytes(arquivo_a_bytes)
    linhas_pdf_2 = ler_linhas_pdf_bytes(arquivo_b_bytes)

    tipo_1 = identificar_tipo_pdf(linhas_pdf_1)
    tipo_2 = identificar_tipo_pdf(linhas_pdf_2)

    if tipo_1 is None or tipo_2 is None:
        raise ValueError(
            "Nao foi possivel identificar o tipo de um dos PDFs. Use um Livro de Registros IOB e um Balancete Tipo 50."
        )
    if tipo_1 == tipo_2:
        raise ValueError("Os dois arquivos parecem ser do mesmo tipo. Envie 1 Livro e 1 Tipo 50.")

    if tipo_1 == "livro":
        livro_nome = arquivo_a_nome
        tipo50_nome = arquivo_b_nome
        linhas_livro = linhas_pdf_1
        linhas_relatorio = linhas_pdf_2
    else:
        livro_nome = arquivo_b_nome
        tipo50_nome = arquivo_a_nome
        linhas_livro = linhas_pdf_2
        linhas_relatorio = linhas_pdf_1

    registros_livro = parse_livro_registro(linhas_livro)
    registros_relatorio = parse_relatorio_tipo50(linhas_relatorio)
    if not registros_livro:
        raise ValueError("Nao consegui extrair dados do Livro de Registros IOB.")
    if not registros_relatorio:
        raise ValueError("Nao consegui extrair dados do Balancete Tipo 50.")

    total_rodape_livro = extrair_total_rodape_livro(linhas_livro)
    total_rodape_relatorio = extrair_total_rodape_relatorio(linhas_relatorio)

    mapa_livro = agrupar_por_nota(registros_livro)
    mapa_relatorio = agrupar_por_nota(registros_relatorio)
    mapa_livro_cfop = agrupar_por_nota_cfop(registros_livro)
    mapa_relatorio_cfop = agrupar_por_nota_cfop(registros_relatorio)

    notas = set(mapa_livro) | set(mapa_relatorio)
    divergencias_xlsx: List[List[object]] = []

    for nota in sorted(notas):
        valor_livro = mapa_livro.get(nota)
        valor_relatorio = mapa_relatorio.get(nota)

        if valor_livro is not None and valor_relatorio is None:
            divergencias_xlsx.append(["SÓ_NO_LIVRO", nota, float(valor_livro)])
            continue

        if valor_relatorio is not None and valor_livro is None:
            divergencias_xlsx.append(["SÓ_NO_RELATÓRIO", nota, float(-valor_relatorio)])
            continue

        if valor_livro is not None and valor_relatorio is not None:
            diferenca = valor_livro - valor_relatorio
            if abs(diferenca) <= TOLERANCIA_DIFERENCA:
                continue
            divergencias_xlsx.append(["VALOR_DIVERGENTE", nota, float(diferenca)])

    chaves_cfop = set(mapa_livro_cfop) | set(mapa_relatorio_cfop)
    divergencias_cfop_xlsx: List[List[object]] = []
    for nota, cfop in sorted(chaves_cfop):
        valor_livro = mapa_livro_cfop.get((nota, cfop), Decimal("0"))
        valor_relatorio = mapa_relatorio_cfop.get((nota, cfop), Decimal("0"))

        if round(float(valor_relatorio), 2) == 0.0:
            continue

        diferenca = valor_livro - valor_relatorio
        if abs(diferenca) <= TOLERANCIA_DIFERENCA:
            continue

        divergencias_cfop_xlsx.append([
            nota,
            cfop,
            float(valor_livro),
            float(valor_relatorio),
        ])

    divergencias_xlsx = anular_pares_opostos_linhas(divergencias_xlsx, 2)
    divergencias_cfop_xlsx = anular_pares_opostos_cfop(divergencias_cfop_xlsx)

    total_parse_livro = sum(mapa_livro.values(), Decimal("0"))
    total_parse_relatorio = sum(mapa_relatorio.values(), Decimal("0"))
    diferenca_parse = total_parse_livro - total_parse_relatorio

    diferenca_rodape = None
    if total_rodape_livro is not None and total_rodape_relatorio is not None:
        diferenca_rodape = total_rodape_livro - total_rodape_relatorio

    xlsx_bytes = gerar_xlsx_bytes(divergencias_xlsx, divergencias_cfop_xlsx)

    resumo = {
        "arquivo_livro": livro_nome,
        "arquivo_tipo50": tipo50_nome,
        "registros_livro": len(registros_livro),
        "registros_relatorio": len(registros_relatorio),
        "total_parse_livro": float(total_parse_livro),
        "total_parse_relatorio": float(total_parse_relatorio),
        "diferenca_parse": float(diferenca_parse),
        "total_rodape_livro": float(total_rodape_livro) if total_rodape_livro is not None else None,
        "total_rodape_relatorio": float(total_rodape_relatorio) if total_rodape_relatorio is not None else None,
        "diferenca_rodape": float(diferenca_rodape) if diferenca_rodape is not None else None,
        "linhas_divergencias": len(divergencias_xlsx),
        "linhas_cfop_diferente": len(divergencias_cfop_xlsx),
        "tolerancia": float(TOLERANCIA_DIFERENCA),
    }

    return {
        "filename": gerar_nome_saida(),
        "xlsxBase64": base64.b64encode(xlsx_bytes).decode("ascii"),
        "resumo": resumo,
        "divergencias": _to_records_div(divergencias_xlsx),
        "cfopDiferente": _to_records_cfop(divergencias_cfop_xlsx),
    }
