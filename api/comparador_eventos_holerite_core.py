from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


MESES_PT_BR = {
    "JANEIRO": "01",
    "FEVEREIRO": "02",
    "MARCO": "03",
    "MARCHO": "03",
    "ABRIL": "04",
    "MAIO": "05",
    "JUNHO": "06",
    "JULHO": "07",
    "AGOSTO": "08",
    "SETEMBRO": "09",
    "OUTUBRO": "10",
    "NOVEMBRO": "11",
    "DEZEMBRO": "12",
}

MESES_NUMERO_PARA_NOME = {
    "01": "Janeiro",
    "02": "Fevereiro",
    "03": "Marco",
    "04": "Abril",
    "05": "Maio",
    "06": "Junho",
    "07": "Julho",
    "08": "Agosto",
    "09": "Setembro",
    "10": "Outubro",
    "11": "Novembro",
    "12": "Dezembro",
}

TEXTOS_IGNORADOS = {
    "",
    "NOME",
    "NOME DO FUNCIONARIO",
    "CODIGO",
    "CBO",
    "EMP.",
    "LOCAL",
    "DEPTO.",
    "SETOR",
    "SECAO",
    "FL.",
    "GERAL",
    "CONTINUA",
    "RECIBO DE PAGAMENTO DE SALARIO",
    "RECIBO DE RETIRADAS",
    "DESCRICAO",
    "REFERENCIA",
    "VENCIMENTOS",
    "DESCONTOS",
    "VALOR LIQUIDO",
    "TOTAL DE VENCIMENTOS",
    "TOTAL DE DESCONTOS",
    "SALARIO BASE",
    "SAL. CONTR. INSS",
    "BASE CALC. FGTS",
    "FGTS DO MES",
    "BASE CALC. IRRF",
    "FAIXA IRRF",
}


@dataclass
class Celula:
    ordem: int
    linha: int
    coluna: int
    valor: str


@dataclass
class EventoParcial:
    codigo: str = ""
    descricao: str = ""
    referencia: str = ""
    valor: str = ""


@dataclass(frozen=True)
class Evento:
    codigo: str
    descricao: str
    referencia: str = ""
    valor: str = ""

    @property
    def chave(self) -> tuple[str, str]:
        return normalizar_codigo(self.codigo), normalizar_texto(self.descricao)

    @property
    def exibicao(self) -> str:
        return f"{self.codigo} - {self.descricao}"


@dataclass
class BlocoHolerite:
    competencia: str
    codigo_funcionario: str = ""
    funcionario: str = ""
    linhas: dict[int, EventoParcial] = field(default_factory=dict)


@dataclass(frozen=True)
class LinhaComparacaoEvento:
    funcionario: str
    codigo_funcionario: str
    nome_funcionario: str
    evento: Evento
    referencia: str
    valor_anterior: str
    valor_atual: str = ""
    situacao: str = ""


def remover_acentos(texto: str) -> str:
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")


def normalizar_texto(texto: str) -> str:
    texto = remover_acentos(texto or "").upper()
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_codigo(texto: str) -> str:
    return re.sub(r"\D", "", texto or "")


def formatar_valor(texto: str) -> str:
    valor = str(texto or "").strip()
    if not valor:
        return ""
    if re.fullmatch(r"-?\d+(?:\.\d+)?", valor):
        numero = float(valor)
        texto_formatado = f"{numero:,.2f}"
        return texto_formatado.replace(",", "X").replace(".", ",").replace("X", ".")
    return valor


def formatar_competencia(valor: str) -> str:
    valor = str(valor or "").strip()
    if re.fullmatch(r"\d{2}/\d{2}", valor):
        return valor
    if re.fullmatch(r"\d{2}/\d{4}", valor):
        return f"{valor[:2]}/{valor[-2:]}"
    if re.fullmatch(r"\d{6}", valor):
        return f"{valor[:2]}/{valor[-2:]}"
    raise ValueError(f"Competencia invalida: {valor}")


def competencia_por_extenso(valor: str) -> str:
    competencia = formatar_competencia(valor)
    mes, ano_curto = competencia.split("/")
    return f"{MESES_NUMERO_PARA_NOME[mes]} / 20{ano_curto}"


def competencia_para_ordenacao(valor: str) -> tuple[int, int]:
    competencia = formatar_competencia(valor)
    mes, ano_curto = competencia.split("/")
    return 2000 + int(ano_curto), int(mes)


def normalizar_competencia_texto(texto_mes: str) -> str:
    texto_limpo = texto_mes.replace("Mes:", "").replace("Mês:", "").strip()
    texto_limpo = re.sub(r"\s+", " ", texto_limpo)

    match_numerico = re.search(r"(\d{2})/(\d{4})", texto_limpo)
    if match_numerico:
        return f"{match_numerico.group(1)}/{match_numerico.group(2)[-2:]}"

    match_texto = re.search(r"([A-Za-zÀ-ÿ]+)\s*/\s*(\d{4})", texto_limpo, flags=re.IGNORECASE)
    if not match_texto:
        raise ValueError(f"Nao foi possivel interpretar a competencia: {texto_mes}")

    mes_nome = normalizar_texto(match_texto.group(1))
    ano = match_texto.group(2)
    if mes_nome not in MESES_PT_BR:
        raise ValueError(f"Mes nao reconhecido: {match_texto.group(1)}")
    return f"{MESES_PT_BR[mes_nome]}/{ano[-2:]}"


def ler_linhas_slk_bytes(arquivo_bytes: bytes) -> list[str]:
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return arquivo_bytes.decode(encoding).splitlines()
        except UnicodeDecodeError:
            continue
    raise ValueError("Falha ao ler o arquivo SLK.")


def extrair_valor_k(valor_bruto: str) -> str:
    valor_bruto = valor_bruto.strip()
    if valor_bruto.startswith('"') and valor_bruto.endswith('"'):
        return valor_bruto[1:-1]
    return valor_bruto


def parsear_celulas_slk_bytes(arquivo_bytes: bytes) -> list[Celula]:
    padrao_f = re.compile(r"^F;(?:[^;]+;)*Y(\d+);X(\d+)$")
    padrao_c = re.compile(r"^C;(?:(?:Y(\d+));)?(?:(?:X(\d+));)?K(.+)$")

    linhas = ler_linhas_slk_bytes(arquivo_bytes)
    celulas: list[Celula] = []
    linha_atual = 0
    coluna_atual = 0

    for ordem, conteudo in enumerate(linhas):
        match_f = padrao_f.match(conteudo)
        if match_f:
            linha_atual = int(match_f.group(1))
            coluna_atual = int(match_f.group(2))
            continue

        match_c = padrao_c.match(conteudo)
        if not match_c:
            continue

        if match_c.group(1):
            linha_atual = int(match_c.group(1))
        if match_c.group(2):
            coluna_atual = int(match_c.group(2))

        if linha_atual <= 0 or coluna_atual <= 0:
            continue

        celulas.append(
            Celula(
                ordem=ordem,
                linha=linha_atual,
                coluna=coluna_atual,
                valor=extrair_valor_k(match_c.group(3)),
            )
        )

    return celulas


def eh_nome_funcionario(texto: str) -> bool:
    texto_norm = normalizar_texto(texto)
    if texto_norm in TEXTOS_IGNORADOS:
        return False
    if texto_norm.startswith("ADMISSAO:") or texto_norm.startswith("BANCO :") or texto_norm.startswith("AGENCIA :"):
        return False
    if texto_norm.startswith("DEPOSITO EFETUADO"):
        return False
    return len(texto_norm) >= 5 and any(char.isalpha() for char in texto_norm)


def eh_codigo_evento(texto: str) -> bool:
    return len(normalizar_codigo(texto)) >= 4


def eh_descricao_evento(texto: str) -> bool:
    texto_norm = normalizar_texto(texto)
    if texto_norm in TEXTOS_IGNORADOS:
        return False
    if texto_norm.startswith("ADMISSAO:") or texto_norm.startswith("BANCO :") or texto_norm.startswith("AGENCIA :"):
        return False
    if texto_norm.startswith("DEPOSITO EFETUADO"):
        return False
    return any(char.isalpha() for char in texto_norm)


def eh_referencia_evento(texto: str) -> bool:
    texto_limpo = str(texto or "").strip()
    if not texto_limpo:
        return False
    texto_norm = normalizar_texto(texto_limpo)
    return texto_norm not in TEXTOS_IGNORADOS


def eh_valor_evento(texto: str) -> bool:
    texto_limpo = str(texto or "").strip()
    if not texto_limpo:
        return False
    texto_norm = normalizar_texto(texto_limpo)
    if texto_norm in TEXTOS_IGNORADOS:
        return False
    return bool(re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+(?:\.\d+)?", texto_limpo))


def formatar_identificacao_funcionario(codigo: str, nome: str) -> str:
    codigo_limpo = str(codigo or "").strip()
    nome_limpo = str(nome or "").strip()
    if codigo_limpo and nome_limpo:
        return f"{codigo_limpo} - {nome_limpo}"
    return nome_limpo or codigo_limpo


def separar_identificacao_funcionario(identificacao: str) -> tuple[str, str]:
    texto = str(identificacao or "").strip()
    if " - " not in texto:
        return "", texto
    codigo, nome = texto.split(" - ", 1)
    return codigo.strip(), nome.strip()


def situacao_para_exibicao(situacao: str) -> str:
    mapa = {
        "Confere": "Confere",
        "Alterado": "Alterado",
        "Nao existe": "Não existe",
    }
    return mapa.get(str(situacao or "").strip(), str(situacao or "").strip())


def adicionar_eventos_do_bloco(bloco: BlocoHolerite, destino: dict[str, dict[str, dict[tuple[str, str], Evento]]]) -> None:
    if not bloco.funcionario:
        return

    funcionario = formatar_identificacao_funcionario(bloco.codigo_funcionario, bloco.funcionario)
    eventos_funcionario = destino.setdefault(bloco.competencia, {}).setdefault(funcionario, {})

    for parcial in bloco.linhas.values():
        if not eh_codigo_evento(parcial.codigo):
            continue
        if not eh_descricao_evento(parcial.descricao):
            continue
        if normalizar_texto(parcial.descricao) == normalizar_texto(funcionario):
            continue
        if parcial.referencia and not eh_referencia_evento(parcial.referencia):
            parcial.referencia = ""
        if parcial.valor and not eh_valor_evento(parcial.valor):
            parcial.valor = ""

        evento = Evento(
            codigo=normalizar_codigo(parcial.codigo),
            descricao=parcial.descricao.strip(),
            referencia=parcial.referencia.strip(),
            valor=parcial.valor.strip(),
        )
        eventos_funcionario[evento.chave] = evento


def extrair_eventos_por_competencia_bytes(arquivo_bytes: bytes) -> dict[str, dict[str, dict[tuple[str, str], Evento]]]:
    resultado: dict[str, dict[str, dict[tuple[str, str], Evento]]] = {}
    bloco_atual: BlocoHolerite | None = None
    aguardando_funcionario = False

    for celula in parsear_celulas_slk_bytes(arquivo_bytes):
        valor = celula.valor.strip()
        if not valor:
            continue

        if valor.startswith("Mes:") or valor.startswith("Mês:"):
            if bloco_atual is not None:
                adicionar_eventos_do_bloco(bloco_atual, resultado)
            bloco_atual = BlocoHolerite(competencia=normalizar_competencia_texto(valor))
            aguardando_funcionario = True
            continue

        if bloco_atual is None:
            continue

        if aguardando_funcionario and celula.coluna == 1 and not bloco_atual.codigo_funcionario and eh_codigo_evento(valor):
            bloco_atual.codigo_funcionario = valor.strip()
            continue

        if aguardando_funcionario and celula.coluna == 2 and eh_nome_funcionario(valor):
            bloco_atual.funcionario = valor.strip()
            aguardando_funcionario = False
            continue

        if not bloco_atual.funcionario:
            continue

        parcial = bloco_atual.linhas.setdefault(celula.linha, EventoParcial())
        if celula.coluna == 1 and eh_codigo_evento(valor):
            parcial.codigo = valor
            continue
        if celula.coluna == 2 and eh_descricao_evento(valor):
            parcial.descricao = valor
            continue
        if celula.coluna == 3 and eh_referencia_evento(valor):
            parcial.referencia = valor.strip()
            continue
        if celula.coluna in (4, 5) and eh_valor_evento(valor):
            parcial.valor = formatar_valor(valor)

    if bloco_atual is not None:
        adicionar_eventos_do_bloco(bloco_atual, resultado)

    return resultado


def comparar_competencias(
    eventos_por_competencia: dict[str, dict[str, dict[tuple[str, str], Evento]]],
    competencia_anterior: str,
    competencia_atual: str,
) -> list[LinhaComparacaoEvento]:
    competencia_inicial = formatar_competencia(competencia_anterior)
    competencia_final = formatar_competencia(competencia_atual)

    if competencia_inicial not in eventos_por_competencia:
        raise ValueError(f"Competencia nao encontrada no arquivo: {competencia_inicial}")
    if competencia_final not in eventos_por_competencia:
        raise ValueError(f"Competencia nao encontrada no arquivo: {competencia_final}")

    comparacao: list[LinhaComparacaoEvento] = []
    eventos_iniciais = eventos_por_competencia[competencia_inicial]
    eventos_finais = eventos_por_competencia[competencia_final]

    for funcionario in sorted(eventos_iniciais):
        chaves_eventos = sorted(
            set(eventos_iniciais[funcionario]),
            key=lambda item: (item[1], item[0]),
        )
        if not chaves_eventos:
            continue

        codigo_funcionario, nome_funcionario = separar_identificacao_funcionario(funcionario)
        for chave in chaves_eventos:
            evento_anterior = eventos_iniciais[funcionario][chave]
            evento_atual = eventos_finais.get(funcionario, {}).get(chave)
            if not evento_atual:
                situacao = "Nao existe"
            elif evento_anterior.valor == evento_atual.valor:
                situacao = "Confere"
            else:
                situacao = "Alterado"

            comparacao.append(
                LinhaComparacaoEvento(
                    funcionario=funcionario,
                    codigo_funcionario=codigo_funcionario,
                    nome_funcionario=nome_funcionario,
                    evento=evento_anterior,
                    referencia=evento_anterior.referencia or evento_atual.referencia if evento_atual else evento_anterior.referencia,
                    valor_anterior=evento_anterior.valor,
                    valor_atual=evento_atual.valor if evento_atual else "",
                    situacao=situacao,
                )
            )

    return comparacao


def detectar_competencias(
    eventos_por_competencia: dict[str, dict[str, dict[tuple[str, str], Evento]]],
    competencia_anterior: str | None = None,
    competencia_atual: str | None = None,
) -> tuple[str, str]:
    if competencia_anterior and competencia_atual:
        return formatar_competencia(competencia_anterior), formatar_competencia(competencia_atual)

    competencias = sorted(eventos_por_competencia.keys(), key=competencia_para_ordenacao)
    if len(competencias) != 2:
        raise ValueError(f"O arquivo deve conter exatamente 2 competencias. Encontradas: {', '.join(competencias) or 'nenhuma'}")
    return competencias[0], competencias[1]


def gerar_excel_bytes(
    linhas_comparacao: list[LinhaComparacaoEvento],
    competencia_anterior: str,
    competencia_atual: str,
) -> bytes:
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Comparacao de Eventos"

    preenchimento_titulo = PatternFill(fill_type="solid", fgColor="1F4E78")
    preenchimento_meta = PatternFill(fill_type="solid", fgColor="EEF4F8")
    preenchimento_confere = PatternFill(fill_type="solid", fgColor="DCFCE7")
    preenchimento_alterado = PatternFill(fill_type="solid", fgColor="FEF3C7")
    preenchimento_nao_existe = PatternFill(fill_type="solid", fgColor="FEE2E2")
    borda = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    planilha["A1"] = "RELATORIO DE COMPARACAO DE EVENTOS"
    planilha["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    planilha["A1"].fill = preenchimento_titulo
    planilha["A1"].alignment = Alignment(horizontal="center", vertical="center")
    planilha.merge_cells("A1:F1")

    for ref, valor in (
        ("A3", "Competencia"),
        ("B3", competencia_por_extenso(competencia_anterior)),
        ("C3", competencia_por_extenso(competencia_atual)),
    ):
        planilha[ref] = valor
        planilha[ref].font = Font(bold=True)
        planilha[ref].fill = preenchimento_meta
        planilha[ref].border = borda
        planilha[ref].alignment = Alignment(horizontal="center", vertical="center")

    linha = 5
    if not linhas_comparacao:
        planilha.cell(linha, 1, "Nenhum evento encontrado para comparacao entre as competencias.")
    else:
        planilha.cell(linha, 1, "Funcionario")
        planilha.cell(linha, 2, "Evento")
        planilha.cell(linha, 3, "Referencia")
        planilha.cell(linha, 4, competencia_por_extenso(competencia_anterior))
        planilha.cell(linha, 5, competencia_por_extenso(competencia_atual))
        planilha.cell(linha, 6, "Situação")
        for coluna in range(1, 7):
            celula = planilha.cell(linha, coluna)
            celula.font = Font(color="FFFFFF", bold=True)
            celula.fill = preenchimento_titulo
            celula.border = borda
            celula.alignment = Alignment(horizontal="center", vertical="center")
        linha += 1

        inicio_grupo = linha
        funcionario_anterior = None
        for item in linhas_comparacao:
            planilha.cell(linha, 1, item.funcionario)
            planilha.cell(linha, 2, item.evento.exibicao)
            planilha.cell(linha, 3, item.referencia)
            planilha.cell(linha, 4, item.valor_anterior)
            planilha.cell(linha, 5, item.valor_atual)
            planilha.cell(linha, 6, situacao_para_exibicao(item.situacao))

            for coluna in range(1, 7):
                planilha.cell(linha, coluna).border = borda
                planilha.cell(linha, coluna).alignment = Alignment(horizontal="left", vertical="center")

            celula_situacao = planilha.cell(linha, 6)
            if item.situacao == "Confere":
                celula_situacao.fill = preenchimento_confere
            elif item.situacao == "Alterado":
                celula_situacao.fill = preenchimento_alterado
            elif item.situacao == "Nao existe":
                celula_situacao.fill = preenchimento_nao_existe

            if funcionario_anterior is None:
                funcionario_anterior = item.funcionario
                inicio_grupo = linha
            elif funcionario_anterior != item.funcionario:
                if linha - 1 > inicio_grupo:
                    planilha.merge_cells(start_row=inicio_grupo, start_column=1, end_row=linha - 1, end_column=1)
                funcionario_anterior = item.funcionario
                inicio_grupo = linha

            linha += 1

        if linha - 1 > inicio_grupo:
            planilha.merge_cells(start_row=inicio_grupo, start_column=1, end_row=linha - 1, end_column=1)

        planilha.column_dimensions["A"].width = 38
        planilha.column_dimensions["B"].width = 34
        planilha.column_dimensions["C"].width = 16
        planilha.column_dimensions["D"].width = 18
        planilha.column_dimensions["E"].width = 18
        planilha.column_dimensions["F"].width = 14

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def aplicar_filtro_eventos(
    linhas_comparacao: list[LinhaComparacaoEvento],
    eventos_ocultos: list[str] | None = None,
) -> list[LinhaComparacaoEvento]:
    eventos_ocultos_set = {str(item or "").strip() for item in (eventos_ocultos or []) if str(item or "").strip()}
    if not eventos_ocultos_set:
        return linhas_comparacao
    return [item for item in linhas_comparacao if item.evento.exibicao not in eventos_ocultos_set]


def processar_comparador_eventos_holerite(
    arquivo_bytes: bytes,
    nome_arquivo: str,
    competencia_anterior: str | None = None,
    competencia_atual: str | None = None,
    eventos_ocultos: list[str] | None = None,
) -> dict:
    eventos_por_competencia = extrair_eventos_por_competencia_bytes(arquivo_bytes)
    competencia_anterior, competencia_atual = detectar_competencias(
        eventos_por_competencia,
        competencia_anterior=competencia_anterior,
        competencia_atual=competencia_atual,
    )
    linhas_comparacao = comparar_competencias(eventos_por_competencia, competencia_anterior, competencia_atual)
    linhas_comparacao = aplicar_filtro_eventos(linhas_comparacao, eventos_ocultos=eventos_ocultos)
    xlsx_bytes = gerar_excel_bytes(linhas_comparacao, competencia_anterior, competencia_atual)

    return {
        "competencias_encontradas": sorted(eventos_por_competencia.keys()),
        "competencia_anterior": formatar_competencia(competencia_anterior),
        "competencia_atual": formatar_competencia(competencia_atual),
        "competencia_anterior_extenso": competencia_por_extenso(competencia_anterior),
        "competencia_atual_extenso": competencia_por_extenso(competencia_atual),
        "preview": [
            {
                "funcionario": item.funcionario,
                "codigo_funcionario": item.codigo_funcionario,
                "nome_funcionario": item.nome_funcionario,
                "evento": item.evento.exibicao,
                "referencia": item.referencia,
                "valor_anterior": item.valor_anterior,
                "valor_atual": item.valor_atual,
                "situacao": item.situacao,
            }
            for item in linhas_comparacao
        ],
        "total_funcionarios_com_diferenca": len({item.funcionario for item in linhas_comparacao}),
        "arquivo_saida": f"{nome_arquivo.rsplit('.', 1)[0]}_eventos_ausentes_{formatar_competencia(competencia_anterior).replace('/', '-')}_vs_{formatar_competencia(competencia_atual).replace('/', '-')}.xlsx",
        "xlsx_bytes": xlsx_bytes,
    }
