from __future__ import annotations

import csv
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pdfplumber

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


TXT_SAIDA_NOME = "LOTD0000.txt"
PENDENCIAS_NOME = "PENDENCIAS_GFBR.csv"
EXCLUSOES_NOME = "EXCLUSOES_GFBR.csv"

TOLERANCIA_PAREAMENTO_VALOR = Decimal("0.01")

PALAVRAS_EXCLUSAO_DIRETA = (
    "notas fiscais de saida",
    "nota fiscal de saida",
    "recebimento",
    "a/r invoice",
)

PALAVRAS_CANCELAMENTO = (
    "cancelamento",
    "cancelado",
    "cancellation",
    "cancelled",
    "anular",
    "anulado",
    "anulada",
    "estorno",
    "estornar",
    "estornado",
    "reversal",
    "reversed",
)

PALAVRAS_RENDA = (
    "renda",
    "rendas",
    "rendimento",
    "rendimentos",
)

PARTNER_PREFIX_TO_ACCOUNT = {
    "TG": "112020401",
    "CL": "112010101",
    "F": "211010101",
    "BF": "112010101",
    "SW": "112020401",
    "DR": "112020401",
}

# Campos do layout IOB usado no lotes-txt (public/js/lotes-txt.js)
IOB_FIELDS = {
    "DTI": 2,
    "DTF": 9,
    "DBI": 10,
    "DBF": 15,
    "CRI": 16,
    "CRF": 21,
    "H3I": 22,
    "H3F": 24,
    "CPI": 25,
    "CPF": 49,
    "VLI": 50,
    "VLF": 64,
    "CDI": 68,
    "CDF": 81,
    "CCI": 82,
    "CCF": 95,
    "SEQI": 96,
    "SEQF": 100,
    "H4I": 533,
    "H4F": 536,
}

HEADER_ALIASES = {
    "seq": {
        "n seq",
        "n seq.",
        "n seq",
        "no seq",
        "no seq.",
        "numero seq",
        "numero de seq",
        "numero de sequencia",
    },
    "numero_transacao": {
        "n transacao",
        "n transacao.",
        "no transacao",
        "no transacao.",
        "numero transacao",
        "numero de transacao",
    },
    "data_lancamento": {
        "data de lancamento",
    },
    "serie": {
        "serie",
    },
    "numero_doc": {
        "n doc",
        "n doc.",
        "no doc",
        "no doc.",
        "numero doc",
        "numero de doc",
    },
    "codigo_conta": {
        "cta contab cod pn",
        "cta contab cod pn.",
        "cta contab cod pn",
    },
    "nome_conta": {
        "cta cont nome pn",
    },
    "valor": {
        "debito credito mc",
        "debito credito (mc)",
    },
    "observacoes": {
        "observacoes",
        "observacao",
    },
    "nome_filial": {
        "nome da filial",
    },
    "marca": {
        "marca",
    },
    "centro_custos": {
        "centro de custos",
    },
}


def normalizar_texto(texto: Any) -> str:
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def texto_ascii(texto: Any, max_len: Optional[int] = None) -> str:
    s = "" if texto is None else str(texto)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^\x20-\x7E]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if max_len is not None:
        return s[:max_len]
    return s


def to_str_limpo(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, Decimal) and valor == valor.to_integral_value():
        return str(int(valor))
    return str(valor).strip()


def remover_siglas_parceiros(texto: Any) -> str:
    t = str(texto) if texto else ""
    if not t:
        return ""
    t = re.sub(r"\b(?:BF|SW|TG|DR|CL|F)\d+\b", "", t, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", t).strip()



def parse_decimal(valor: Any) -> Optional[Decimal]:
    if valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, int):
        return Decimal(valor)
    if isinstance(valor, float):
        return Decimal(str(valor))

    texto = str(valor).strip()
    if not texto:
        return None
    texto = re.sub(r"[^\d,.\-]", "", texto)
    if not texto or texto in {"-", ",", "."}:
        return None
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    elif "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def parse_data(valor: Any) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    if not texto:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    return None


def data_para_yyyymmdd(valor: Optional[date]) -> str:
    if not valor:
        return ""
    return valor.strftime("%Y-%m-%d")


def data_para_ddmmyyyy(valor: Optional[date]) -> str:
    if not valor:
        return datetime.now().strftime("%d%m%Y")
    return valor.strftime("%d%m%Y")


def localizar_linha_cabecalho(
    ws: Worksheet,
) -> Tuple[int, Dict[str, int]]:
    alias_to_key: Dict[str, str] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_to_key[normalizar_texto(alias)] = key

    ultimo_mapa: Dict[str, int] = {}
    required = {"seq", "numero_transacao", "data_lancamento", "codigo_conta", "valor"}

    for idx_linha in range(1, min(ws.max_row, 40) + 1):
        mapa: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            nome = normalizar_texto(ws.cell(row=idx_linha, column=col).value)
            key = alias_to_key.get(nome)
            if key and key not in mapa:
                mapa[key] = col
        ultimo_mapa = mapa
        if required.issubset(set(mapa.keys())):
            return idx_linha, mapa

    faltantes = sorted(required.difference(set(ultimo_mapa.keys())))
    raise ValueError(f"Cabecalhos obrigatorios nao encontrados: {faltantes}")


def valor_coluna(ws: Worksheet, idx_linha: int, mapa: Dict[str, int], key: str) -> Any:
    col = mapa.get(key)
    if not col:
        return None
    return ws.cell(row=idx_linha, column=col).value


def agrupar_lancamentos(
    ws: Worksheet, linha_cabecalho: int, colunas: Dict[str, int]
) -> List[Dict[str, Any]]:
    lancamentos: List[Dict[str, Any]] = []
    atual: Optional[Dict[str, Any]] = None
    contador_id = 0

    for idx_linha in range(linha_cabecalho + 1, ws.max_row + 1):
        seq_raw = to_str_limpo(valor_coluna(ws, idx_linha, colunas, "seq"))
        nova_operacao = seq_raw != ""

        if nova_operacao:
            if atual:
                lancamentos.append(atual)
            contador_id += 1
            atual = {
                "_id": contador_id,
                "seq": seq_raw,
                "numero_transacao": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "numero_transacao")
                ),
                "data_lancamento": parse_data(
                    valor_coluna(ws, idx_linha, colunas, "data_lancamento")
                ),
                "serie": to_str_limpo(valor_coluna(ws, idx_linha, colunas, "serie")),
                "numero_doc": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "numero_doc")
                ),
                "observacao_principal": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "observacoes")
                ),
                "nome_filial": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "nome_filial")
                ),
                "marca": to_str_limpo(valor_coluna(ws, idx_linha, colunas, "marca")),
                "centro_custos": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "centro_custos")
                ),
                "itens_contabeis": [],
            }

        if atual is None:
            continue

        codigo = to_str_limpo(valor_coluna(ws, idx_linha, colunas, "codigo_conta"))
        descricao = to_str_limpo(valor_coluna(ws, idx_linha, colunas, "nome_conta"))
        valor = parse_decimal(valor_coluna(ws, idx_linha, colunas, "valor"))
        observacao_linha = to_str_limpo(
            valor_coluna(ws, idx_linha, colunas, "observacoes")
        )

        if not codigo and not descricao and valor is None:
            continue

        atual["itens_contabeis"].append(
            {
                "codigo_original": codigo,
                "descricao_original": descricao,
                "valor": valor,
                "observacao": observacao_linha,
                "marca": to_str_limpo(valor_coluna(ws, idx_linha, colunas, "marca"))
                or atual["marca"],
                "centro_custos": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "centro_custos")
                )
                or atual["centro_custos"],
                "nome_filial": to_str_limpo(
                    valor_coluna(ws, idx_linha, colunas, "nome_filial")
                )
                or atual["nome_filial"],
            }
        )

        if not atual["observacao_principal"] and observacao_linha:
            atual["observacao_principal"] = observacao_linha

    if atual:
        lancamentos.append(atual)

    return lancamentos


def texto_observacoes_lancamento(lancamento: Dict[str, Any]) -> str:
    partes = [lancamento.get("observacao_principal", "")]
    partes.extend(item.get("observacao", "") for item in lancamento.get("itens_contabeis", []))
    return " | ".join(p for p in partes if p)


def texto_descritivo_lancamento(lancamento: Dict[str, Any]) -> str:
    partes = [
        lancamento.get("observacao_principal", ""),
        lancamento.get("historico", ""),
    ]
    for item in lancamento.get("itens_contabeis", []) or lancamento.get("itens", []):
        partes.append(item.get("descricao_original", "") or item.get("descricao", ""))
        partes.append(item.get("observacao", ""))
    return " | ".join(p for p in partes if p)


def contem_termo(texto: str, termos: Iterable[str]) -> bool:
    texto_norm = normalizar_texto(texto)
    return any(normalizar_texto(t) in texto_norm for t in termos)


def tem_conta_com_prefixo_cl(lancamento: Dict[str, Any]) -> bool:
    for item in lancamento.get("itens_contabeis", []):
        codigo = re.sub(r"\s+", "", to_str_limpo(item.get("codigo_original", ""))).upper()
        if codigo.startswith("CL"):
            return True
    return False


def eh_lancamento_de_exclusao_direta(lancamento: Dict[str, Any]) -> bool:
    return contem_termo(texto_observacoes_lancamento(lancamento), PALAVRAS_EXCLUSAO_DIRETA)


def conta_com_prefixo_11102(lancamento: Dict[str, Any]) -> bool:
    for item in lancamento.get("itens_contabeis", []) or lancamento.get("itens", []):
        conta = re.sub(r"\D", "", to_str_limpo(item.get("conta_final", "")))
        if conta.startswith("11102"):
            return True
    return False


def eh_lancamento_renda_11102(lancamento: Dict[str, Any]) -> bool:
    if not conta_com_prefixo_11102(lancamento):
        return False
    return contem_termo(texto_descritivo_lancamento(lancamento), PALAVRAS_RENDA)


def identificar_lancamentos_cancelatorios(
    lancamentos: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    saida: List[Dict[str, Any]] = []
    for lanc in lancamentos:
        if contem_termo(texto_observacoes_lancamento(lanc), PALAVRAS_CANCELAMENTO):
            saida.append(lanc)
    return saida


def total_lancamento(lancamento: Dict[str, Any]) -> Decimal:
    total = Decimal("0")
    for item in lancamento.get("itens_contabeis", []):
        valor = item.get("valor")
        if isinstance(valor, Decimal):
            total += valor
    return total


def construir_registro_exclusao(
    lancamento: Dict[str, Any], motivo: str, transacao_par: str = ""
) -> Dict[str, Any]:
    return {
        "seq": lancamento.get("seq", ""),
        "numero_transacao": lancamento.get("numero_transacao", ""),
        "numero_doc": lancamento.get("numero_doc", ""),
        "data": data_para_yyyymmdd(lancamento.get("data_lancamento")),
        "observacao": lancamento.get("observacao_principal", "") or lancamento.get("historico", ""),
        "motivo_exclusao": motivo,
        "transacao_par_vinculada": transacao_par,
    }


def marcar_pares_para_exclusao(
    lancamentos: List[Dict[str, Any]],
) -> Tuple[set, List[Dict[str, Any]]]:
    excluidos_ids: set = set()
    exclusoes: List[Dict[str, Any]] = []
    por_id: Dict[int, Dict[str, Any]] = {}

    for lanc in lancamentos:
        if eh_lancamento_de_exclusao_direta(lanc):
            excluidos_ids.add(lanc["_id"])
            por_id[lanc["_id"]] = construir_registro_exclusao(
                lanc, "exclusao_direta_tipo_operacao"
            )
        elif tem_conta_com_prefixo_cl(lanc):
            excluidos_ids.add(lanc["_id"])
            por_id[lanc["_id"]] = construir_registro_exclusao(
                lanc, "exclusao_direta_recebimento_cl"
            )

    # Pareamento simples por modulo de valor para cancelamentos/estornos.
    # Mantem comportamento tolerante sem bloquear processamento.
    candidatos = [x for x in lancamentos if x["_id"] not in excluidos_ids]
    cancelatorios = identificar_lancamentos_cancelatorios(candidatos)
    usados = set()

    for est in cancelatorios:
        if est["_id"] in excluidos_ids or est["_id"] in usados:
            continue
        total_est = abs(total_lancamento(est))
        melhor = None
        melhor_dif = None
        for cand in candidatos:
            if cand["_id"] == est["_id"]:
                continue
            if cand["_id"] in excluidos_ids or cand["_id"] in usados:
                continue
            total_cand = abs(total_lancamento(cand))
            dif = abs(total_est - total_cand)
            if dif <= TOLERANCIA_PAREAMENTO_VALOR:
                if melhor is None or dif < (melhor_dif or Decimal("999999")):
                    melhor = cand
                    melhor_dif = dif
        if melhor:
            excluidos_ids.add(est["_id"])
            excluidos_ids.add(melhor["_id"])
            usados.add(est["_id"])
            usados.add(melhor["_id"])
            por_id[est["_id"]] = construir_registro_exclusao(
                est,
                "cancelamento_estorno_anulacao",
                transacao_par=melhor.get("numero_transacao", ""),
            )
            por_id[melhor["_id"]] = construir_registro_exclusao(
                melhor,
                "par_do_cancelamento_estorno_anulacao",
                transacao_par=est.get("numero_transacao", ""),
            )
        else:
            excluidos_ids.add(est["_id"])
            usados.add(est["_id"])
            por_id[est["_id"]] = construir_registro_exclusao(
                est,
                "cancelamento_estorno_sem_par_confirmado",
            )

    exclusoes.extend(por_id.values())
    return excluidos_ids, exclusoes


def eh_conta_contabil(codigo: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)+", codigo.strip()))


def eh_codigo_parceiro_generico(codigo: str) -> bool:
    codigo_limpo = re.sub(r"\s+", "", codigo).upper()
    return bool(re.fullmatch(r"[A-Z]{1,4}\d{2,}", codigo_limpo))


def substituir_codigo_parceiro_por_prefixo(
    codigo: str,
) -> Tuple[Optional[str], Optional[str]]:
    codigo_limpo = re.sub(r"\s+", "", to_str_limpo(codigo)).upper()
    if not codigo_limpo:
        return None, "codigo_vazio"

    for prefixo in sorted(PARTNER_PREFIX_TO_ACCOUNT.keys(), key=len, reverse=True):
        if codigo_limpo.startswith(prefixo):
            return PARTNER_PREFIX_TO_ACCOUNT[prefixo], None

    if eh_conta_contabil(codigo_limpo):
        return codigo_limpo, None

    if eh_codigo_parceiro_generico(codigo_limpo):
        return None, f"prefixo_nao_mapeado:{codigo_limpo}"

    return None, f"codigo_sem_classificacao:{codigo_limpo}"


def resolver_conta_final_item(
    item: Dict[str, Any],
) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    conta_final, erro = substituir_codigo_parceiro_por_prefixo(item.get("codigo_original", ""))

    if erro:
        return None, {
            "codigo_parceiro": item.get("codigo_original", ""),
            "nome_parceiro": item.get("descricao_original", ""),
            "valor": item.get("valor", ""),
            "motivo_da_pendencia": erro,
        }

    if item.get("valor") is None:
        return None, {
            "codigo_parceiro": item.get("codigo_original", ""),
            "nome_parceiro": item.get("descricao_original", ""),
            "valor": "",
            "motivo_da_pendencia": "item_sem_valor",
        }

    return {
        "conta_original": item.get("codigo_original", ""),
        "conta_final": conta_final,
        "descricao": item.get("descricao_original", ""),
        "valor": item.get("valor"),
        "observacao": item.get("observacao", ""),
        "marca": item.get("marca", ""),
        "centro_custos": item.get("centro_custos", ""),
        "nome_filial": item.get("nome_filial", ""),
    }, None


def normalizar_lancamento(
    bruto: Dict[str, Any], pendencias: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    normalizado = {
        "seq": bruto.get("seq", ""),
        "numero_transacao": bruto.get("numero_transacao", ""),
        "data_lancamento": bruto.get("data_lancamento"),
        "data": data_para_yyyymmdd(bruto.get("data_lancamento")),
        "serie": bruto.get("serie", ""),
        "numero_doc": bruto.get("numero_doc", ""),
        "historico": texto_ascii(remover_siglas_parceiros(bruto.get("observacao_principal", "")), 120),
        "nome_filial": bruto.get("nome_filial", ""),
        "marca": bruto.get("marca", ""),
        "centro_custos": bruto.get("centro_custos", ""),
        "itens": [],
    }

    for item in bruto.get("itens_contabeis", []):
        resolvido, pendencia = resolver_conta_final_item(item)
        if pendencia:
            pendencias.append(
                {
                    "seq": bruto.get("seq", ""),
                    "numero_transacao": bruto.get("numero_transacao", ""),
                    "numero_doc": bruto.get("numero_doc", ""),
                    "data": data_para_yyyymmdd(bruto.get("data_lancamento")),
                    "observacao": bruto.get("observacao_principal", ""),
                    **pendencia,
                }
            )
            continue
        normalizado["itens"].append(resolvido)

    if not normalizado["itens"]:
        pendencias.append(
            {
                "seq": bruto.get("seq", ""),
                "numero_transacao": bruto.get("numero_transacao", ""),
                "numero_doc": bruto.get("numero_doc", ""),
                "data": data_para_yyyymmdd(bruto.get("data_lancamento")),
                "observacao": bruto.get("observacao_principal", ""),
                "codigo_parceiro": "",
                "nome_parceiro": "",
                "valor": "",
                "motivo_da_pendencia": "lancamento_sem_itens_validos",
            }
        )
        return None

    return normalizado


def conta_para_campos(conta: str) -> Tuple[str, str]:
    c = re.sub(r"\s+", "", to_str_limpo(conta)).upper()
    if re.fullmatch(r"\d{1,6}", c):
        return c.zfill(6), ""
    c_limpo = re.sub(r"[^A-Z0-9]", "", c)
    classificacao = texto_ascii(c_limpo, 14).upper()
    return "000000", classificacao


def set_range(
    base: List[str], inicio: int, fim: int, valor: str, *, align_right: bool = False, fill: str = " "
) -> None:
    width = fim - inicio + 1
    v = str(valor or "")
    if align_right:
        v = v.rjust(width, fill)[-width:]
    else:
        v = v[:width].ljust(width, fill)
    for i, ch in enumerate(v, start=inicio - 1):
        base[i] = ch


def decimal_to_cents_str(valor: Decimal, width: int = 15) -> str:
    cents = int((valor.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)) * 100)
    cents = abs(cents)
    return str(cents).rjust(width, "0")[-width:]


def build_l_line(entry: Dict[str, Any], sequencia: int) -> str:
    line = [" "] * 536
    line[0] = "L"

    db_red, db_cls = conta_para_campos(entry["debito_conta"])
    cr_red, cr_cls = conta_para_campos(entry["credito_conta"])

    set_range(line, IOB_FIELDS["DTI"], IOB_FIELDS["DTF"], entry["data_ddmmyyyy"])
    set_range(line, IOB_FIELDS["DBI"], IOB_FIELDS["DBF"], db_red, align_right=True, fill="0")
    set_range(line, IOB_FIELDS["CRI"], IOB_FIELDS["CRF"], cr_red, align_right=True, fill="0")
    set_range(line, IOB_FIELDS["H3I"], IOB_FIELDS["H3F"], "000", align_right=True, fill="0")
    set_range(line, IOB_FIELDS["CPI"], IOB_FIELDS["CPF"], texto_ascii(entry.get("complemento", ""), 25))
    set_range(
        line,
        IOB_FIELDS["VLI"],
        IOB_FIELDS["VLF"],
        decimal_to_cents_str(entry["valor"]),
        align_right=True,
        fill="0",
    )
    set_range(line, IOB_FIELDS["CDI"], IOB_FIELDS["CDF"], db_cls)
    set_range(line, IOB_FIELDS["CCI"], IOB_FIELDS["CCF"], cr_cls)
    set_range(
        line,
        IOB_FIELDS["SEQI"],
        IOB_FIELDS["SEQF"],
        str(sequencia),
        align_right=True,
        fill="0",
    )
    set_range(line, IOB_FIELDS["H4I"], IOB_FIELDS["H4F"], "0000", align_right=True, fill="0")
    return "".join(line)


def build_h_line(texto_hist: str, sequencia: int) -> str:
    line = [" "] * 100
    line[0] = "H"
    set_range(line, 2, 51, texto_ascii(texto_hist, 50))
    set_range(line, 96, 100, str(sequencia), align_right=True, fill="0")
    return "".join(line)


def build_c_line(data_ddmmyyyy: str, total_debitos: Decimal) -> str:
    line = [" "] * 100
    line[0] = "C"
    line[1] = "M"
    set_range(line, 3, 10, data_ddmmyyyy)
    set_range(line, 11, 25, decimal_to_cents_str(total_debitos, width=15), align_right=True, fill="0")
    set_range(line, 26, 50, "DIARIO SW GFBR OUT")
    set_range(line, 96, 100, "00002", align_right=True, fill="0")
    return "".join(line)


def criar_partidas_iob(
    lancamentos: List[Dict[str, Any]], pendencias: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    saida: List[Dict[str, Any]] = []

    for lanc in lancamentos:
        debitos: List[Dict[str, Any]] = []
        creditos: List[Dict[str, Any]] = []
        data_ddmmyyyy = data_para_ddmmyyyy(lanc.get("data_lancamento"))
        complemento = texto_ascii(lanc.get("historico", ""), 25)

        for item in lanc.get("itens", []):
            valor = item.get("valor")
            if not isinstance(valor, Decimal):
                continue
            conta = item.get("conta_final", "")
            payload = {
                "conta": conta,
                "restante": abs(valor),
                "descricao": texto_ascii(remover_siglas_parceiros(item.get("descricao", "")), 40),
                "observacao": texto_ascii(remover_siglas_parceiros(item.get("observacao", "")), 50),
                "marca": texto_ascii(item.get("marca", ""), 20),
                "centro_custos": texto_ascii(item.get("centro_custos", ""), 20),
            }
            if valor > 0:
                debitos.append(payload)
            elif valor < 0:
                creditos.append(payload)

        i = 0
        j = 0
        while i < len(debitos) and j < len(creditos):
            d = debitos[i]
            c = creditos[j]
            valor_partida = min(d["restante"], c["restante"])
            if valor_partida > Decimal("0"):
                hist_h = " | ".join(
                    x
                    for x in [
                        lanc.get("numero_doc", ""),
                        d.get("descricao", ""),
                        c.get("descricao", ""),
                        d.get("observacao", ""),
                    ]
                    if x
                )
                saida.append(
                    {
                        "data_ddmmyyyy": data_ddmmyyyy,
                        "debito_conta": d["conta"],
                        "credito_conta": c["conta"],
                        "valor": valor_partida,
                        "complemento": complemento,
                        "historico_h": texto_ascii(hist_h, 50),
                    }
                )
            d["restante"] -= valor_partida
            c["restante"] -= valor_partida
            if d["restante"] <= TOLERANCIA_PAREAMENTO_VALOR:
                i += 1
            if c["restante"] <= TOLERANCIA_PAREAMENTO_VALOR:
                j += 1

        for idx in range(i, len(debitos)):
            d = debitos[idx]
            if d["restante"] > TOLERANCIA_PAREAMENTO_VALOR:
                pendencias.append(
                    {
                        "seq": lanc.get("seq", ""),
                        "numero_transacao": lanc.get("numero_transacao", ""),
                        "numero_doc": lanc.get("numero_doc", ""),
                        "data": lanc.get("data", ""),
                        "observacao": lanc.get("historico", ""),
                        "codigo_parceiro": d.get("conta", ""),
                        "nome_parceiro": d.get("descricao", ""),
                        "valor": d.get("restante", ""),
                        "motivo_da_pendencia": "lancamento_nao_balanceado_debito",
                    }
                )

        for idx in range(j, len(creditos)):
            c = creditos[idx]
            if c["restante"] > TOLERANCIA_PAREAMENTO_VALOR:
                pendencias.append(
                    {
                        "seq": lanc.get("seq", ""),
                        "numero_transacao": lanc.get("numero_transacao", ""),
                        "numero_doc": lanc.get("numero_doc", ""),
                        "data": lanc.get("data", ""),
                        "observacao": lanc.get("historico", ""),
                        "codigo_parceiro": c.get("conta", ""),
                        "nome_parceiro": c.get("descricao", ""),
                        "valor": c.get("restante", ""),
                        "motivo_da_pendencia": "lancamento_nao_balanceado_credito",
                    }
                )

    return saida


def gerar_txt_iob(
    partidas: List[Dict[str, Any]], caminho_saida: Path
) -> Dict[str, int]:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    data_header = partidas[0]["data_ddmmyyyy"] if partidas else datetime.now().strftime("%d%m%Y")
    total_debitos = sum((p["valor"] for p in partidas), Decimal("0"))

    linhas: List[str] = [build_c_line(data_header, total_debitos)]
    seq = 3
    total_l = 0
    total_h = 0

    for partida in partidas:
        linhas.append(build_l_line(partida, seq))
        total_l += 1
        seq += 1

        linhas.append(build_h_line(partida.get("historico_h", ""), seq))
        total_h += 1
        seq += 1

    with caminho_saida.open("w", encoding="latin1", newline="\n") as fh:
        conteudo = "\n".join(linhas) + "\n"
        fh.write(conteudo.upper())

    return {
        "linhas_l": total_l,
        "linhas_h": total_h,
        "linhas_total": len(linhas),
    }


def gerar_relatorio_pendencias(pendencias: List[Dict[str, Any]], caminho_saida: Path) -> None:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    campos = [
        "seq",
        "numero_transacao",
        "numero_doc",
        "data",
        "observacao",
        "codigo_parceiro",
        "nome_parceiro",
        "valor",
        "motivo_da_pendencia",
    ]
    with caminho_saida.open("w", encoding="utf-8", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=campos, delimiter=";")
        writer.writeheader()
        for pend in pendencias:
            registro = dict(pend)
            valor = registro.get("valor")
            if isinstance(valor, Decimal):
                registro["valor"] = f"{valor:.2f}"
            writer.writerow(registro)


def gerar_relatorio_exclusoes(exclusoes: List[Dict[str, Any]], caminho_saida: Path) -> None:
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    campos = [
        "seq",
        "numero_transacao",
        "numero_doc",
        "data",
        "observacao",
        "motivo_exclusao",
        "transacao_par_vinculada",
    ]
    with caminho_saida.open("w", encoding="utf-8", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=campos, delimiter=";")
        writer.writeheader()
        for exc in exclusoes:
            writer.writerow(exc)


ITAU_MESES_ANO_REGEX = re.compile(
    r"\b(?:jan|fev|mar|abr|mai|jun|jul|ago|set|out|nov|dez)\s+(\d{4})\b",
    re.IGNORECASE,
)
ITAU_RESUMO_ANO_REGEX = re.compile(r"resumo\s*-\s*m[êe]s\s+\d{2}/(\d{4})", re.IGNORECASE)
ITAU_MOVIMENTACAO_HEADER_REGEX = re.compile(
    r"movimenta[cç][aã]o\s*-\s*aplica[cç][oõ]es/resgates antecipados e vencimentos",
    re.IGNORECASE,
)
ITAU_MOVIMENTACAO_ROW_REGEX = re.compile(
    r"^(?P<data>\d{2}/\d{2})\s+"
    r"(?P<aplicacao>[\d.]+,\d{2})\s+"
    r"(?P<resgate_principal>[\d.]+,\d{2})\s+"
    r"(?P<rend_bruto>[\d.]+,\d{2})\s+"
    r"(?P<iof>[\d.]+,\d{2})\s+"
    r"(?P<irrf>[\d.]+,\d{2})\s+"
    r"(?P<rend_liquido>[\d.]+,\d{2})"
    r"(?:\s+(?P<saldo_principal>[\d.]+,\d{2}))?$"
)


def extrair_ano_itau(linha: str, ano_atual: str) -> str:
    m_resumo = ITAU_RESUMO_ANO_REGEX.search(linha)
    if m_resumo:
        return m_resumo.group(1)
    m_mes = ITAU_MESES_ANO_REGEX.search(linha)
    if m_mes:
        return m_mes.group(1)
    return ano_atual


def montar_partida_itau(
    data_lancamento: date,
    debito_conta: str,
    credito_conta: str,
    valor: Decimal,
    complemento: str,
    historico_h: str,
) -> Dict[str, Any]:
    return {
        "data_ddmmyyyy": data_para_ddmmyyyy(data_lancamento),
        "debito_conta": debito_conta,
        "credito_conta": credito_conta,
        "valor": valor,
        "complemento": texto_ascii(complemento, 25),
        "historico_h": texto_ascii(historico_h, 50),
    }


def processar_pdf_itau_aplicacoes_interno(
    pdf_path: Path, conta_aplicacao: Optional[str], conta_corrente: Optional[str]
) -> List[Dict[str, Any]]:
    if not conta_aplicacao or not conta_corrente:
        return []

    conta_aplicacao = str(conta_aplicacao).strip()
    conta_corrente = str(conta_corrente).strip()
    partidas: List[Dict[str, Any]] = []
    dentro_quadro = False

    ano_atual = str(datetime.now().year)

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto = page.extract_text() or ""
            linhas = texto.split("\n")

            for linha in linhas:
                linha = linha.strip()
                if not linha:
                    continue

                ano_atual = extrair_ano_itau(linha, ano_atual)

                if ITAU_MOVIMENTACAO_HEADER_REGEX.search(linha):
                    dentro_quadro = True
                    continue

                if not dentro_quadro:
                    continue

                m = ITAU_MOVIMENTACAO_ROW_REGEX.match(linha)
                if not m:
                    continue

                data_bruta = m.group("data")
                data_str = f"{ano_atual}-{data_bruta[3:5]}-{data_bruta[:2]}"
                dt = parse_data(data_str)
                if not dt:
                    continue

                aplicacao = parse_decimal(m.group("aplicacao")) or Decimal("0")
                resgate_principal = parse_decimal(m.group("resgate_principal")) or Decimal("0")
                rend_bruto = parse_decimal(m.group("rend_bruto")) or Decimal("0")
                iof = parse_decimal(m.group("iof")) or Decimal("0")
                irrf = parse_decimal(m.group("irrf")) or Decimal("0")

                historico_base = f"ITAU APLICACOES {dt.strftime('%d/%m/%Y')}"

                if aplicacao > 0:
                    partidas.append(
                        montar_partida_itau(
                            dt,
                            conta_aplicacao,
                            conta_corrente,
                            aplicacao,
                            "APLICACAO ITAU",
                            historico_base,
                        )
                    )

                if resgate_principal > 0:
                    partidas.append(
                        montar_partida_itau(
                            dt,
                            conta_corrente,
                            conta_aplicacao,
                            resgate_principal,
                            "RESGATE ITAU",
                            historico_base,
                        )
                    )

                if rend_bruto > 0:
                    partidas.append(
                        montar_partida_itau(
                            dt,
                            conta_aplicacao,
                            "514010201",
                            rend_bruto,
                            "REND APLIC",
                            historico_base,
                        )
                    )

                if iof > 0:
                    partidas.append(
                        montar_partida_itau(
                            dt,
                            "514010103",
                            conta_aplicacao,
                            iof,
                            "IOF RETIDO",
                            historico_base,
                        )
                    )

                if irrf > 0:
                    partidas.append(
                        montar_partida_itau(
                            dt,
                            "112030107",
                            conta_aplicacao,
                            irrf,
                            "IRRF RETIDO",
                            historico_base,
                        )
                    )

    return partidas


def processar_gfbr_gerador_txt(
    input_path: Optional[str] = None,
    aba_origem: Optional[str] = None,
    pdf_itau_1_path: Optional[str] = None,
    conta_aplicacao_1: Optional[str] = None,
    conta_corrente_1: Optional[str] = None,
    pdf_itau_2_path: Optional[str] = None,
    conta_aplicacao_2: Optional[str] = None,
    conta_corrente_2: Optional[str] = None,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    # Determine base directory
    base_path = input_path or pdf_itau_1_path or pdf_itau_2_path
    if not base_path:
        raise ValueError("Nenhum arquivo de entrada fornecido.")
    
    caminho_base = Path(base_path)
    pasta_saida = Path(output_dir) if output_dir else caminho_base.parent
    pasta_saida.mkdir(parents=True, exist_ok=True)

    todas_partidas: List[Dict[str, Any]] = []
    pendencias: List[Dict[str, Any]] = []
    exclusoes: List[Dict[str, Any]] = []
    
    lancamentos_lidos = 0
    lancamentos_excluidos = 0
    lancamentos_exportados = 0

    # 1. Process Excel main file
    if input_path and Path(input_path).exists():
        caminho_entrada = Path(input_path)
        wb = load_workbook(caminho_entrada, data_only=True, read_only=False)
        nome_aba = aba_origem.strip() if isinstance(aba_origem, str) else ""
        if nome_aba:
            if nome_aba not in wb.sheetnames:
                raise ValueError(f"Aba '{nome_aba}' nao encontrada. Abas disponiveis: {wb.sheetnames}")
        else:
            nome_aba = wb.sheetnames[0]
        ws = wb[nome_aba]

        linha_cabecalho, colunas = localizar_linha_cabecalho(ws)
        lancamentos_brutos = agrupar_lancamentos(ws, linha_cabecalho, colunas)
        
        excluidos_ids, excl = marcar_pares_para_exclusao(lancamentos_brutos)
        exclusoes.extend(excl)
        
        validos_brutos = [l for l in lancamentos_brutos if l["_id"] not in excluidos_ids]
        
        lancamentos_norm: List[Dict[str, Any]] = []
        for bruto in validos_brutos:
            normalizado = normalizar_lancamento(bruto, pendencias)
            if normalizado:
                lancamentos_norm.append(normalizado)

        lancamentos_norm_filtrados: List[Dict[str, Any]] = []
        for lancamento_norm in lancamentos_norm:
            if eh_lancamento_renda_11102(lancamento_norm):
                lancamentos_excluidos += 1
                exclusoes.append(
                    construir_registro_exclusao(
                        lancamento_norm,
                        "exclusao_renda_classificacao_11102",
                    )
                )
                continue
            lancamentos_norm_filtrados.append(lancamento_norm)

        partidas_excel = criar_partidas_iob(lancamentos_norm_filtrados, pendencias)
        todas_partidas.extend(partidas_excel)
        
        lancamentos_lidos += len(lancamentos_brutos)
        lancamentos_excluidos += len(excluidos_ids)
        lancamentos_exportados += len(lancamentos_norm_filtrados)

    # 2. Process Itaú PDF 1
    if pdf_itau_1_path and Path(pdf_itau_1_path).exists() and conta_aplicacao_1:
        if not conta_corrente_1:
            raise ValueError("Informe a conta contábil da conta corrente do PDF Itaú 1.")
        partidas_pdf1 = processar_pdf_itau_aplicacoes_interno(
            Path(pdf_itau_1_path),
            conta_aplicacao_1,
            conta_corrente_1,
        )
        todas_partidas.extend(partidas_pdf1)
        lancamentos_lidos += len(partidas_pdf1)
        lancamentos_exportados += len(partidas_pdf1)

    # 3. Process Itaú PDF 2
    if pdf_itau_2_path and Path(pdf_itau_2_path).exists() and conta_aplicacao_2:
        if not conta_corrente_2:
            raise ValueError("Informe a conta contábil da conta corrente do PDF Itaú 2.")
        partidas_pdf2 = processar_pdf_itau_aplicacoes_interno(
            Path(pdf_itau_2_path),
            conta_aplicacao_2,
            conta_corrente_2,
        )
        todas_partidas.extend(partidas_pdf2)
        lancamentos_lidos += len(partidas_pdf2)
        lancamentos_exportados += len(partidas_pdf2)

    # 4. Validations (Warnings for same debit/credit accounts)
    advertencias: List[str] = []
    for idx, p in enumerate(todas_partidas):
        debito = str(p.get("debito_conta", "")).strip()
        credito = str(p.get("credito_conta", "")).strip()
        if debito and credito and debito == credito:
            data_str = p.get("data_ddmmyyyy", "")
            hist = p.get("historico_h", "")
            vl = p.get("valor", Decimal("0"))
            advertencias.append(
                f"Lançamento em {data_str[:2]}/{data_str[2:4]}/{data_str[4:]}: Débito e Crédito na conta '{debito}'. R$ {vl:.2f} - {hist[:30]}"
            )

    # 5. Sort by date (optional, but ensures chronological consistency if we mixed files)
    try:
        todas_partidas.sort(key=lambda x: datetime.strptime(x["data_ddmmyyyy"], "%d%m%Y"))
    except Exception:
        pass  # ignore if sort fails, we just don't sort chronologically

    arquivo_txt = pasta_saida / TXT_SAIDA_NOME
    arquivo_pendencias = pasta_saida / PENDENCIAS_NOME
    
    if todas_partidas:
        estat = gerar_txt_iob(todas_partidas, arquivo_txt)
    else:
        # Cria um arquivo vazio apenas com C caso nao encontre nada
        with arquivo_txt.open("w", encoding="latin1") as f:
            f.write(build_c_line(datetime.now().strftime("%d%m%Y"), Decimal("0")) + "\n")
        estat = {"linhas_l": 0, "linhas_h": 0}

    gerar_relatorio_pendencias(pendencias, arquivo_pendencias)
    
    return {
        "lancamentos_lidos": lancamentos_lidos,
        "lancamentos_excluidos": lancamentos_excluidos,
        "lancamentos_exportados": lancamentos_exportados,
        "partidas_geradas": len(todas_partidas),
        "linhas_l_exportadas": estat["linhas_l"],
        "linhas_h_exportadas": estat["linhas_h"],
        "pendencias": len(pendencias),
        "arquivo_entrada": str(base_path),
        "aba_utilizada": aba_origem or "Múltipla",
        "arquivo_txt": str(arquivo_txt),
        "arquivo_pendencias": str(arquivo_pendencias),
        "advertencias": advertencias,
    }
