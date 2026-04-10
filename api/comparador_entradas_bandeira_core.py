from __future__ import annotations

import io
import math
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment

TOLERANCIA_DIFERENCA = 0.01
STATUS_DIVERGENTE = "DIVERGENTE ENTRE ARQUIVOS"
STATUS_SO_FSIST = "SÓ NO FSIST"
COLUNA_NUMERO_SAIDA = "NÚMERO"
COLUNA_DIFERENCA_SAIDA = "DIFERENÇA"


def _para_float(valor: object) -> float:
    if valor is None or (isinstance(valor, float) and math.isnan(valor)):
        return 0.0
    try:
        return float(valor)
    except (TypeError, ValueError):
        texto = str(valor).strip().replace(".", "").replace(",", ".")
        try:
            return float(texto)
        except ValueError:
            return 0.0


def _normalizar_numero(valor: object) -> str:
    if valor is None or (isinstance(valor, float) and math.isnan(valor)):
        return ""
    texto = str(valor).strip()
    if not texto:
        return ""
    if texto.endswith(".0"):
        texto = texto[:-2]
    return "".join(ch for ch in texto if ch.isdigit())


def _classificar_status(valor_fsist: float, valor_entradas: float) -> str:
    existe_fsist = pd.notna(valor_fsist)
    existe_entradas = pd.notna(valor_entradas)

    if existe_fsist and existe_entradas:
        if abs(float(valor_fsist) - float(valor_entradas)) <= TOLERANCIA_DIFERENCA:
            return "OK"
        return STATUS_DIVERGENTE
    if existe_fsist:
        return STATUS_SO_FSIST
    return "SOMENTE_ENTRADAS"


def _carregar_fsist(arquivo_bytes: bytes, nome_arquivo: str) -> pd.DataFrame:
    try:
        df = pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name=0)
    except Exception as exc:
        raise ValueError(
            f"Arquivo do FSIST corrompido, inválido ou em formato não suportado: {nome_arquivo}."
        ) from exc

    coluna_numero = "Número" if "Número" in df.columns else "NÚMERO"
    colunas_necessarias = {coluna_numero, "Valor"}
    faltando = colunas_necessarias - set(df.columns)
    if faltando:
        raise ValueError(
            "Formato errado no arquivo do FSIST. "
            f"Colunas obrigatórias: {coluna_numero} e Valor."
        )

    base = pd.DataFrame()
    base["numero"] = df[coluna_numero].apply(_normalizar_numero)
    base["valor_fsist"] = df["Valor"].apply(_para_float)
    base = base[base["numero"] != ""]
    return base.groupby("numero", as_index=False)["valor_fsist"].sum()


def _carregar_entradas_iob(arquivo_bytes: bytes, nome_arquivo: str) -> pd.DataFrame:
    try:
        bruto = pd.read_excel(io.BytesIO(arquivo_bytes), sheet_name=0, header=None)
    except Exception as exc:
        raise ValueError(
            f"Arquivo de Entradas IOB corrompido, inválido ou em formato não suportado: {nome_arquivo}."
        ) from exc

    notas: list[dict[str, Any]] = []
    numero_atual: str | None = None
    valor_atual = 0.0

    for _, row in bruto.iterrows():
        numero = _normalizar_numero(row[3] if 3 in bruto.columns else None)
        tipo_doc = (
            str(row[2]).strip().upper()
            if 2 in bruto.columns and pd.notna(row[2])
            else ""
        )
        linha_documento = tipo_doc in {"NF", "NFE", "NFCE", "CTE", "CT-E"}
        valor_col7 = _para_float(row[7] if 7 in bruto.columns else None)
        valor_col8 = _para_float(row[8] if 8 in bruto.columns else None)

        if numero and linha_documento:
            if numero_atual is not None:
                notas.append({"numero": numero_atual, "valor_entradas": valor_atual})
            numero_atual = numero
            valor_atual = valor_col7 + valor_col8
            continue

        if numero_atual is not None:
            valor_atual += valor_col7 + valor_col8

    if numero_atual is not None:
        notas.append({"numero": numero_atual, "valor_entradas": valor_atual})

    if not notas:
        raise ValueError(
            "Formato errado no arquivo de Entradas IOB. "
            "Não foi possível identificar as notas."
        )

    df = pd.DataFrame(notas)
    df = df[df["numero"] != ""]
    return df.groupby("numero", as_index=False)["valor_entradas"].sum()


def _gerar_comparativo(df_fsist: pd.DataFrame, df_entradas: pd.DataFrame) -> pd.DataFrame:
    comparativo = df_fsist.merge(df_entradas, on="numero", how="outer")
    comparativo["status"] = comparativo.apply(
        lambda linha: _classificar_status(
            linha.get("valor_fsist"), linha.get("valor_entradas")
        ),
        axis=1,
    )
    comparativo["diferenca"] = (
        comparativo["valor_fsist"].fillna(0) - comparativo["valor_entradas"].fillna(0)
    )
    comparativo = comparativo[
        comparativo["status"].isin([STATUS_DIVERGENTE, STATUS_SO_FSIST])
    ].copy()

    comparativo = comparativo.rename(
        columns={
            "numero": COLUNA_NUMERO_SAIDA,
            "valor_fsist": "VALOR FSIST",
            "valor_entradas": "VALOR ENTRADAS",
            "status": "STATUS",
            "diferenca": COLUNA_DIFERENCA_SAIDA,
        }
    )
    comparativo = comparativo.sort_values(by=["STATUS", COLUNA_NUMERO_SAIDA]).reset_index(
        drop=True
    )
    return comparativo


def _aplicar_estilo_planilha(worksheet) -> None:
    formato_brl = r'"R$" #,##0.00'
    alinhamento_direita = Alignment(horizontal="right")

    for coluna in ("B", "C", "E"):
        for linha in range(2, worksheet.max_row + 1):
            celula = worksheet[f"{coluna}{linha}"]
            if isinstance(celula.value, (int, float)):
                celula.number_format = formato_brl

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for celula in row:
            celula.alignment = alinhamento_direita


def _gerar_excel_bytes(comparativo: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        comparativo.to_excel(writer, index=False, sheet_name="comparativo")
        worksheet = writer.book["comparativo"]
        _aplicar_estilo_planilha(worksheet)
    return buffer.getvalue()


def _preview_records(df: pd.DataFrame, limite: int = 200) -> list[dict[str, Any]]:
    preview = df.head(limite).copy()
    registros: list[dict[str, Any]] = []

    for _, linha in preview.iterrows():
        item: dict[str, Any] = {}
        for coluna, valor in linha.items():
            if pd.isna(valor):
                item[coluna] = None
            else:
                valor_json = valor
                if hasattr(valor_json, "item"):
                    try:
                        valor_json = valor_json.item()
                    except Exception:
                        pass
                if isinstance(valor_json, pd.Timestamp):
                    valor_json = valor_json.isoformat()
                item[coluna] = valor_json
        registros.append(item)

    return registros


def processar_comparador_entradas_bandeira(
    arquivo_fsist_bytes: bytes,
    arquivo_fsist_nome: str,
    arquivo_entradas_bytes: bytes,
    arquivo_entradas_nome: str,
) -> dict[str, Any]:
    if not arquivo_fsist_bytes:
        raise ValueError("Arquivo do FSIST vazio.")
    if not arquivo_entradas_bytes:
        raise ValueError("Arquivo de Entradas IOB vazio.")

    df_fsist = _carregar_fsist(arquivo_fsist_bytes, arquivo_fsist_nome)
    df_entradas = _carregar_entradas_iob(arquivo_entradas_bytes, arquivo_entradas_nome)
    comparativo = _gerar_comparativo(df_fsist, df_entradas)
    excel_bytes = _gerar_excel_bytes(comparativo)

    total_divergentes = int((comparativo["STATUS"] == STATUS_DIVERGENTE).sum())
    total_so_fsist = int((comparativo["STATUS"] == STATUS_SO_FSIST).sum())
    nome_saida = (
        "comparativo_entradas_bandeira_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return {
        "arquivo_saida": nome_saida,
        "xlsx_bytes": excel_bytes,
        "preview": _preview_records(comparativo, limite=200),
        "total_linhas": int(len(comparativo)),
        "total_divergente_entre_arquivos": total_divergentes,
        "total_so_no_fsist": total_so_fsist,
        "status_validos": [STATUS_DIVERGENTE, STATUS_SO_FSIST],
    }
